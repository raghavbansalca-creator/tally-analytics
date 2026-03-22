"""
Seven Labs Vision -- Narration Audit Engine
Reads voucher narrations from Tally data, classifies transactions,
flags suspicious entries, and auto-generates audit comments.
"""

import sqlite3
import re
import os
from datetime import datetime

# ── CATEGORY DEFINITIONS ─────────────────────────────────────────────────────

CATEGORIES = [
    {
        "name": "Related Party",
        "patterns": [
            r"\bdirector\b", r"\bpromoter\b", r"\brelative\b",
            # "family" — exclude shop/business names containing "family"
            r"\bfamily\s+(?:member|relation|trust|concern)\b",
            r"\bspouse\b",
            # Family members — require context (not standalone product names)
            r"\b(?:his|her|my|the|to|of|for)\s+(?:son|daughter|brother|sister|father|mother)\b",
            r"\bdirector'?s?\s+(?:wife|husband|son|daughter|brother|sister|relative)\b",
            r"\bpartner'?s?\s+(?:wife|husband|capital|drawing|current|loan)\b",
        ],
        "comment": "Possible related party transaction - verify arm's length pricing",
        "severity": "HIGH",
    },
    {
        "name": "Cash Transactions",
        "patterns": [
            r"\bcash\s+(?:payment|receipt|deposit|withdrawal|paid|received|purchase)\b",
            r"\b(?:paid|received|deposited|withdrawn)\s+(?:in\s+)?cash\b",
            r"\bcash\s+(?:a/?c|account)\b",
            r"(?:^|\s)cash(?:\s|$)",  # standalone "cash" but not inside words
        ],
        "comment": "Cash transaction - verify Sec 269ST/269SS compliance",
        "severity": "MEDIUM",
    },
    {
        "name": "Loan/Advance",
        "patterns": [
            r"\bloan\b", r"\blent\b", r"\bborrowed\b",
            r"\bemi\b", r"(?:^|[\s_])emi(?:[\s_]|$)",  # EMI with underscores too (EMI_CEL...)
            r"\brepayment\b",
            # "advance" — broader matching, exclude "advance tax"
            r"\badvances?\s+(?:to|from|given|received|paid|against|payment|amount)\b",
            r"\b(?:staff|employee|salary|personal)\s+advances?\b",
            r"\b(?:transfer|trf|paid\s+for)\s+advances?\b",
            r"\badvance\s+(?:payment|amount|for|to|from)\b",
            r"(?:^|\s)advance(?:\s|$)",  # standalone "advance" at word boundary
            # "interest" — broader matching for SAVC patterns (common typos: intrest, interst)
            r"\bint(?:e?re?|re)st\s+(?:on|paid|received|charged|due|capitali[sz]ed|for)\b",
            r"\bbeing\s+int(?:e?re?|re)st\b",
            r"\bdebit\s+int(?:e?re?|re)st\b",
            r"InterestCharged",  # no space variant
            r"\bcasa\s+(?:debit\s+)?int(?:e?re?|re)st\b",
            r"\b(?:paid\s+for|for)\s+int(?:e?re?|re)st\b",
            r"\bsec\s+\d+[a-z]?\s+int(?:e?re?|re)st\b",  # SEC 94C INTEREST etc.
        ],
        "comment": "Loan/advance - verify Sec 185/186 compliance, TDS applicability",
        "severity": "MEDIUM",
    },
    {
        "name": "Capital Expenditure",
        "patterns": [
            r"\bpurchase of\b", r"\bacquisition\b",
            # "machinery" — exclude shop names like "MACHINERY STORE", "MACHINE TOOLS"
            r"\b(?:purchase|bought|acquired|new)\s+(?:of\s+)?machinery\b",
            r"\bmachinery\s+(?:purchase|bought|acquired|cost)\b",
            # "equipment" — exclude company names like "BALA EQUIPMENTS"
            r"\b(?:purchase|bought|acquired|new)\s+(?:of\s+)?equipments?\b",
            r"\bequipments?\s+(?:purchase|bought|acquired|cost)\b",
            # "vehicle" — exclude "VEHICLE NO." references and repair/consumable contexts
            r"\b(?:purchase|bought|acquired|new)\s+(?:of\s+)?vehicles?\b",
            r"\bvehicles?\s+(?:purchase|bought|acquired|cost)\b",
            r"\bcomputer\b", r"\blaptop\b",
            r"\bfurniture\b", r"\bair\s*conditioner\b",
            # "building" — exclude company names like "DEEP BUILDING MATERIAL"
            r"\b(?:purchase|construction|acquired)\s+(?:of\s+)?building\b",
            r"\bbuilding\s+(?:purchase|construction|cost|work)\b",
            r"\bfixed\s+asset\b", r"\bcapital\s+(?:asset|goods|item)\b",
        ],
        # BUG 5 FIX: Removed \bac\b (too short, matches "account")
        "comment": "Capital expenditure - verify capitalization vs revenue treatment",
        "severity": "MEDIUM",
    },
    {
        "name": "Salary/Wages",
        "patterns": [
            r"\bsalary\b", r"\bwages\b", r"\bbonus\b", r"\bincentive\b",
            r"\bgratuity\b", r"\bleave encashment\b",
            r"\bpf\b", r"\besi\b", r"\bprofessional tax\b",
            r"\bstipend\b", r"\bhonorarium\b",
            # "commission" removed — too generic (sales commission vs employee commission)
        ],
        "comment": "Employee payment - verify TDS u/s 192, PF/ESI compliance",
        "severity": "LOW",
    },
    {
        "name": "Rent Payments",
        "patterns": [
            r"\b(?:office|warehouse|godown|shop|factory|premises|property|building|guest\s*house)\s+rent(?:ed|ing|al|s)?\b",
            r"\brent(?:ed|ing|al)?\s+(?:paid|payment|for|of|a/?c|account)\b",
            r"\blease\s+(?:rent|payment|amount)\b", r"\blicense fee\b",
            r"\bmonthly\s+rent\b",
            r"\brent(?:ed|ing)?\s+(?:office|shop|warehouse|godown|premises|space|room|flat|house)\b",
            # Construction/site equipment rent (Poclain, JCB, machine, crane, etc.)
            # Allow words between equipment name and "rent" (e.g. "JCB June Month rent")
            r"\b(?:poclain|jcb|crane|excavator|hydra|loader|mixer|generator|d\.?g\.?|dg|machine|binding\s+machine|exhaust\s+fan)\b.{0,40}\brent\b",
            r"\brent\s+(?:for|of)\b",  # generic "rent for/of"
            r"\bhire\s+(?:charges?|rent)\b", r"\b(?:on|for)\s+hire\b",
            r"\broom\s+rent\b",  # room rent
            # "for Rent" or "Rent ... Days" — standalone rent with context
            r"\bfor\s+rent\b", r"\brent\s+\d+\s+days?\b",
            r"\bpathak\s+dg\b.{0,20}\brent\b",  # Pathak DG rent
        ],
        "comment": "Rent payment - verify TDS u/s 194I, GST RCM if applicable",
        "severity": "LOW",
    },
    {
        "name": "Professional/Consultancy",
        "patterns": [
            r"\bprofessional\s+(?:fees?|charges?|services?)\b",  # fees plural
            r"\bconsultancy\b", r"\bconsulting\b", r"\bconsultants?\b",  # consultant/consultants
            r"\blegal\s+(?:fees?|charges?|services?|expenses?)\b",
            r"\baudit\s*fees?\b", r"\bca\s+fees?\b", r"\bcs\s+fees?\b",
            r"\badvocate\b", r"\blawyer\b", r"\badvisory\b",
            r"\binternal\s+audit\b",
        ],
        "comment": "Professional fee - verify TDS u/s 194J",
        "severity": "LOW",
    },
    {
        "name": "Contractor Payments",
        "patterns": [
            r"\bcontractor\b", r"\bsub-contractor\b",
            # "labour" — broader matching: loading labour, labour at, etc.
            r"\blabour\b",
            r"\bworks contract\b", r"\bjob work\b", r"\bfabrication\b",
            r"\bconstruction\s+(?:work|site|charge|cost|material)\b",
        ],
        "comment": "Contractor payment - verify TDS u/s 194C",
        "severity": "LOW",
    },
    {
        "name": "Insurance",
        "patterns": [
            r"insurance", r"\bpremium\b",  # BUG 8 FIX: no word boundary at end for "insurance"
            r"\binsurance\s+(?:policy|premium|renewal|claim)\b",
            r"\bmediclaim\b", r"\blic\b", r"\bgeneral insurance\b",
            r"\bhealth\s+insurance\b", r"\bfire\s+insurance\b",
        ],
        "comment": "Insurance - verify prepaid treatment if multi-year",
        "severity": "LOW",
    },
    {
        "name": "Provision/Write-off",
        "patterns": [
            r"\bprovision\s+(?:for|against|made)\b",
            r"\bwrite\s*off\b", r"\bwritten\s*off\b",
            r"\bbad\s+debt\b", r"\bdoubtful\b", r"\bnpa\b", r"\bwaiver\b",
            # Exclude "provisional" (too generic)
        ],
        "comment": "Provision/write-off - verify board resolution and documentation",
        "severity": "HIGH",
    },
    {
        "name": "Inter-company/Branch",
        "patterns": [
            r"\binter[\s-]?company\b", r"\bbranch\s+transfer\b",
            # "head office" — only flag when combined with transfer/trf, not salary/payment location
            r"\b(?:transfer|trf)\s+(?:to|from|for)\s+head\s+office\b",
            r"\bhead\s+office\s+(?:transfer|trf)\b",
            r"\binter[\s-]?unit\b",
            r"\b(?:transfer\s+(?:to|from)\s+)?branch\s+(?:office|unit|location)\b",
            # BUG 4 FIX: removed standalone \bbranch\b and \bho\b (matches bank ATM branches)
        ],
        "comment": "Inter-company/branch - verify transfer pricing, GST implications",
        "severity": "MEDIUM",
    },
    {
        "name": "Reversal/Correction",
        "patterns": [
            r"\breversal\b", r"\breversed\b",
            # "correction" — require accounting context, exclude "registration & correction"
            r"\bcorrection\s+(?:entry|entries|voucher|journal|of\s+entry)\b",
            r"\b(?:entry|entries|voucher|journal)\s+correction\b",
            r"\brectification\b", r"\bmistake\b",
            r"\bwrong\s+entry\b", r"\bentry\s+(?:reversed|corrected)\b",
            # Removed \berror\b (too generic) and \badjusted\b (too generic)
        ],
        "comment": "Reversal/correction entry - verify original entry and authorization",
        "severity": "HIGH",
    },
    {
        "name": "Year-end Adjustments",
        "patterns": [
            r"\bclosing\s+(?:entry|entries|balance|stock|adjustment)\b",
            r"\bopening\s+(?:entry|entries|balance|stock|adjustment)\b",
            r"\byear[\s-]?end\b",
            r"\b(?:accrual|accrued)\b",
            r"\bprepaid\s+(?:expense|rent|insurance|amount)\b",
            # BUG 3 FIX: removed standalone \bprepaid\b (matches "Amazon Prepaid")
            r"\boutstanding\s+(?:expense|salary|rent|liability)\b",
        ],
        "comment": "Year-end adjustment - verify cut-off and supporting documentation",
        "severity": "MEDIUM",
    },
    {
        "name": "Suspense/Clearing",
        "patterns": [
            r"\bsuspense\b", r"\bclearing\s+(?:a/?c|account|entry)\b",
            r"\btemporary\s+(?:a/?c|account|entry|posting)\b",
            r"\bunidentified\b", r"\bunknown\b",
            # Removed \bpending\b and \bto be\b (too generic)
            r"\btba\b",
        ],
        "comment": "Suspense/clearing entry - must be cleared before year-end",
        "severity": "HIGH",
    },
    {
        "name": "Donation/CSR",
        "patterns": [
            r"\bdonation\b", r"\bcharity\b", r"\bcsr\b",
            r"\bcontribution\b", r"\bcorpus\b",
            # Removed \btrust\b (too generic — many businesses are trusts)
        ],
        "comment": "Donation/CSR - verify Sec 80G eligibility, Sec 135 compliance",
        "severity": "MEDIUM",
    },
    {
        "name": "Foreign/Forex",
        "patterns": [
            r"\bforeign\s+(?:currency|exchange|remittance|payment|receipt)\b",
            r"\bforex\b", r"\busd\b", r"\beur\b", r"\bgbp\b",
            r"\bremittance\b", r"\bswift\b", r"\bwire\s+transfer\b",
            r"\bcurrency\s+exchange\b",
            # BUG 2 FIX: removed standalone \bexchange\b (matches product exchange/return)
        ],
        "comment": "Foreign transaction - verify FEMA compliance, withholding tax",
        "severity": "MEDIUM",
    },
]

# Pre-compile all regex patterns for performance
for cat in CATEGORIES:
    cat["_compiled"] = [re.compile(p, re.IGNORECASE) for p in cat["patterns"]]

# Severity ordering for sorting
SEVERITY_ORDER = {"HIGH": 0, "MEDIUM": 1, "LOW": 2}


# ── CLASSIFICATION ───────────────────────────────────────────────────────────

def classify_narration(narration: str) -> list[dict]:
    """Classify a single narration. Returns list of matched categories with comments.

    Each match dict has: category, comment, severity
    A single narration can match multiple categories.
    """
    results = []

    # Handle no narration
    if not narration or not narration.strip():
        results.append({
            "category": "No Narration",
            "comment": "WARNING: No narration - transaction purpose unclear",
            "severity": "HIGH",
        })
        return results

    text = narration.strip()

    # Check unusual length
    if len(text) < 3:
        results.append({
            "category": "Unusually Short Narration",
            "comment": "Verify: Unusually short narration",
            "severity": "MEDIUM",
        })
    elif len(text) > 200:
        results.append({
            "category": "Unusually Long Narration",
            "comment": "Verify: Unusually long narration",
            "severity": "LOW",
        })

    # Match against all categories
    for cat in CATEGORIES:
        for regex in cat["_compiled"]:
            if regex.search(text):
                results.append({
                    "category": cat["name"],
                    "comment": cat["comment"],
                    "severity": cat["severity"],
                })
                break  # one match per category is enough

    return results


# ── FULL ANALYSIS ────────────────────────────────────────────────────────────

def analyze_all_narrations(db_path: str, from_date=None, to_date=None) -> dict:
    """Analyze all voucher narrations in the database.

    Args:
        db_path: Path to the SQLite database.
        from_date: Optional start date filter (YYYYMMDD string).
        to_date: Optional end date filter (YYYYMMDD string).

    Returns dict with total_vouchers, narrations_analyzed, category_summary,
    flagged_vouchers, risk_summary.
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    # Build query
    date_filter = ""
    params = []
    if from_date:
        date_filter += " AND v.DATE >= ?"
        params.append(from_date)
    if to_date:
        date_filter += " AND v.DATE <= ?"
        params.append(to_date)

    sql = f"""
    SELECT
        v.GUID,
        v.DATE,
        v.VOUCHERTYPENAME,
        v.VOUCHERNUMBER,
        v.PARTYLEDGERNAME,
        v.NARRATION,
        COALESCE(
            (SELECT SUM(ABS(CAST(a.AMOUNT AS REAL)))
             FROM trn_accounting a
             WHERE a.VOUCHER_GUID = v.GUID AND CAST(a.AMOUNT AS REAL) > 0),
            0
        ) as amount
    FROM trn_voucher v
    WHERE 1=1{date_filter}
    ORDER BY v.DATE
    """

    rows = conn.execute(sql, params).fetchall()
    conn.close()

    total_vouchers = len(rows)
    narrations_analyzed = 0
    no_narration_count = 0

    # Category summary: {name: {count, total_amount, vouchers[]}}
    category_summary = {}
    flagged_vouchers = []

    for row in rows:
        guid = row["GUID"]
        date_raw = row["DATE"] or ""
        vtype = row["VOUCHERTYPENAME"] or ""
        vnum = row["VOUCHERNUMBER"] or ""
        party = row["PARTYLEDGERNAME"] or ""
        narration = row["NARRATION"] or ""
        amount = float(row["amount"] or 0)

        # Format date for display
        display_date = _format_date(date_raw)

        narrations_analyzed += 1
        matches = classify_narration(narration)

        # BUG 9 FIX: Also flag if party ledger is "Cash" (even if narration doesn't say "cash")
        if party and party.strip().upper() in ("CASH", "CASH A/C", "CASH ACCOUNT", "CASH IN HAND", "CASH-IN-HAND", "PETTY CASH"):
            cash_already = any(m["category"] == "Cash Transactions" for m in matches)
            if not cash_already:
                matches.append({
                    "category": "Cash Transactions",
                    "comment": "Cash transaction (party is Cash ledger) - verify Sec 269ST/269SS compliance",
                    "severity": "MEDIUM",
                })

        if not matches:
            continue

        # Check if "No Narration" is among matches
        cat_names = [m["category"] for m in matches]
        if "No Narration" in cat_names:
            no_narration_count += 1

        # Determine highest severity among matches
        highest_severity = _highest_severity(matches)

        # For "No Narration" on large amounts, bump to HIGH
        if "No Narration" in cat_names and amount >= 100000:
            highest_severity = "HIGH"

        comments = list(dict.fromkeys(m["comment"] for m in matches))
        categories = list(dict.fromkeys(m["category"] for m in matches))

        # Update category summary
        for m in matches:
            cname = m["category"]
            if cname not in category_summary:
                category_summary[cname] = {"count": 0, "total_amount": 0.0, "vouchers": []}
            category_summary[cname]["count"] += 1
            category_summary[cname]["total_amount"] += amount
            # Limit stored vouchers per category to avoid memory bloat
            if len(category_summary[cname]["vouchers"]) < 500:
                category_summary[cname]["vouchers"].append({
                    "guid": guid,
                    "date": display_date,
                    "voucher_type": vtype,
                    "voucher_number": vnum,
                    "party": party,
                    "amount": amount,
                    "narration": narration[:300],
                })

        flagged_vouchers.append({
            "guid": guid,
            "date": display_date,
            "voucher_type": vtype,
            "voucher_number": vnum,
            "party": party,
            "amount": amount,
            "narration": narration[:300],
            "categories": categories,
            "comments": comments,
            "severity": highest_severity,
        })

    # Sort flagged vouchers: severity (HIGH first), then amount descending
    flagged_vouchers.sort(
        key=lambda v: (SEVERITY_ORDER.get(v["severity"], 9), -v["amount"])
    )

    # Risk summary
    risk_summary = {
        "high": sum(1 for v in flagged_vouchers if v["severity"] == "HIGH"),
        "medium": sum(1 for v in flagged_vouchers if v["severity"] == "MEDIUM"),
        "low": sum(1 for v in flagged_vouchers if v["severity"] == "LOW"),
    }

    return {
        "total_vouchers": total_vouchers,
        "narrations_analyzed": narrations_analyzed,
        "no_narration_count": no_narration_count,
        "category_summary": category_summary,
        "flagged_vouchers": flagged_vouchers,
        "risk_summary": risk_summary,
    }


# ── SINGLE VOUCHER COMMENTS ─────────────────────────────────────────────────

def get_voucher_comments(db_path: str, voucher_guid: str) -> dict:
    """Get auto-generated comments for a specific voucher.

    Returns dict with guid, narration, categories, comments, severity.
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    row = conn.execute(
        "SELECT GUID, NARRATION, VOUCHERTYPENAME, VOUCHERNUMBER, "
        "PARTYLEDGERNAME, DATE FROM trn_voucher WHERE GUID = ?",
        (voucher_guid,)
    ).fetchone()
    conn.close()

    if not row:
        return {"error": "Voucher not found"}

    narration = row["NARRATION"] or ""
    matches = classify_narration(narration)

    return {
        "guid": voucher_guid,
        "date": _format_date(row["DATE"] or ""),
        "voucher_type": row["VOUCHERTYPENAME"] or "",
        "voucher_number": row["VOUCHERNUMBER"] or "",
        "party": row["PARTYLEDGERNAME"] or "",
        "narration": narration,
        "categories": [m["category"] for m in matches],
        "comments": [m["comment"] for m in matches],
        "severity": _highest_severity(matches) if matches else "LOW",
    }


# ── EXCEL EXPORT ─────────────────────────────────────────────────────────────

def export_narration_report(analysis_result: dict, output_path: str):
    """Export analysis to Excel with sheets: Summary, Flagged Vouchers, Category Details.

    Args:
        analysis_result: Output from analyze_all_narrations().
        output_path: Path for the .xlsx file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        # Fallback: try pandas Excel writer
        _export_with_pandas(analysis_result, output_path)
        return

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    severity_fills = {
        "HIGH": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "MEDIUM": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
        "LOW": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
    }
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws_sum.merge_cells("A1:D1")
    ws_sum["A1"] = "Narration Audit Report"
    ws_sum["A1"].font = Font(bold=True, size=14)

    ws_sum["A3"] = "Total Vouchers"
    ws_sum["B3"] = analysis_result.get("total_vouchers", 0)
    ws_sum["A4"] = "Narrations Analyzed"
    ws_sum["B4"] = analysis_result.get("narrations_analyzed", 0)
    ws_sum["A5"] = "No Narration Count"
    ws_sum["B5"] = analysis_result.get("no_narration_count", 0)

    risk = analysis_result.get("risk_summary", {})
    ws_sum["A7"] = "Risk Summary"
    ws_sum["A7"].font = Font(bold=True, size=11)
    ws_sum["A8"] = "HIGH"
    ws_sum["B8"] = risk.get("high", 0)
    ws_sum["A8"].fill = severity_fills["HIGH"]
    ws_sum["A9"] = "MEDIUM"
    ws_sum["B9"] = risk.get("medium", 0)
    ws_sum["A9"].fill = severity_fills["MEDIUM"]
    ws_sum["A10"] = "LOW"
    ws_sum["B10"] = risk.get("low", 0)
    ws_sum["A10"].fill = severity_fills["LOW"]

    # Category summary table
    ws_sum["A12"] = "Category"
    ws_sum["B12"] = "Count"
    ws_sum["C12"] = "Total Amount"
    for cell in [ws_sum["A12"], ws_sum["B12"], ws_sum["C12"]]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row_idx = 13
    cat_summary = analysis_result.get("category_summary", {})
    for cname in sorted(cat_summary.keys(), key=lambda k: -cat_summary[k]["count"]):
        cdata = cat_summary[cname]
        ws_sum.cell(row=row_idx, column=1, value=cname).border = thin_border
        ws_sum.cell(row=row_idx, column=2, value=cdata["count"]).border = thin_border
        ws_sum.cell(row=row_idx, column=3, value=round(cdata["total_amount"], 2)).border = thin_border
        ws_sum.cell(row=row_idx, column=3).number_format = "#,##0.00"
        row_idx += 1

    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 15
    ws_sum.column_dimensions["C"].width = 20

    # ── Sheet 2: Flagged Vouchers ────────────────────────────────────────
    ws_flag = wb.create_sheet("Flagged Vouchers")
    flag_headers = ["Date", "Voucher Type", "Number", "Party", "Amount",
                    "Narration", "Categories", "Comments", "Severity"]
    for ci, h in enumerate(flag_headers, 1):
        cell = ws_flag.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for ri, v in enumerate(analysis_result.get("flagged_vouchers", []), 2):
        ws_flag.cell(row=ri, column=1, value=v.get("date", "")).border = thin_border
        ws_flag.cell(row=ri, column=2, value=v.get("voucher_type", "")).border = thin_border
        ws_flag.cell(row=ri, column=3, value=v.get("voucher_number", "")).border = thin_border
        ws_flag.cell(row=ri, column=4, value=v.get("party", "")).border = thin_border
        amt_cell = ws_flag.cell(row=ri, column=5, value=v.get("amount", 0))
        amt_cell.number_format = "#,##0.00"
        amt_cell.border = thin_border
        ws_flag.cell(row=ri, column=6, value=v.get("narration", "")).border = thin_border
        ws_flag.cell(row=ri, column=7, value=", ".join(v.get("categories", []))).border = thin_border
        ws_flag.cell(row=ri, column=8, value="; ".join(v.get("comments", []))).border = thin_border
        sev_cell = ws_flag.cell(row=ri, column=9, value=v.get("severity", ""))
        sev_cell.border = thin_border
        sev = v.get("severity", "")
        if sev in severity_fills:
            sev_cell.fill = severity_fills[sev]

    for ci, w in enumerate([12, 14, 12, 25, 15, 50, 30, 50, 10], 1):
        ws_flag.column_dimensions[chr(64 + ci)].width = w

    # ── Sheet 3: Category Details ────────────────────────────────────────
    ws_cat = wb.create_sheet("Category Details")
    cat_headers = ["Category", "Date", "Voucher Type", "Number", "Party", "Amount", "Narration"]
    for ci, h in enumerate(cat_headers, 1):
        cell = ws_cat.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    ri = 2
    for cname in sorted(cat_summary.keys()):
        for v in cat_summary[cname].get("vouchers", []):
            ws_cat.cell(row=ri, column=1, value=cname).border = thin_border
            ws_cat.cell(row=ri, column=2, value=v.get("date", "")).border = thin_border
            ws_cat.cell(row=ri, column=3, value=v.get("voucher_type", "")).border = thin_border
            ws_cat.cell(row=ri, column=4, value=v.get("voucher_number", "")).border = thin_border
            ws_cat.cell(row=ri, column=5, value=v.get("party", "")).border = thin_border
            amt_cell = ws_cat.cell(row=ri, column=6, value=v.get("amount", 0))
            amt_cell.number_format = "#,##0.00"
            amt_cell.border = thin_border
            ws_cat.cell(row=ri, column=7, value=v.get("narration", "")).border = thin_border
            ri += 1

    for ci, w in enumerate([25, 12, 14, 12, 25, 15, 50], 1):
        ws_cat.column_dimensions[chr(64 + ci)].width = w

    wb.save(output_path)


def _export_with_pandas(analysis_result: dict, output_path: str):
    """Fallback export using pandas if openpyxl is not available standalone."""
    import pandas as pd

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Summary sheet
        summary_data = {
            "Metric": ["Total Vouchers", "Narrations Analyzed", "No Narration Count",
                        "HIGH Severity", "MEDIUM Severity", "LOW Severity"],
            "Value": [
                analysis_result.get("total_vouchers", 0),
                analysis_result.get("narrations_analyzed", 0),
                analysis_result.get("no_narration_count", 0),
                analysis_result.get("risk_summary", {}).get("high", 0),
                analysis_result.get("risk_summary", {}).get("medium", 0),
                analysis_result.get("risk_summary", {}).get("low", 0),
            ],
        }
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)

        # Flagged vouchers
        flagged = analysis_result.get("flagged_vouchers", [])
        if flagged:
            df_flag = pd.DataFrame(flagged)
            df_flag["categories"] = df_flag["categories"].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)
            df_flag["comments"] = df_flag["comments"].apply(lambda x: "; ".join(x) if isinstance(x, list) else x)
            cols = ["date", "voucher_type", "voucher_number", "party", "amount",
                    "narration", "categories", "comments", "severity"]
            df_flag = df_flag[[c for c in cols if c in df_flag.columns]]
            df_flag.to_excel(writer, sheet_name="Flagged Vouchers", index=False)

        # Category details
        cat_rows = []
        for cname, cdata in analysis_result.get("category_summary", {}).items():
            for v in cdata.get("vouchers", []):
                cat_rows.append({"category": cname, **v})
        if cat_rows:
            pd.DataFrame(cat_rows).to_excel(writer, sheet_name="Category Details", index=False)


# ── HELPERS ──────────────────────────────────────────────────────────────────

def _format_date(date_str: str) -> str:
    """Convert Tally date format (YYYYMMDD) to display format (DD-Mon-YYYY)."""
    if not date_str or len(date_str) < 8:
        return date_str
    try:
        dt = datetime.strptime(date_str[:8], "%Y%m%d")
        return dt.strftime("%d-%b-%Y")
    except (ValueError, TypeError):
        return date_str


def _highest_severity(matches: list[dict]) -> str:
    """Return highest severity among a list of match dicts."""
    if not matches:
        return "LOW"
    best = "LOW"
    for m in matches:
        sev = m.get("severity", "LOW")
        if SEVERITY_ORDER.get(sev, 9) < SEVERITY_ORDER.get(best, 9):
            best = sev
    return best


# ── CLI TEST ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    db = os.path.join(os.path.dirname(__file__), "tally_data.db")
    if not os.path.exists(db):
        print(f"Database not found at {db}")
        exit(1)

    print("Running narration analysis...")
    result = analyze_all_narrations(db)

    print(f"\nTotal vouchers:       {result['total_vouchers']}")
    print(f"Narrations analyzed:  {result['narrations_analyzed']}")
    print(f"No narration:         {result['no_narration_count']}")
    print(f"\nRisk summary: {result['risk_summary']}")

    print("\nCategory summary:")
    for cat, data in sorted(result["category_summary"].items(), key=lambda x: -x[1]["count"]):
        print(f"  {cat:30s}  Count: {data['count']:>5d}  Amount: {data['total_amount']:>15,.2f}")

    print(f"\nFlagged vouchers (top 10):")
    for v in result["flagged_vouchers"][:10]:
        print(f"  [{v['severity']}] {v['date']} {v['voucher_type']:12s} {v['voucher_number']:>8s} "
              f"Rs {v['amount']:>12,.2f}  {v['narration'][:60]}")
        print(f"         Categories: {', '.join(v['categories'])}")
