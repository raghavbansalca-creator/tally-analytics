"""
Seven Labs Vision — Narration Cross-Check Dashboard
Compares GROUP classification vs NARRATION classification independently.
Flags disagreements as potential misclassifications for partner review.
"""

import streamlit as st
import pandas as pd
import sys
import os
import io
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from narration_classifier import classify_all
from styles import (
    inject_base_styles, page_header, section_header, metric_card,
    badge, footer, fmt,
)

st.set_page_config(page_title="Narration Cross-Check", page_icon="", layout="wide")
inject_base_styles()

DB_PATH = st.session_state.get("db_path", "") or os.path.join(os.path.dirname(os.path.dirname(__file__)), "tally_data.db")

if not os.path.exists(DB_PATH):
    page_header("Narration Cross-Check", "No database loaded")
    st.warning("No Tally database found. Please load a database from the Setup page.")
    st.stop()


# ── Helpers ──────────────────────────────────────────────────────────────────

def fmt_inr(amount):
    if amount is None:
        return "Rs 0"
    try:
        amount = float(amount)
    except (ValueError, TypeError):
        return "Rs 0"
    abs_amt = abs(amount)
    sign = "-" if amount < 0 else ""
    if abs_amt >= 10000000:
        return f"{sign}Rs {abs_amt / 10000000:.2f} Cr"
    elif abs_amt >= 100000:
        return f"{sign}Rs {abs_amt / 100000:.2f} L"
    elif abs_amt >= 1000:
        return f"{sign}Rs {abs_amt:,.0f}"
    else:
        return f"{sign}Rs {abs_amt:.2f}"


def fmt_date(d):
    if not d or len(d) < 8:
        return d or ""
    try:
        return f"{d[6:8]}-{d[4:6]}-{d[:4]}"
    except Exception:
        return d


def verdict_badge(verdict):
    v = (verdict or "").upper()
    if "AGREE" in v:
        return badge("AGREE", "green")
    elif "DISAGREE" in v or "CONTRADICT" in v:
        return badge("CONFLICT", "red")
    elif "NO_NARRATION" in v:
        return badge("NO NARR", "amber")
    elif "GROUP_ONLY" in v:
        return badge("GRP ONLY", "blue")
    return badge(verdict or "?", "gray")


def severity_badge(severity):
    s = (severity or "").upper()
    if s == "HIGH":
        return badge("HIGH", "red")
    elif s == "MEDIUM":
        return badge("MED", "amber")
    return badge(s or "N/A", "gray")


# ── Company name ─────────────────────────────────────────────────────────────

try:
    import sqlite3
    _conn = sqlite3.connect(DB_PATH)
    # Check _metadata table exists before querying
    _tbl_check = _conn.execute(
        "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='_metadata'"
    ).fetchone()
    if _tbl_check and _tbl_check[0] > 0:
        _row = _conn.execute("SELECT value FROM _metadata WHERE key='company_name'").fetchone()
        COMPANY = _row[0] if _row else "Company"
    else:
        COMPANY = "Company"
    _conn.close()
except Exception:
    COMPANY = "Company"


# ── Header ───────────────────────────────────────────────────────────────────

page_header(
    "Narration Cross-Check",
    f"Independent comparison: Group classification vs Narration classification | {COMPANY}",
)


# ── Session state ────────────────────────────────────────────────────────────

if "crosscheck_results" not in st.session_state:
    st.session_state.crosscheck_results = None


# ── Run button ───────────────────────────────────────────────────────────────

col_btn, col_status = st.columns([1, 4])
with col_btn:
    run_clicked = st.button("Run Cross-Check Analysis", type="primary")
with col_status:
    if st.session_state.crosscheck_results is None:
        st.info("Click to run the multi-layer narration vs group cross-check on all vouchers.")

if run_clicked:
    with st.spinner("Running 4-layer classification pipeline on all vouchers..."):
        st.session_state.crosscheck_results = classify_all(DB_PATH)

pipeline = st.session_state.crosscheck_results
if pipeline is None:
    st.stop()

stats = pipeline["stats"]
results = pipeline["results"]


# ── Summary Metrics ──────────────────────────────────────────────────────────

section_header("OVERVIEW")

m1, m2, m3, m4, m5 = st.columns(5)
with m1:
    metric_card("Total Vouchers", f"{stats['total']:,}", color_class="blue")
with m2:
    agree_count = stats["by_verdict"].get("AGREE", 0) + stats["by_verdict"].get("AGREE_NARRATION_MORE_SPECIFIC", 0)
    metric_card("Group & Narration Agree", f"{agree_count:,}", color_class="green")
with m3:
    conflict_count = stats.get("group_narration_disagreements", 0)
    cc = "red" if conflict_count > 0 else "green"
    metric_card("Conflicts Found", f"{conflict_count:,}", color_class=cc)
with m4:
    no_narr = stats.get("no_narration", 0)
    metric_card("No Narration", f"{no_narr:,}", color_class="amber")
with m5:
    group_only = stats["by_verdict"].get("GROUP_ONLY_NARRATION_UNCLEAR", 0)
    metric_card("Group Only", f"{group_only:,}", color_class="blue")

st.markdown("")


# ── Verdict breakdown bar ────────────────────────────────────────────────────

section_header("VERDICT BREAKDOWN")

verdict_df_data = []
for verdict, count in sorted(stats["by_verdict"].items(), key=lambda x: -x[1]):
    pct = count / stats["total"] * 100 if stats["total"] > 0 else 0
    verdict_df_data.append({"Verdict": verdict, "Count": count, "Percentage": f"{pct:.1f}%"})

if verdict_df_data:
    st.dataframe(
        pd.DataFrame(verdict_df_data),
        use_container_width=True,
        hide_index=True,
        height=min(300, 40 + len(verdict_df_data) * 35),
    )

st.markdown("")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION: CONFLICTS (the audit gold)
# ═══════════════════════════════════════════════════════════════════════════════

conflicts = [r for r in results if r.get("verdict") in (
    "DISAGREE_POSSIBLE_MISCLASSIFICATION", "NARRATION_CONTRADICTS_GROUP"
)]

section_header(f"CONFLICTS — Group vs Narration Disagree ({len(conflicts)} found)")

if conflicts:
    st.error(
        f"**{len(conflicts)} entries** where the narration implies a different classification "
        f"than what the ledger grouping suggests. These are potential misclassifications "
        f"requiring partner review."
    )

    conflict_rows = []
    for r in conflicts:
        severity = r.get("cross_check", {}).get("severity", "MEDIUM") if r.get("cross_check") else "MEDIUM"
        conflict_rows.append({
            "Date": fmt_date(r["date"]),
            "Vch No": r["voucher_number"],
            "Type": r["voucher_type"],
            "Party": r["party"],
            "Amount": r["amount"],
            "GROUP Says": r.get("group_says", ""),
            "NARRATION Says": r.get("narration_says", ""),
            "Severity": severity,
            "Suggested Group": r.get("suggested_correct_group", ""),
            "Narration": r["narration"][:120],
            "Audit Comment": r["comment"][:150],
        })

    df_conflicts = pd.DataFrame(conflict_rows)
    df_conflicts["Amount"] = df_conflicts["Amount"].apply(fmt_inr)

    st.dataframe(
        df_conflicts,
        use_container_width=True,
        hide_index=True,
        height=min(500, 40 + len(df_conflicts) * 35),
        column_config={
            "Narration": st.column_config.TextColumn(width="large"),
            "Audit Comment": st.column_config.TextColumn(width="large"),
        },
    )
else:
    st.success("No conflicts found — all classifiable vouchers have consistent grouping and narration.")

st.markdown("")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION: NO NARRATION (highlighted for review)
# ═══════════════════════════════════════════════════════════════════════════════

no_narr_items = [r for r in results if r.get("verdict") == "NO_NARRATION"]

section_header(f"NO NARRATION — Group Only ({len(no_narr_items)} vouchers)")

if no_narr_items:
    st.warning(
        f"**{len(no_narr_items)} vouchers** have no narration. Classification is based on "
        f"ledger grouping only. Transaction purpose cannot be independently verified."
    )

    # Show top by amount
    no_narr_sorted = sorted(no_narr_items, key=lambda x: -abs(x.get("amount", 0)))
    display_limit = min(200, len(no_narr_sorted))

    narr_rows = []
    for r in no_narr_sorted[:display_limit]:
        narr_rows.append({
            "Date": fmt_date(r["date"]),
            "Vch No": r["voucher_number"],
            "Type": r["voucher_type"],
            "Party": r["party"],
            "Amount": r["amount"],
            "Group Classification": r["category"],
            "Confidence": f"{r['confidence']:.0%}",
            "Method": r["method"],
        })

    df_nonarr = pd.DataFrame(narr_rows)
    df_nonarr["Amount"] = df_nonarr["Amount"].apply(fmt_inr)

    st.dataframe(
        df_nonarr,
        use_container_width=True,
        hide_index=True,
        height=min(500, 40 + len(df_nonarr) * 35),
    )

    if len(no_narr_sorted) > display_limit:
        st.caption(f"Showing top {display_limit} by amount. Download Excel for all {len(no_narr_sorted)}.")

st.markdown("")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION: CATEGORY DISTRIBUTION
# ═══════════════════════════════════════════════════════════════════════════════

section_header("CLASSIFICATION SUMMARY")

cat_data = []
for cat, count in sorted(stats["by_category"].items(), key=lambda x: -x[1]):
    pct = count / stats["total"] * 100 if stats["total"] > 0 else 0
    cat_data.append({"Category": cat, "Count": count, "Percentage": f"{pct:.1f}%"})

if cat_data:
    col_chart, col_table = st.columns([1, 1])

    with col_table:
        st.dataframe(
            pd.DataFrame(cat_data),
            use_container_width=True,
            hide_index=True,
            height=min(500, 40 + len(cat_data) * 35),
        )

    with col_chart:
        chart_df = pd.DataFrame(cat_data[:10])
        st.bar_chart(chart_df.set_index("Category")["Count"])

st.markdown("")


# ═══════════════════════════════════════════════════════════════════════════════
# EXPORT: Review Excel with 5 tabs
# ═══════════════════════════════════════════════════════════════════════════════

section_header("EXPORT REVIEW FILE")

st.markdown(
    "Download the complete review Excel with 5 tabs: "
    "**AGREE** (green), **CONFLICT** (red — needs partner sign-off), "
    "**NO NARRATION** (amber), **GROUP ONLY** (blue), and **SUMMARY**."
)

if st.button("Generate Review Excel", type="primary"):
    with st.spinner("Generating audit review file..."):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()

        hdr_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        hdr_fill_green = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        hdr_fill_red = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
        hdr_fill_amber = PatternFill(start_color="EF6C00", end_color="EF6C00", fill_type="solid")
        hdr_fill_blue = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        hdr_fill_navy = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
        data_font = Font(name="Arial", size=10)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        def write_header(ws, row, cols, fill):
            for col_idx, col_name in enumerate(cols, 1):
                cell = ws.cell(row=row, column=col_idx, value=col_name)
                cell.font = hdr_font
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = thin_border

        def write_row(ws, row, data):
            for col_idx, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # ═══ SHEET 1: AGREE ═══
        ws1 = wb.active
        ws1.title = "AGREE - No Issues"
        cols1 = ["Date", "Vch No", "Vch Type", "Party", "Amount", "Group Says",
                 "Narration Says", "Final Category", "Confidence", "Method", "Narration"]
        write_header(ws1, 1, cols1, hdr_fill_green)

        agree_items = [r for r in results if r.get("verdict") in ("AGREE", "AGREE_NARRATION_MORE_SPECIFIC")]
        row = 2
        for r in agree_items:
            write_row(ws1, row, [
                fmt_date(r["date"]), r["voucher_number"], r["voucher_type"], r["party"],
                round(r["amount"], 2), r.get("group_says", ""), r.get("narration_says", ""),
                r["category"], r["confidence"], r["method"], (r["narration"] or "")[:100],
            ])
            row += 1

        # ═══ SHEET 2: CONFLICT ═══
        ws2 = wb.create_sheet("CONFLICT - Misclassification")
        cols2 = ["Date", "Vch No", "Vch Type", "Party", "Amount", "GROUP Says",
                 "NARRATION Says", "Severity", "Suggested Correct Group",
                 "Audit Comment", "Partner Decision", "Narration"]
        write_header(ws2, 1, cols2, hdr_fill_red)

        conflict_items = [r for r in results if r.get("verdict") in (
            "DISAGREE_POSSIBLE_MISCLASSIFICATION", "NARRATION_CONTRADICTS_GROUP"
        )]
        row = 2
        for r in conflict_items:
            sev = r.get("cross_check", {}).get("severity", "MEDIUM") if r.get("cross_check") else "MEDIUM"
            write_row(ws2, row, [
                fmt_date(r["date"]), r["voucher_number"], r["voucher_type"], r["party"],
                round(r["amount"], 2), r.get("group_says", ""), r.get("narration_says", ""),
                sev, r.get("suggested_correct_group", ""), (r["comment"] or "")[:150],
                "",  # Partner Decision — blank for sign-off
                (r["narration"] or "")[:120],
            ])
            if sev == "HIGH":
                ws2.cell(row=row, column=8).fill = PatternFill(start_color="FFCDD2", fill_type="solid")
            row += 1

        # ═══ SHEET 3: NO NARRATION ═══
        ws3 = wb.create_sheet("NO NARRATION - Group Only")
        cols3 = ["Date", "Vch No", "Vch Type", "Party", "Amount",
                 "Group Classification", "Confidence", "Method", "Comment", "Reviewer Note"]
        write_header(ws3, 1, cols3, hdr_fill_amber)

        no_narr_list = [r for r in results if r.get("verdict") == "NO_NARRATION"]
        row = 2
        for r in no_narr_list:
            write_row(ws3, row, [
                fmt_date(r["date"]), r["voucher_number"], r["voucher_type"], r["party"],
                round(r["amount"], 2), r["category"], r["confidence"],
                r["method"], (r["comment"] or "")[:120], "",
            ])
            row += 1

        # ═══ SHEET 4: GROUP ONLY ═══
        ws4 = wb.create_sheet("GROUP ONLY - Narration Unclear")
        cols4 = ["Date", "Vch No", "Vch Type", "Party", "Amount",
                 "Group Classification", "Confidence", "Method", "Narration", "Comment"]
        write_header(ws4, 1, cols4, hdr_fill_blue)

        group_only_list = [r for r in results if r.get("verdict") == "GROUP_ONLY_NARRATION_UNCLEAR"]
        row = 2
        for r in group_only_list:
            write_row(ws4, row, [
                fmt_date(r["date"]), r["voucher_number"], r["voucher_type"], r["party"],
                round(r["amount"], 2), r["category"], r["confidence"],
                r["method"], (r["narration"] or "")[:100], (r["comment"] or "")[:100],
            ])
            row += 1

        # ═══ SHEET 5: SUMMARY ═══
        ws5 = wb.create_sheet("SUMMARY")
        ws5.sheet_properties.tabColor = "1565C0"
        summary_rows = [
            (f"NARRATION AUDIT REVIEW — {COMPANY}", "", ""),
            ("", "", ""),
            ("Total Vouchers", stats["total"], ""),
            ("", "", ""),
            ("SECTION", "COUNT", "PERCENTAGE"),
            ("AGREE (group & narration match)", len(agree_items),
             f'{len(agree_items)/max(stats["total"],1)*100:.1f}%'),
            ("CONFLICT (group vs narration disagree)", len(conflict_items),
             f'{len(conflict_items)/max(stats["total"],1)*100:.1f}%'),
            ("NO NARRATION (group-only, highlighted)", len(no_narr_list),
             f'{len(no_narr_list)/max(stats["total"],1)*100:.1f}%'),
            ("GROUP ONLY (narration unclear, group used)", len(group_only_list),
             f'{len(group_only_list)/max(stats["total"],1)*100:.1f}%'),
            ("", "", ""),
            ("AUDIT METRICS", "", ""),
            ("Needs Manual Review", stats["needs_review"], ""),
            ("Group-Narration Disagreements", stats.get("group_narration_disagreements", 0), ""),
            ("HIGH Severity Mismatches", stats.get("high_severity_mismatches", 0), ""),
            ("No Narration Vouchers", stats.get("no_narration", 0), ""),
            ("Bank Statement Narrations", stats.get("bank_narrations", 0), ""),
        ]
        for i, (a, b, c) in enumerate(summary_rows, 1):
            ws5.cell(row=i, column=1, value=a).font = Font(name="Arial", size=11, bold=(i <= 1 or i == 5 or i == 11))
            ws5.cell(row=i, column=2, value=b).font = Font(name="Arial", size=11)
            ws5.cell(row=i, column=3, value=c).font = Font(name="Arial", size=11)

        # Column widths
        for ws in [ws1, ws2, ws3, ws4]:
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 25
            ws.column_dimensions["E"].width = 14
            for col in "FGHIJKL":
                ws.column_dimensions[col].width = 22
        ws5.column_dimensions["A"].width = 45
        ws5.column_dimensions["B"].width = 15
        ws5.column_dimensions["C"].width = 15

        # Save to buffer
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        st.download_button(
            label="Download Narration Cross-Check Review (.xlsx)",
            data=buf.getvalue(),
            file_name=f"Narration_CrossCheck_Review_{COMPANY.replace(' ', '_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.success(
            f"Review file generated: {len(agree_items)} AGREE | "
            f"{len(conflict_items)} CONFLICT | "
            f"{len(no_narr_list)} NO NARRATION | "
            f"{len(group_only_list)} GROUP ONLY"
        )


# ── Footer ───────────────────────────────────────────────────────────────────

footer(COMPANY)
