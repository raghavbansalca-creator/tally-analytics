"""
Seven Labs Vision -- Companies Act Schedule III Financial Statement Exporter
Generates Balance Sheet (Part I), P&L Statement (Part II), and Notes to Accounts
in Excel format with professional formatting compliant with Companies Act, 2013.
"""

import sqlite3
import os
from datetime import datetime
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

DB_PATH = os.path.join(os.path.dirname(__file__), "tally_data.db")


# ═══════════════════════════════════════════════════════════════════════════════
# TALLY GROUP → SCHEDULE III MAPPING
# ═══════════════════════════════════════════════════════════════════════════════

# --- Balance Sheet: Equity & Liabilities ---
SCHEDULE_III_LIABILITIES = OrderedDict([
    ("shareholders_funds", OrderedDict([
        ("share_capital", {
            "label": "Share Capital",
            "note": "2",
            "groups": ["Share Capital Account"],
            "exclude_groups": ["Reserves & Surplus"],
        }),
        ("reserves_surplus", {
            "label": "Reserves and Surplus",
            "note": "3",
            "groups": ["Reserves & Surplus"],
            "include_pl": True,
        }),
        ("money_against_warrants", {
            "label": "Money received against share warrants",
            "note": "",
            "groups": [],
        }),
    ])),
    ("share_application", OrderedDict([
        ("share_app_money", {
            "label": "Share application money pending allotment",
            "note": "",
            "groups": [],
        }),
    ])),
    ("non_current_liabilities", OrderedDict([
        ("long_term_borrowings", {
            "label": "Long-term borrowings",
            "note": "4",
            "groups": ["Secured Loans", "Unsecured Loans"],
        }),
        ("deferred_tax_liabilities", {
            "label": "Deferred tax liabilities (Net)",
            "note": "5",
            "groups": [],
            "special": "deferred_tax_liability",
        }),
        ("other_long_term_liabilities", {
            "label": "Other Long term liabilities",
            "note": "",
            "groups": [],
        }),
        ("long_term_provisions", {
            "label": "Long term provisions",
            "note": "",
            "groups": [],
        }),
    ])),
    ("current_liabilities", OrderedDict([
        ("short_term_borrowings", {
            "label": "Short-term borrowings",
            "note": "6",
            "groups": ["Bank OD A/c"],
        }),
        ("trade_payables", {
            "label": "Trade payables",
            "note": "7",
            "groups": ["Sundry Creditors"],
        }),
        ("other_current_liabilities", {
            "label": "Other current liabilities",
            "note": "8",
            "groups": ["Duties & Taxes", "Salary Payable"],
            "_direct_groups": ["Current Liabilities"],
            "exclude_ledgers_containing": ["deferred tax"],
        }),
        ("short_term_provisions", {
            "label": "Short-term provisions",
            "note": "9",
            "groups": ["Provisions"],
        }),
    ])),
])

# --- Balance Sheet: Assets ---
SCHEDULE_III_ASSETS = OrderedDict([
    ("non_current_assets", OrderedDict([
        ("ppe", {
            "label": "Property, Plant and Equipment",
            "note": "10",
            "groups": ["Fixed Assets"],
        }),
        ("non_current_investments", {
            "label": "Non-current investments",
            "note": "11",
            "groups": ["Investments"],
        }),
        ("deferred_tax_assets", {
            "label": "Deferred tax assets (net)",
            "note": "5",
            "groups": [],
            "special": "deferred_tax_asset",
        }),
        ("long_term_loans_advances", {
            "label": "Long term loans and advances",
            "note": "",
            "groups": [],
        }),
        ("other_non_current_assets", {
            "label": "Other non-current assets",
            "note": "",
            "groups": ["Misc. Expenses (ASSET)"],
        }),
    ])),
    ("current_assets", OrderedDict([
        ("current_investments", {
            "label": "Current investments",
            "note": "",
            "groups": [],
        }),
        ("inventories", {
            "label": "Inventories",
            "note": "12",
            "groups": ["Stock-in-Hand"],
        }),
        ("trade_receivables", {
            "label": "Trade receivables",
            "note": "13",
            "groups": ["Sundry Debtors"],
        }),
        ("cash_equivalents", {
            "label": "Cash and cash equivalents",
            "note": "14",
            "groups": ["Cash-in-Hand", "Bank Accounts"],
        }),
        ("short_term_loans_advances", {
            "label": "Short-term loans and advances",
            "note": "15",
            "groups": ["Loans & Advances (Asset)", "Loan to Employees", "Deposits (Asset)"],
        }),
        ("other_current_assets", {
            "label": "Other current assets",
            "note": "16",
            "groups": [],
            "_direct_groups": ["Current Assets"],
            "exclude_ledgers_containing": ["deferred tax"],
        }),
    ])),
])

# --- P&L Mapping ---
SCHEDULE_III_PL = OrderedDict([
    ("revenue_from_operations", {
        "label": "Revenue from operations",
        "note": "17",
        "groups": ["Sales Accounts"],
    }),
    ("other_income", {
        "label": "Other Income",
        "note": "18",
        "groups": ["Direct Incomes", "Indirect Incomes"],
    }),
])

SCHEDULE_III_EXPENSES = OrderedDict([
    ("cost_of_materials", {
        "label": "Cost of materials consumed",
        "note": "19",
        "groups": ["Purchase Accounts"],
    }),
    ("purchase_stock_in_trade", {
        "label": "Purchase of Stock-in-Trade",
        "note": "",
        "groups": [],
    }),
    ("changes_in_inventories", {
        "label": "Changes in inventories of finished goods, WIP and Stock-in-Trade",
        "note": "20",
        "groups": [],
        "special": "inventory_change",
    }),
    ("employee_benefit", {
        "label": "Employee benefit expense",
        "note": "21",
        "groups": ["Salary Expenses"],
    }),
    ("finance_costs", {
        "label": "Finance costs",
        "note": "22",
        "groups": [],
        "match_names": ["interest", "bank charges", "bank od interest"],
    }),
    ("depreciation", {
        "label": "Depreciation and amortization expense",
        "note": "10",
        "groups": [],
        "match_names": ["depreciation", "amortization", "amortisation"],
    }),
    ("other_expenses", {
        "label": "Other expenses",
        "note": "23",
        "groups": ["Direct Expenses", "Indirect Expenses"],
        "exclude_groups": ["Salary Expenses"],
        "exclude_names": ["interest", "bank charges", "bank od interest",
                          "depreciation", "amortization", "amortisation"],
    }),
])


# ═══════════════════════════════════════════════════════════════════════════════
# DATABASE HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _get_conn(db_path=None):
    return sqlite3.connect(db_path or DB_PATH)


def _get_all_groups_under(conn, root_groups):
    """Recursively get all group names under root groups."""
    all_groups = set()
    queue = list(root_groups)
    while queue:
        parent = queue.pop(0)
        all_groups.add(parent)
        children = conn.execute(
            "SELECT NAME FROM mst_group WHERE PARENT = ?", (parent,)
        ).fetchall()
        for (child,) in children:
            if child not in all_groups:
                queue.append(child)
    return all_groups


def _get_ledger_balances(conn, group_names, exclude_groups=None, exclude_names=None):
    """Get sum of closing balances for ledgers under given groups."""
    if not group_names:
        return 0.0

    all_groups = set()
    for g in group_names:
        all_groups |= _get_all_groups_under(conn, [g])

    # Remove excluded groups and their children
    if exclude_groups:
        for eg in exclude_groups:
            excluded = _get_all_groups_under(conn, [eg])
            all_groups -= excluded

    if not all_groups:
        return 0.0

    placeholders = ",".join(["?"] * len(all_groups))
    sql = f"""
    SELECT COALESCE(SUM(CAST(CLOSINGBALANCE AS REAL)), 0)
    FROM mst_ledger
    WHERE PARENT IN ({placeholders})
    """
    params = list(all_groups)

    if exclude_names:
        for en in exclude_names:
            sql += " AND LOWER(NAME) NOT LIKE ?"
            params.append(f"%{en.lower()}%")

    row = conn.execute(sql, params).fetchone()
    return row[0] if row else 0.0


def _get_direct_ledger_balances(conn, group_names, exclude_names=None):
    """Get sum of closing balances for ledgers DIRECTLY under given groups."""
    if not group_names:
        return 0.0
    placeholders = ",".join(["?"] * len(group_names))
    sql = f"""
    SELECT COALESCE(SUM(CAST(CLOSINGBALANCE AS REAL)), 0)
    FROM mst_ledger
    WHERE PARENT IN ({placeholders})
    """
    params = list(group_names)
    if exclude_names:
        for en in exclude_names:
            sql += " AND LOWER(NAME) NOT LIKE ?"
            params.append(f"%{en.lower()}%")
    row = conn.execute(sql, params).fetchone()
    return row[0] if row else 0.0


def _get_ledger_balances_by_name(conn, match_names, parent_groups=None):
    """Get sum of closing balances for ledgers matching name patterns."""
    if not match_names:
        return 0.0
    conditions = " OR ".join(["LOWER(NAME) LIKE ?" for _ in match_names])
    params = [f"%{n.lower()}%" for n in match_names]
    sql = f"SELECT COALESCE(SUM(CAST(CLOSINGBALANCE AS REAL)), 0) FROM mst_ledger WHERE ({conditions})"
    if parent_groups:
        all_groups = set()
        conn2 = conn
        for g in parent_groups:
            all_groups |= _get_all_groups_under(conn2, [g])
        if all_groups:
            ph = ",".join(["?"] * len(all_groups))
            sql += f" AND PARENT IN ({ph})"
            params.extend(list(all_groups))
    row = conn.execute(sql, params).fetchone()
    return row[0] if row else 0.0


def _get_ledger_details(conn, group_names, exclude_groups=None, exclude_names=None):
    """Get individual ledger details for given groups."""
    if not group_names:
        return []
    all_groups = set()
    for g in group_names:
        all_groups |= _get_all_groups_under(conn, [g])
    if exclude_groups:
        for eg in exclude_groups:
            excluded = _get_all_groups_under(conn, [eg])
            all_groups -= excluded
    if not all_groups:
        return []
    placeholders = ",".join(["?"] * len(all_groups))
    sql = f"""
    SELECT NAME, PARENT, CAST(OPENINGBALANCE AS REAL) as ob,
           CAST(CLOSINGBALANCE AS REAL) as cb
    FROM mst_ledger WHERE PARENT IN ({placeholders})
    """
    params = list(all_groups)
    if exclude_names:
        for en in exclude_names:
            sql += " AND LOWER(NAME) NOT LIKE ?"
            params.append(f"%{en.lower()}%")
    sql += " ORDER BY PARENT, NAME"
    return conn.execute(sql, params).fetchall()


def _get_direct_ledger_details(conn, group_names, exclude_names=None):
    """Get individual ledger details DIRECTLY under given groups."""
    if not group_names:
        return []
    placeholders = ",".join(["?"] * len(group_names))
    sql = f"""
    SELECT NAME, PARENT, CAST(OPENINGBALANCE AS REAL) as ob,
           CAST(CLOSINGBALANCE AS REAL) as cb
    FROM mst_ledger WHERE PARENT IN ({placeholders})
    """
    params = list(group_names)
    if exclude_names:
        for en in exclude_names:
            sql += " AND LOWER(NAME) NOT LIKE ?"
            params.append(f"%{en.lower()}%")
    sql += " ORDER BY PARENT, NAME"
    return conn.execute(sql, params).fetchall()


def _get_metadata(conn):
    """Fetch company metadata from _metadata table."""
    try:
        rows = conn.execute("SELECT * FROM _metadata").fetchall()
        return {r[0]: r[1] for r in rows}
    except Exception:
        return {}


def _get_stock_ledger_balances(conn):
    """Get opening and closing stock from Stock-in-Hand ledgers."""
    rows = conn.execute("""
        SELECT NAME, CAST(OPENINGBALANCE AS REAL) as ob,
               CAST(CLOSINGBALANCE AS REAL) as cb
        FROM mst_ledger WHERE PARENT IN (
            SELECT NAME FROM mst_group
            WHERE NAME = 'Stock-in-Hand'
            UNION SELECT 'Stock-in-Hand'
        )
    """).fetchall()
    opening = sum(abs(r[1]) for r in rows if r[1])
    closing = sum(abs(r[2]) for r in rows if r[2])
    details = [(r[0], abs(r[1]) if r[1] else 0, abs(r[2]) if r[2] else 0) for r in rows]
    return opening, closing, details


# ═══════════════════════════════════════════════════════════════════════════════
# DATA EXTRACTION
# ═══════════════════════════════════════════════════════════════════════════════

def _compute_item_amount(conn, config, side="liability"):
    """Compute amount for a Schedule III line item.
    side='liability': positive (credit) = amount to show
    side='asset': negate (debit negative -> positive asset)
    """
    amt = 0.0

    if config.get("groups"):
        amt += _get_ledger_balances(
            conn, config["groups"],
            exclude_groups=config.get("exclude_groups"),
            exclude_names=config.get("exclude_names"),
        )

    if config.get("_direct_groups"):
        amt += _get_direct_ledger_balances(
            conn, config["_direct_groups"],
            exclude_names=config.get("exclude_ledgers_containing"),
        )

    if config.get("match_names"):
        parent_filter = config.get("match_parent_groups")
        amt += _get_ledger_balances_by_name(conn, config["match_names"], parent_filter)

    if side == "asset":
        amt = -amt  # flip: debit (negative) -> positive

    return amt


def extract_balance_sheet_data(conn):
    """Extract all Balance Sheet line items mapped to Schedule III."""

    # Get deferred tax raw value (sum of ALL deferred tax ledgers)
    dt_row = conn.execute(
        "SELECT COALESCE(SUM(CAST(CLOSINGBALANCE AS REAL)), 0) FROM mst_ledger WHERE LOWER(NAME) LIKE '%deferred tax%'"
    ).fetchone()
    dt_value = dt_row[0] if dt_row else 0.0

    # Extract liabilities
    liabilities = OrderedDict()
    for section_key, items in SCHEDULE_III_LIABILITIES.items():
        section_data = OrderedDict()
        for item_key, config in items.items():
            special = config.get("special", "")

            if special == "deferred_tax_liability":
                amt = dt_value if dt_value > 0 else 0.0
            elif config.get("include_pl"):
                # Reserves & Surplus + P&L balance
                amt = _compute_item_amount(conn, config, side="liability")
                pl_row = conn.execute(
                    "SELECT CAST(CLOSINGBALANCE AS REAL) FROM mst_ledger WHERE NAME = 'Profit & Loss A/c'"
                ).fetchone()
                if pl_row and pl_row[0]:
                    amt += pl_row[0]
            elif config.get("exclude_groups"):
                # Share Capital excluding Reserves & Surplus sub-groups
                all_grps = set()
                for g in config["groups"]:
                    all_grps |= _get_all_groups_under(conn, [g])
                for eg in config.get("exclude_groups", []):
                    excluded = _get_all_groups_under(conn, [eg])
                    all_grps -= excluded
                if all_grps:
                    ph = ",".join(["?"] * len(all_grps))
                    row = conn.execute(
                        f"SELECT COALESCE(SUM(CAST(CLOSINGBALANCE AS REAL)), 0) FROM mst_ledger WHERE PARENT IN ({ph})",
                        list(all_grps)
                    ).fetchone()
                    amt = row[0] if row else 0.0
                else:
                    amt = 0.0
            else:
                amt = _compute_item_amount(conn, config, side="liability")

            # Note: Deferred tax is already excluded from OCL via exclude_ledgers_containing
            # in the config, so no manual subtraction needed here.

            section_data[item_key] = {
                "label": config["label"],
                "note": config.get("note", ""),
                "amount": amt,
            }
        liabilities[section_key] = section_data

    # Extract assets
    assets = OrderedDict()
    for section_key, items in SCHEDULE_III_ASSETS.items():
        section_data = OrderedDict()
        for item_key, config in items.items():
            special = config.get("special", "")

            if special == "deferred_tax_asset":
                amt = abs(dt_value) if dt_value < 0 else 0.0
            else:
                amt = _compute_item_amount(conn, config, side="asset")

            section_data[item_key] = {
                "label": config["label"],
                "note": config.get("note", ""),
                "amount": amt,
            }
        assets[section_key] = section_data

    # Handle Suspense A/c
    suspense_raw = _get_ledger_balances(conn, ["Suspense A/c"])
    if suspense_raw != 0:
        if suspense_raw < 0:  # debit = asset
            assets["current_assets"]["other_current_assets"]["amount"] += abs(suspense_raw)
        else:  # credit = liability
            liabilities["current_liabilities"]["other_current_liabilities"]["amount"] += suspense_raw

    total_liabilities = sum(
        item["amount"] for section in liabilities.values() for item in section.values()
    )
    total_assets = sum(
        item["amount"] for section in assets.values() for item in section.values()
    )

    return {
        "liabilities": liabilities,
        "assets": assets,
        "total_liabilities": total_liabilities,
        "total_assets": total_assets,
    }


def extract_pl_data(conn):
    """Extract all P&L Statement line items."""

    # Revenue (credit balances -> positive for display)
    revenue_raw = _get_ledger_balances(conn, SCHEDULE_III_PL["revenue_from_operations"]["groups"])
    revenue = abs(revenue_raw)

    other_income_raw = _get_ledger_balances(conn, SCHEDULE_III_PL["other_income"]["groups"])
    other_income = abs(other_income_raw)

    total_income = revenue + other_income

    # Expenses
    expenses = OrderedDict()
    for key, config in SCHEDULE_III_EXPENSES.items():
        special = config.get("special", "")

        if special == "inventory_change":
            opening, closing, _ = _get_stock_ledger_balances(conn)
            amt = opening - closing  # decrease in stock = expense
            expenses[key] = {"label": config["label"], "note": config.get("note", ""), "amount": amt}
        elif config.get("match_names") and not config.get("groups"):
            # Finance costs, depreciation: match by name across all expense groups
            raw = _get_ledger_balances_by_name(conn, config["match_names"])
            expenses[key] = {"label": config["label"], "note": config.get("note", ""), "amount": abs(raw)}
        elif key == "other_expenses":
            raw = _get_ledger_balances(
                conn, config["groups"],
                exclude_groups=config.get("exclude_groups"),
                exclude_names=config.get("exclude_names"),
            )
            expenses[key] = {"label": config["label"], "note": config.get("note", ""), "amount": abs(raw)}
        else:
            raw = _get_ledger_balances(
                conn, config.get("groups", []),
                exclude_groups=config.get("exclude_groups"),
                exclude_names=config.get("exclude_names"),
            )
            expenses[key] = {"label": config["label"], "note": config.get("note", ""), "amount": abs(raw)}

    total_expenses = sum(item["amount"] for item in expenses.values())
    profit_before_tax = total_income - total_expenses

    # Tax
    dt_row = conn.execute(
        "SELECT COALESCE(SUM(CAST(CLOSINGBALANCE AS REAL)), 0) FROM mst_ledger WHERE LOWER(NAME) LIKE '%deferred tax%'"
    ).fetchone()
    dt_value = dt_row[0] if dt_row else 0.0

    # For deferred tax in P&L: change in DT during the year
    dt_ob_row = conn.execute(
        "SELECT COALESCE(SUM(CAST(OPENINGBALANCE AS REAL)), 0) FROM mst_ledger WHERE LOWER(NAME) LIKE '%deferred tax%'"
    ).fetchone()
    dt_ob = dt_ob_row[0] if dt_ob_row else 0.0
    dt_change = dt_value - dt_ob  # positive = increase in liability = tax expense

    # Current tax from Provision for Corporate Tax or similar
    ct_row = conn.execute(
        "SELECT CAST(CLOSINGBALANCE AS REAL) FROM mst_ledger WHERE LOWER(NAME) LIKE '%provision for%tax%'"
    ).fetchone()
    current_tax = abs(ct_row[0]) if ct_row and ct_row[0] else 0.0

    profit_after_tax = profit_before_tax - current_tax - dt_change

    return {
        "revenue": revenue,
        "other_income": other_income,
        "total_income": total_income,
        "expenses": expenses,
        "total_expenses": total_expenses,
        "profit_before_tax": profit_before_tax,
        "tax_current": current_tax,
        "tax_deferred": dt_change,
        "profit_after_tax": profit_after_tax,
    }


def extract_notes_data(conn, bs_data, pl_data):
    """Extract all Notes to Accounts data."""
    notes = OrderedDict()

    # Note 1: Significant Accounting Policies (template text)
    notes["1"] = {
        "title": "Significant Accounting Policies",
        "type": "text",
        "content": _get_accounting_policies_text(),
    }

    # Note 2: Share Capital
    share_cap_ledgers = _get_ledger_details(conn, ["Share Capital Account"],
                                             exclude_groups=["Reserves & Surplus"])
    # Get share info
    notes["2"] = {
        "title": "Share Capital",
        "type": "share_capital",
        "ledgers": [(l[0], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in share_cap_ledgers],
        "total_ob": sum(abs(l[2]) for l in share_cap_ledgers if l[2]),
        "total_cb": sum(abs(l[3]) for l in share_cap_ledgers if l[3]),
    }

    # Note 3: Reserves and Surplus
    rs_ledgers = _get_ledger_details(conn, ["Reserves & Surplus"])
    pl_row = conn.execute(
        "SELECT CAST(OPENINGBALANCE AS REAL), CAST(CLOSINGBALANCE AS REAL) FROM mst_ledger WHERE NAME = 'Profit & Loss A/c'"
    ).fetchone()
    pl_ob = pl_row[0] if pl_row and pl_row[0] else 0.0
    pl_cb = pl_row[1] if pl_row and pl_row[1] else 0.0

    notes["3"] = {
        "title": "Reserves and Surplus",
        "type": "reserves",
        "ledgers": [(l[0], l[2] or 0, l[3] or 0) for l in rs_ledgers],
        "pl_opening": pl_ob,
        "pl_closing": pl_cb,
        "pl_profit": pl_data["profit_after_tax"],
    }

    # Note 4: Long-term Borrowings
    lt_ledgers = _get_ledger_details(conn, ["Secured Loans", "Unsecured Loans"])
    notes["4"] = {
        "title": "Long-term Borrowings",
        "type": "ledger_list",
        "ledgers": [(l[0], l[1], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in lt_ledgers if l[3] and abs(l[3]) > 0.01],
        "total_cb": abs(bs_data["liabilities"]["non_current_liabilities"]["long_term_borrowings"]["amount"]),
    }

    # Note 5: Deferred Tax
    dt_ledgers = conn.execute(
        "SELECT NAME, CAST(OPENINGBALANCE AS REAL), CAST(CLOSINGBALANCE AS REAL) FROM mst_ledger WHERE LOWER(NAME) LIKE '%deferred tax%'"
    ).fetchall()
    notes["5"] = {
        "title": "Deferred Tax Liability / (Asset)",
        "type": "deferred_tax",
        "ledgers": dt_ledgers,
    }

    # Note 6: Short-term Borrowings
    stb_ledgers = _get_ledger_details(conn, ["Bank OD A/c"])
    notes["6"] = {
        "title": "Short-term Borrowings",
        "type": "ledger_list",
        "ledgers": [(l[0], l[1], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in stb_ledgers if l[3] and abs(l[3]) > 0.01],
        "total_cb": abs(bs_data["liabilities"]["current_liabilities"]["short_term_borrowings"]["amount"]),
    }

    # Note 7: Trade Payables
    tp_ledgers = _get_ledger_details(conn, ["Sundry Creditors"])
    tp_credit = [(l[0], l[1], l[2] or 0, l[3] or 0) for l in tp_ledgers if l[3] and l[3] > 0]
    tp_debit = [(l[0], l[1], l[2] or 0, l[3] or 0) for l in tp_ledgers if l[3] and l[3] < 0]
    notes["7"] = {
        "title": "Trade Payables",
        "type": "trade_payables",
        "credit_ledgers": tp_credit,  # actual payables (credit balance)
        "debit_ledgers": tp_debit,    # advances to suppliers
        "total_cb": bs_data["liabilities"]["current_liabilities"]["trade_payables"]["amount"],
    }

    # Note 8: Other Current Liabilities
    ocl_ledgers_tax = _get_ledger_details(conn, ["Duties & Taxes"])
    ocl_ledgers_salary = _get_ledger_details(conn, ["Salary Payable"])
    ocl_ledgers_direct = _get_direct_ledger_details(
        conn, ["Current Liabilities"],
        exclude_names=["deferred tax"]
    )
    notes["8"] = {
        "title": "Other Current Liabilities",
        "type": "other_current_liabilities",
        "tax_ledgers": [(l[0], l[1], l[2] or 0, l[3] or 0) for l in ocl_ledgers_tax if l[3] and abs(l[3]) > 0.01],
        "salary_ledgers": [(l[0], l[1], l[2] or 0, l[3] or 0) for l in ocl_ledgers_salary if l[3] and abs(l[3]) > 0.01],
        "other_ledgers": [(l[0], l[1], l[2] or 0, l[3] or 0) for l in ocl_ledgers_direct if l[3] and abs(l[3]) > 0.01],
        "total_cb": bs_data["liabilities"]["current_liabilities"]["other_current_liabilities"]["amount"],
    }

    # Note 9: Short-term Provisions
    prov_ledgers = _get_ledger_details(conn, ["Provisions"])
    notes["9"] = {
        "title": "Short-term Provisions",
        "type": "ledger_list",
        "ledgers": [(l[0], l[1], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in prov_ledgers],
        "total_cb": abs(bs_data["liabilities"]["current_liabilities"]["short_term_provisions"]["amount"]),
    }

    # Note 10: Property, Plant and Equipment
    fa_ledgers = _get_ledger_details(conn, ["Fixed Assets"])
    notes["10"] = {
        "title": "Property, Plant and Equipment",
        "type": "ppe",
        "ledgers": [(l[0], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in fa_ledgers if (l[2] and abs(l[2]) > 0.01) or (l[3] and abs(l[3]) > 0.01)],
    }

    # Note 11: Non-current Investments
    inv_ledgers = _get_ledger_details(conn, ["Investments"])
    notes["11"] = {
        "title": "Non-current Investments",
        "type": "ledger_list",
        "ledgers": [(l[0], l[1], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in inv_ledgers if l[3] and abs(l[3]) > 0.01],
        "total_cb": abs(bs_data["assets"]["non_current_assets"]["non_current_investments"]["amount"]),
    }

    # Note 12: Inventories
    _, _, stock_details = _get_stock_ledger_balances(conn)
    notes["12"] = {
        "title": "Inventories",
        "type": "inventories",
        "items": stock_details,
        "total_cb": abs(bs_data["assets"]["current_assets"]["inventories"]["amount"]),
    }

    # Note 13: Trade Receivables
    tr_ledgers = _get_ledger_details(conn, ["Sundry Debtors"])
    tr_debit = [(l[0], l[1], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in tr_ledgers if l[3] and l[3] < 0]
    tr_credit = [(l[0], l[1], l[2] or 0, l[3] or 0) for l in tr_ledgers if l[3] and l[3] > 0]
    notes["13"] = {
        "title": "Trade Receivables",
        "type": "trade_receivables",
        "debit_ledgers": tr_debit,  # actual receivables (debit balance in Tally = negative CB)
        "credit_ledgers": tr_credit,  # advances from customers
        "total_cb": bs_data["assets"]["current_assets"]["trade_receivables"]["amount"],
    }

    # Note 14: Cash and Cash Equivalents
    cash_ledgers = _get_ledger_details(conn, ["Cash-in-Hand"])
    bank_ledgers = _get_ledger_details(conn, ["Bank Accounts"])
    notes["14"] = {
        "title": "Cash and Cash Equivalents",
        "type": "cash",
        "cash_ledgers": [(l[0], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in cash_ledgers if l[3] and abs(l[3]) > 0.01],
        "bank_ledgers": [(l[0], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in bank_ledgers if l[3] and abs(l[3]) > 0.01],
        "total_cb": abs(bs_data["assets"]["current_assets"]["cash_equivalents"]["amount"]),
    }

    # Note 15: Short-term Loans and Advances
    la_ledgers = _get_ledger_details(conn, ["Loans & Advances (Asset)", "Loan to Employees", "Deposits (Asset)"])
    notes["15"] = {
        "title": "Short-term Loans and Advances",
        "type": "ledger_list",
        "ledgers": [(l[0], l[1], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in la_ledgers if l[3] and abs(l[3]) > 0.01],
        "total_cb": abs(bs_data["assets"]["current_assets"]["short_term_loans_advances"]["amount"]),
    }

    # Note 16: Other Current Assets
    oca_ledgers = _get_direct_ledger_details(conn, ["Current Assets"],
                                              exclude_names=["deferred tax"])
    notes["16"] = {
        "title": "Other Current Assets",
        "type": "ledger_list",
        "ledgers": [(l[0], l[1], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in oca_ledgers if l[3] and abs(l[3]) > 0.01],
        "total_cb": abs(bs_data["assets"]["current_assets"]["other_current_assets"]["amount"]),
    }

    # Note 17: Revenue from Operations
    rev_ledgers = _get_ledger_details(conn, ["Sales Accounts"])
    notes["17"] = {
        "title": "Revenue from Operations",
        "type": "ledger_list_abs",
        "ledgers": [(l[0], l[1], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in rev_ledgers if l[3] and abs(l[3]) > 0.01],
        "total_cb": pl_data["revenue"],
    }

    # Note 18: Other Income
    oi_ledgers = _get_ledger_details(conn, ["Direct Incomes", "Indirect Incomes"])
    notes["18"] = {
        "title": "Other Income",
        "type": "ledger_list_abs",
        "ledgers": [(l[0], l[1], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in oi_ledgers if l[3] and abs(l[3]) > 0.01],
        "total_cb": pl_data["other_income"],
    }

    # Note 19: Cost of Materials Consumed
    purch_ledgers = _get_ledger_details(conn, ["Purchase Accounts"])
    notes["19"] = {
        "title": "Cost of Materials Consumed",
        "type": "ledger_list_abs",
        "ledgers": [(l[0], l[1], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in purch_ledgers if l[3] and abs(l[3]) > 0.01],
        "total_cb": pl_data["expenses"]["cost_of_materials"]["amount"],
    }

    # Note 20: Changes in Inventories
    opening, closing, stock_det = _get_stock_ledger_balances(conn)
    notes["20"] = {
        "title": "Changes in Inventories of Finished Goods",
        "type": "inventory_change",
        "items": stock_det,
        "opening": opening,
        "closing": closing,
        "change": opening - closing,
    }

    # Note 21: Employee Benefit Expense
    emp_ledgers = _get_ledger_details(conn, ["Salary Expenses"])
    notes["21"] = {
        "title": "Employee Benefit Expense",
        "type": "ledger_list_abs",
        "ledgers": [(l[0], l[1], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in emp_ledgers if l[3] and abs(l[3]) > 0.01],
        "total_cb": pl_data["expenses"]["employee_benefit"]["amount"],
    }

    # Note 22: Finance Costs
    fc_names = ["interest", "bank charges", "bank od interest"]
    fc_conditions = " OR ".join(["LOWER(NAME) LIKE ?" for _ in fc_names])
    fc_params = [f"%{n}%" for n in fc_names]
    fc_ledgers = conn.execute(
        f"SELECT NAME, PARENT, CAST(OPENINGBALANCE AS REAL), CAST(CLOSINGBALANCE AS REAL) FROM mst_ledger WHERE ({fc_conditions}) AND CAST(CLOSINGBALANCE AS REAL) != 0",
        fc_params
    ).fetchall()
    notes["22"] = {
        "title": "Finance Costs",
        "type": "ledger_list_abs",
        "ledgers": [(l[0], l[1], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in fc_ledgers if l[3] and abs(l[3]) > 0.01],
        "total_cb": pl_data["expenses"]["finance_costs"]["amount"],
    }

    # Note 23: Other Expenses
    oe_ledgers = _get_ledger_details(
        conn, ["Direct Expenses", "Indirect Expenses"],
        exclude_groups=["Salary Expenses"],
        exclude_names=["interest", "bank charges", "bank od interest",
                       "depreciation", "amortization", "amortisation"]
    )
    notes["23"] = {
        "title": "Other Expenses",
        "type": "other_expenses",
        "ledgers": [(l[0], l[1], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in oe_ledgers if l[3] and abs(l[3]) > 0.01],
        "total_cb": pl_data["expenses"]["other_expenses"]["amount"],
    }

    # Note 24: Related Party Disclosures
    # Directors from Unsecured Loans
    ul_ledgers = _get_ledger_details(conn, ["Unsecured Loans"])
    notes["24"] = {
        "title": "Related Party Disclosures",
        "type": "related_party",
        "loans": [(l[0], abs(l[2]) if l[2] else 0, abs(l[3]) if l[3] else 0) for l in ul_ledgers if l[3] and abs(l[3]) > 0.01],
    }

    # Note 25: Segment Information
    notes["25"] = {
        "title": "Segment Information",
        "type": "text",
        "content": "The Company has one reportable segment during the financial year, which is revenue from selling of products.",
    }

    # Note 26: Figures rounded off
    notes["26"] = {
        "title": "Additional Information",
        "type": "text",
        "content": "Previous year figures have been regrouped/reclassified wherever necessary to make them comparable with the current year figures.",
    }

    return notes


def _get_accounting_policies_text():
    """Return standard Significant Accounting Policies text."""
    return """1. BASIS OF PREPARATION
The financial statements have been prepared in accordance with the Generally Accepted Accounting Principles in India under the historical cost convention on accrual basis. These financial statements have been prepared to comply in all material aspects with the Accounting Standards notified under Section 133 of the Companies Act, 2013.

2. USE OF ESTIMATES
The preparation of financial statements requires estimates and assumptions that affect the reported amounts of assets, liabilities, revenue and expenses during the reporting period. Although these estimates are based on management's best knowledge, actual results could differ from these estimates.

3. REVENUE RECOGNITION
Revenue is recognized to the extent that it is probable that the economic benefits will flow to the Company and the revenue can be reliably measured. Revenue from sale of goods is recognized when the significant risks and rewards of ownership have been transferred to the buyer.

4. PROPERTY, PLANT AND EQUIPMENT
Property, Plant and Equipment are stated at cost of acquisition less accumulated depreciation. Depreciation is provided on Written Down Value (WDV) method at the rates prescribed under Schedule II of the Companies Act, 2013.

5. INVENTORIES
Inventories are valued at cost or net realizable value, whichever is lower. Cost is determined on FIFO basis.

6. EMPLOYEE BENEFITS
Short-term employee benefits are recognized as expense at undiscounted amount in the Statement of Profit and Loss of the year in which the related service is rendered.

7. TAXATION
Tax expense comprises current tax and deferred tax. Current tax is measured at the amount expected to be paid to the tax authorities. Deferred tax is recognized on timing differences between the taxable income and accounting income using the tax rates that have been enacted or substantively enacted.

8. PROVISIONS AND CONTINGENT LIABILITIES
Provisions are recognized when the Company has a present obligation as a result of past events and it is probable that an outflow of resources will be required to settle the obligation. Contingent liabilities are disclosed by way of notes to accounts.

9. CASH AND CASH EQUIVALENTS
Cash and cash equivalents comprise cash at bank and in hand and short-term deposits with an original maturity of three months or less."""


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL FORMATTING
# ═══════════════════════════════════════════════════════════════════════════════

FONT_COMPANY = Font(name="Times New Roman", size=14, bold=True)
FONT_SUBTITLE = Font(name="Times New Roman", size=11, bold=True)
FONT_NORMAL_BOLD = Font(name="Times New Roman", size=10, bold=True)
FONT_NORMAL = Font(name="Times New Roman", size=10)
FONT_SECTION = Font(name="Times New Roman", size=10, bold=True)
FONT_HEADER = Font(name="Times New Roman", size=10, bold=True)
FONT_NOTE = Font(name="Times New Roman", size=9)
FONT_SIGNATORY = Font(name="Times New Roman", size=9)
FONT_SIGNATORY_BOLD = Font(name="Times New Roman", size=9, bold=True)
FONT_POLICY = Font(name="Times New Roman", size=9)
FONT_POLICY_BOLD = Font(name="Times New Roman", size=9, bold=True)
FONT_SMALL = Font(name="Times New Roman", size=8)
FONT_NOTE_TITLE = Font(name="Times New Roman", size=11, bold=True)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_LEFT_INDENT1 = Alignment(horizontal="left", vertical="center", indent=2)
ALIGN_LEFT_INDENT2 = Alignment(horizontal="left", vertical="center", indent=4)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_NOTE_CENTER = Alignment(horizontal="center", vertical="center")

THIN_BORDER = Side(style="thin")
HAIR_BORDER = Side(style="hair")
BOTTOM_BORDER = Border(bottom=THIN_BORDER)
TOP_BOTTOM_BORDER = Border(top=THIN_BORDER, bottom=THIN_BORDER)
DOUBLE_BOTTOM_BORDER = Border(bottom=Side(style="double"))
HEADER_BORDER = Border(bottom=Side(style="medium"))

HEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

AMT_FMT = '#,##0.00;(#,##0.00);"-"'
AMT_FMT_INT = '#,##0;(#,##0);"-"'


def _write_company_header(ws, company_info, row_start, title, max_col=4):
    """Write company header block."""
    row = row_start

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=company_info.get("name", ""))
    cell.font = FONT_COMPANY
    cell.alignment = ALIGN_CENTER
    row += 1

    cin = company_info.get("cin", "")
    if cin:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        cell = ws.cell(row=row, column=1, value=f"CIN : {cin}")
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_CENTER
        row += 1

    address = company_info.get("address", "")
    if address:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
        cell = ws.cell(row=row, column=1, value=f"Regd Office: {address}")
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_CENTER
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = FONT_SUBTITLE
    cell.alignment = ALIGN_CENTER
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=f"Amount (in {company_info.get('denomination', 'INR')})")
    cell.font = Font(name="Times New Roman", size=9, italic=True)
    cell.alignment = Alignment(horizontal="right")
    row += 1

    return row


def _write_column_headers(ws, row, headers):
    """Write column headers."""
    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER if col_idx > 1 else ALIGN_LEFT
        cell.border = HEADER_BORDER
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    return row + 1


def _write_row(ws, row, label, note="", amount=None, prev_amount=None,
               font=None, alignment=None, is_total=False, num_fmt=AMT_FMT):
    """Write a single line item row."""
    f = font or FONT_NORMAL
    a = alignment or ALIGN_LEFT_INDENT2

    cell = ws.cell(row=row, column=1, value=label)
    cell.font = f
    cell.alignment = a

    if note:
        cell = ws.cell(row=row, column=2, value=note)
        cell.font = FONT_NOTE
        cell.alignment = ALIGN_NOTE_CENTER

    if amount is not None:
        cell = ws.cell(row=row, column=3, value=amount if amount else None)
        cell.font = f
        cell.alignment = ALIGN_RIGHT
        cell.number_format = num_fmt

    if prev_amount is not None:
        cell = ws.cell(row=row, column=4, value=prev_amount if prev_amount else None)
        cell.font = f
        cell.alignment = ALIGN_RIGHT
        cell.number_format = num_fmt

    if is_total:
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = TOP_BOTTOM_BORDER
            ws.cell(row=row, column=col).font = FONT_NORMAL_BOLD

    return row + 1


def _write_signatory_block(ws, row, company_info, max_col=4):
    """Write signatory block."""
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    cell = ws.cell(row=row, column=1, value="As per our Report of even date attached")
    cell.font = FONT_SIGNATORY_BOLD
    row += 2

    auditor_name = company_info.get("auditor_name", "")
    if auditor_name:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1, value=f"For {auditor_name}").font = FONT_SIGNATORY_BOLD
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1, value="Chartered Accountants").font = FONT_SIGNATORY
        row += 1
        frn = company_info.get("auditor_frn", "")
        if frn:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.cell(row=row, column=1, value=f"FRN: {frn}").font = FONT_SIGNATORY
            row += 1

    # Directors on right side
    dir1 = company_info.get("director1_name", "")
    dir1_din = company_info.get("director1_din", "")
    dir2 = company_info.get("director2_name", "")
    dir2_din = company_info.get("director2_din", "")

    dir_row = row
    if dir1:
        c = ws.cell(row=dir_row, column=max_col, value=dir1)
        c.font = FONT_SIGNATORY_BOLD
        c.alignment = ALIGN_RIGHT
        dir_row += 1
        c = ws.cell(row=dir_row, column=max_col, value="Director")
        c.font = FONT_SIGNATORY
        c.alignment = ALIGN_RIGHT
        dir_row += 1
        if dir1_din:
            c = ws.cell(row=dir_row, column=max_col, value=f"DIN: {dir1_din}")
            c.font = FONT_SIGNATORY
            c.alignment = ALIGN_RIGHT
            dir_row += 1

    row = max(row + 2, dir_row)

    # Member info (left) and Director 2 (right)
    member = company_info.get("auditor_member", "")
    if member:
        ws.cell(row=row, column=1, value=f"Membership No: {member}").font = FONT_SIGNATORY

    if dir2:
        c = ws.cell(row=row, column=max_col, value=dir2)
        c.font = FONT_SIGNATORY_BOLD
        c.alignment = ALIGN_RIGHT
        row += 1
        c = ws.cell(row=row, column=max_col, value="Director")
        c.font = FONT_SIGNATORY
        c.alignment = ALIGN_RIGHT
        row += 1
        if dir2_din:
            c = ws.cell(row=row, column=max_col, value=f"DIN: {dir2_din}")
            c.font = FONT_SIGNATORY
            c.alignment = ALIGN_RIGHT
            row += 1

    row += 1
    ws.cell(row=row, column=1, value="Place:").font = FONT_SIGNATORY
    ws.cell(row=row, column=max_col, value="Date:").font = FONT_SIGNATORY

    return row


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL SHEET BUILDERS
# ═══════════════════════════════════════════════════════════════════════════════

def _build_balance_sheet(wb, bs_data, company_info):
    """Build Balance Sheet worksheet."""
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    year_end = company_info.get("year_end", "31.03.2026")
    prev_year = company_info.get("prev_year_end", "31.03.2025")

    row = _write_company_header(ws, company_info, 1, f"Balance Sheet as at {year_end}")

    headers = [
        ("Particulars", 55),
        ("Note", 8),
        (f"As at {year_end}", 22),
        (f"As at {prev_year}", 22),
    ]
    row = _write_column_headers(ws, row, headers)

    # I. EQUITY AND LIABILITIES
    row += 1
    ws.cell(row=row, column=1, value="I. EQUITY AND LIABILITIES").font = FONT_SECTION
    row += 1

    liabilities = bs_data["liabilities"]

    # (1) Shareholder's Funds
    ws.cell(row=row, column=1, value="(1) Shareholder's Funds").font = FONT_NORMAL_BOLD
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT_INDENT1
    row += 1

    sf = liabilities["shareholders_funds"]
    for idx, (key, item) in enumerate(sf.items()):
        letter = chr(ord('a') + idx)
        row = _write_row(ws, row, f"({letter}) {item['label']}", item["note"], item["amount"])

    # (2) Share application money
    ws.cell(row=row, column=1,
            value="(2) Share application money pending allotment").font = FONT_NORMAL_BOLD
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT_INDENT1
    row += 1

    # (3) Non-Current Liabilities
    ws.cell(row=row, column=1, value="(3) Non-Current Liabilities").font = FONT_NORMAL_BOLD
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT_INDENT1
    row += 1

    ncl = liabilities["non_current_liabilities"]
    for idx, (key, item) in enumerate(ncl.items()):
        letter = chr(ord('a') + idx)
        row = _write_row(ws, row, f"({letter}) {item['label']}", item["note"], item["amount"])

    # (4) Current Liabilities
    ws.cell(row=row, column=1, value="(4) Current Liabilities").font = FONT_NORMAL_BOLD
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT_INDENT1
    row += 1

    cl = liabilities["current_liabilities"]
    for idx, (key, item) in enumerate(cl.items()):
        letter = chr(ord('a') + idx)
        row = _write_row(ws, row, f"({letter}) {item['label']}", item["note"], item["amount"])

    # Total
    row = _write_row(ws, row, "Total", "", bs_data["total_liabilities"],
                     font=FONT_NORMAL_BOLD, alignment=ALIGN_LEFT, is_total=True)

    # II. ASSETS
    row += 1
    ws.cell(row=row, column=1, value="II. ASSETS").font = FONT_SECTION
    row += 1

    assets = bs_data["assets"]

    # (1) Non-current assets
    ws.cell(row=row, column=1, value="(1) Non-current assets").font = FONT_NORMAL_BOLD
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT_INDENT1
    row += 1

    nca = assets["non_current_assets"]
    for idx, (key, item) in enumerate(nca.items()):
        letter = chr(ord('a') + idx)
        row = _write_row(ws, row, f"({letter}) {item['label']}", item["note"], item["amount"])

    # (2) Current assets
    ws.cell(row=row, column=1, value="(2) Current assets").font = FONT_NORMAL_BOLD
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT_INDENT1
    row += 1

    ca = assets["current_assets"]
    for idx, (key, item) in enumerate(ca.items()):
        letter = chr(ord('a') + idx)
        row = _write_row(ws, row, f"({letter}) {item['label']}", item["note"], item["amount"])

    # Notes reference
    row += 1
    ws.cell(row=row, column=1,
            value="Notes to accounts & Significant accounting policies").font = FONT_NORMAL
    row += 1

    # Total
    row = _write_row(ws, row, "Total", "", bs_data["total_assets"],
                     font=FONT_NORMAL_BOLD, alignment=ALIGN_LEFT, is_total=True)

    _write_signatory_block(ws, row, company_info)


def _build_pl_statement(wb, pl_data, company_info):
    """Build P&L Statement worksheet."""
    ws = wb.create_sheet("Profit and Loss")
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    year_end = company_info.get("year_end", "31.03.2026")
    prev_year = company_info.get("prev_year_end", "31.03.2025")

    row = _write_company_header(ws, company_info, 1,
                                 f"Statement of Profit and Loss for the Year Ended {year_end}")

    headers = [
        ("Particulars", 60),
        ("Note", 8),
        (f"Year ended {year_end}", 22),
        (f"Year ended {prev_year}", 22),
    ]
    row = _write_column_headers(ws, row, headers)
    row += 1

    # Revenue
    row = _write_row(ws, row, "I. Revenue from operations", "17", pl_data["revenue"])
    row = _write_row(ws, row, "II. Other Income", "18", pl_data["other_income"])
    row = _write_row(ws, row, "III. Total Income (I + II)", "",
                     pl_data["total_income"], font=FONT_NORMAL_BOLD, alignment=ALIGN_LEFT,
                     is_total=True)

    row += 1
    ws.cell(row=row, column=1, value="IV. Expenses:").font = FONT_SECTION
    row += 1

    expense_order = ["cost_of_materials", "purchase_stock_in_trade", "changes_in_inventories",
                     "employee_benefit", "finance_costs", "depreciation", "other_expenses"]
    for key in expense_order:
        if key in pl_data["expenses"]:
            item = pl_data["expenses"][key]
            row = _write_row(ws, row, item["label"], item.get("note", ""), item["amount"],
                             alignment=ALIGN_LEFT_INDENT1)

    row = _write_row(ws, row, "Total Expenses", "", pl_data["total_expenses"],
                     font=FONT_NORMAL_BOLD, alignment=ALIGN_LEFT, is_total=True)

    row += 1
    row = _write_row(ws, row,
                     "V. Profit/(Loss) before exceptional items and tax (III - IV)", "",
                     pl_data["profit_before_tax"], font=FONT_NORMAL_BOLD, alignment=ALIGN_LEFT)

    row = _write_row(ws, row, "VI. Exceptional Items", "", 0)
    row = _write_row(ws, row, "VII. Profit/(Loss) before tax (V - VI)", "",
                     pl_data["profit_before_tax"], font=FONT_NORMAL_BOLD, alignment=ALIGN_LEFT)

    row += 1
    ws.cell(row=row, column=1, value="VIII. Tax expense:").font = FONT_NORMAL_BOLD
    row += 1
    row = _write_row(ws, row, "(1) Current tax", "", pl_data["tax_current"],
                     alignment=ALIGN_LEFT_INDENT1)
    row = _write_row(ws, row, "(2) Deferred tax", "5", pl_data["tax_deferred"],
                     alignment=ALIGN_LEFT_INDENT1)

    row += 1
    row = _write_row(ws, row,
                     "IX. Profit/(Loss) for the period", "",
                     pl_data["profit_after_tax"], font=FONT_NORMAL_BOLD, alignment=ALIGN_LEFT,
                     is_total=True)

    row += 1
    ws.cell(row=row, column=1, value="X. Earnings per equity share:").font = FONT_NORMAL_BOLD
    row += 1
    row = _write_row(ws, row, "(1) Basic", "", None, alignment=ALIGN_LEFT_INDENT1)
    row = _write_row(ws, row, "(2) Diluted", "", None, alignment=ALIGN_LEFT_INDENT1)

    row += 1
    ws.cell(row=row, column=1,
            value="Notes to accounts & Significant accounting policies").font = FONT_NORMAL
    row += 1

    _write_signatory_block(ws, row, company_info)


def _build_notes_sheets(wb, notes_data, company_info):
    """Build Notes to Accounts worksheets."""

    year_end = company_info.get("year_end", "31.03.2026")
    prev_year = company_info.get("prev_year_end", "31.03.2025")

    # ── Note 1: Accounting Policies ──
    ws = wb.create_sheet("Note 1 - Accounting Policies")
    ws.column_dimensions['A'].width = 90
    row = 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
    ws.cell(row=row, column=1, value=company_info.get("name", "")).font = FONT_COMPANY
    ws.cell(row=row, column=1).alignment = ALIGN_CENTER
    row += 2

    ws.cell(row=row, column=1, value="Note 1: Significant Accounting Policies").font = FONT_NOTE_TITLE
    row += 2

    policies = notes_data["1"]["content"]
    for line in policies.split("\n"):
        line = line.strip()
        if not line:
            row += 1
            continue
        if line and line[0].isdigit() and "." in line[:3]:
            ws.cell(row=row, column=1, value=line).font = FONT_POLICY_BOLD
        else:
            ws.cell(row=row, column=1, value=line).font = FONT_POLICY
        ws.cell(row=row, column=1).alignment = ALIGN_LEFT_WRAP
        ws.row_dimensions[row].height = 30
        row += 1

    # ── Notes 2-16 (Balance Sheet notes) ──
    ws = wb.create_sheet("Notes 2-16 (BS)")
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    row = 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value=company_info.get("name", "")).font = FONT_COMPANY
    ws.cell(row=row, column=1).alignment = ALIGN_CENTER
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1,
            value=f"Notes forming part of Financial Statements as on {year_end}").font = FONT_SUBTITLE
    ws.cell(row=row, column=1).alignment = ALIGN_CENTER
    row += 2

    for note_num_str in ["2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12", "13", "14", "15", "16"]:
        note = notes_data.get(note_num_str)
        if not note:
            continue

        # Note header
        ws.cell(row=row, column=1, value="NOTE").font = FONT_NORMAL_BOLD
        ws.cell(row=row, column=2, value=f"{note_num_str}. {note['title']}").font = FONT_NORMAL_BOLD
        row += 1

        # Sub-headers
        ws.cell(row=row, column=2, value="Particulars").font = FONT_HEADER
        ws.cell(row=row, column=2).border = HEADER_BORDER
        ws.cell(row=row, column=3, value=f"As at {year_end}").font = FONT_HEADER
        ws.cell(row=row, column=3).alignment = ALIGN_CENTER
        ws.cell(row=row, column=3).border = HEADER_BORDER
        ws.cell(row=row, column=4, value=f"As at {prev_year}").font = FONT_HEADER
        ws.cell(row=row, column=4).alignment = ALIGN_CENTER
        ws.cell(row=row, column=4).border = HEADER_BORDER
        row += 1

        ntype = note.get("type", "")

        if ntype == "share_capital":
            for name, ob, cb in note.get("ledgers", []):
                ws.cell(row=row, column=2, value=name).font = FONT_NORMAL
                ws.cell(row=row, column=3, value=cb).font = FONT_NORMAL
                ws.cell(row=row, column=3).number_format = AMT_FMT
                ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
                ws.cell(row=row, column=4, value=ob).font = FONT_NORMAL
                ws.cell(row=row, column=4).number_format = AMT_FMT
                ws.cell(row=row, column=4).alignment = ALIGN_RIGHT
                row += 1
            ws.cell(row=row, column=2, value="Total").font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3, value=note["total_cb"]).font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            ws.cell(row=row, column=3).border = TOP_BOTTOM_BORDER
            row += 1

        elif ntype == "reserves":
            ws.cell(row=row, column=2, value="A. Surplus in Statement of P&L").font = FONT_NORMAL_BOLD
            row += 1
            ws.cell(row=row, column=2, value="Opening balance").font = FONT_NORMAL
            ws.cell(row=row, column=3, value=note["pl_opening"]).font = FONT_NORMAL
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            row += 1
            ws.cell(row=row, column=2, value="Profit/(Loss) for the year").font = FONT_NORMAL
            ws.cell(row=row, column=3, value=note["pl_profit"]).font = FONT_NORMAL
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            row += 1
            ws.cell(row=row, column=2, value="Closing Balance").font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3, value=note["pl_closing"]).font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            ws.cell(row=row, column=3).border = TOP_BOTTOM_BORDER
            row += 1

            # Other reserves
            for name, ob, cb in note.get("ledgers", []):
                if abs(cb) > 0.01 or abs(ob) > 0.01:
                    ws.cell(row=row, column=2, value=name).font = FONT_NORMAL
                    ws.cell(row=row, column=3, value=cb).font = FONT_NORMAL
                    ws.cell(row=row, column=3).number_format = AMT_FMT
                    ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
                    ws.cell(row=row, column=4, value=ob).font = FONT_NORMAL
                    ws.cell(row=row, column=4).number_format = AMT_FMT
                    ws.cell(row=row, column=4).alignment = ALIGN_RIGHT
                    row += 1

        elif ntype == "deferred_tax":
            for l in note.get("ledgers", []):
                ws.cell(row=row, column=2, value=l[0]).font = FONT_NORMAL
                cb = abs(l[2]) if l[2] else 0
                ob = abs(l[1]) if l[1] else 0
                ws.cell(row=row, column=3, value=cb).font = FONT_NORMAL
                ws.cell(row=row, column=3).number_format = AMT_FMT
                ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
                ws.cell(row=row, column=4, value=ob).font = FONT_NORMAL
                ws.cell(row=row, column=4).number_format = AMT_FMT
                ws.cell(row=row, column=4).alignment = ALIGN_RIGHT
                row += 1

        elif ntype == "ppe":
            # PPE schedule
            ws2 = ws  # staying on same sheet
            row -= 2  # go back to overwrite the sub-headers for PPE
            ws2.cell(row=row, column=2, value="Particulars").font = FONT_HEADER
            ws2.cell(row=row, column=2).border = HEADER_BORDER
            ws2.cell(row=row, column=3, value="Opening Balance").font = FONT_HEADER
            ws2.cell(row=row, column=3).alignment = ALIGN_CENTER
            ws2.cell(row=row, column=3).border = HEADER_BORDER
            ws2.cell(row=row, column=4, value="Closing Balance").font = FONT_HEADER
            ws2.cell(row=row, column=4).alignment = ALIGN_CENTER
            ws2.cell(row=row, column=4).border = HEADER_BORDER
            row += 1

            total_ob = 0
            total_cb = 0
            for name, ob, cb in note.get("ledgers", []):
                ws2.cell(row=row, column=2, value=name).font = FONT_NORMAL
                ws2.cell(row=row, column=3, value=ob).font = FONT_NORMAL
                ws2.cell(row=row, column=3).number_format = AMT_FMT
                ws2.cell(row=row, column=3).alignment = ALIGN_RIGHT
                ws2.cell(row=row, column=4, value=cb).font = FONT_NORMAL
                ws2.cell(row=row, column=4).number_format = AMT_FMT
                ws2.cell(row=row, column=4).alignment = ALIGN_RIGHT
                total_ob += ob
                total_cb += cb
                row += 1
            ws2.cell(row=row, column=2, value="Total").font = FONT_NORMAL_BOLD
            ws2.cell(row=row, column=3, value=total_ob).font = FONT_NORMAL_BOLD
            ws2.cell(row=row, column=3).number_format = AMT_FMT
            ws2.cell(row=row, column=3).alignment = ALIGN_RIGHT
            ws2.cell(row=row, column=3).border = TOP_BOTTOM_BORDER
            ws2.cell(row=row, column=4, value=total_cb).font = FONT_NORMAL_BOLD
            ws2.cell(row=row, column=4).number_format = AMT_FMT
            ws2.cell(row=row, column=4).alignment = ALIGN_RIGHT
            ws2.cell(row=row, column=4).border = TOP_BOTTOM_BORDER
            row += 1

        elif ntype == "trade_payables":
            # MSME vs Others (simplified - all as others since no MSME flag in Tally)
            ws.cell(row=row, column=2, value="Micro and Small Enterprises").font = FONT_NORMAL
            ws.cell(row=row, column=3, value=0).font = FONT_NORMAL
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            row += 1
            ws.cell(row=row, column=2, value="Others").font = FONT_NORMAL
            ws.cell(row=row, column=3, value=note["total_cb"]).font = FONT_NORMAL
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            row += 1
            ws.cell(row=row, column=2, value="Total").font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3, value=note["total_cb"]).font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            ws.cell(row=row, column=3).border = TOP_BOTTOM_BORDER
            row += 1

        elif ntype == "trade_receivables":
            total = note.get("total_cb", 0)
            ws.cell(row=row, column=2, value="Unsecured, considered good").font = FONT_NORMAL
            ws.cell(row=row, column=3, value=total).font = FONT_NORMAL
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            row += 1
            ws.cell(row=row, column=2, value="Unsecured, considered doubtful").font = FONT_NORMAL
            ws.cell(row=row, column=3, value=0).font = FONT_NORMAL
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            row += 1
            ws.cell(row=row, column=2, value="Total").font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3, value=total).font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            ws.cell(row=row, column=3).border = TOP_BOTTOM_BORDER
            row += 1

        elif ntype == "other_current_liabilities":
            all_items = note.get("tax_ledgers", []) + note.get("salary_ledgers", []) + note.get("other_ledgers", [])
            total = 0
            for name, parent, ob, cb in all_items:
                ws.cell(row=row, column=2, value=name).font = FONT_NORMAL
                ws.cell(row=row, column=3, value=cb).font = FONT_NORMAL
                ws.cell(row=row, column=3).number_format = AMT_FMT
                ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
                total += cb
                row += 1
            ws.cell(row=row, column=2, value="Total").font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3, value=note["total_cb"]).font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            ws.cell(row=row, column=3).border = TOP_BOTTOM_BORDER
            row += 1

        elif ntype == "inventories":
            for name, ob, cb in note.get("items", []):
                ws.cell(row=row, column=2, value=name).font = FONT_NORMAL
                ws.cell(row=row, column=3, value=cb).font = FONT_NORMAL
                ws.cell(row=row, column=3).number_format = AMT_FMT
                ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
                ws.cell(row=row, column=4, value=ob).font = FONT_NORMAL
                ws.cell(row=row, column=4).number_format = AMT_FMT
                ws.cell(row=row, column=4).alignment = ALIGN_RIGHT
                row += 1
            ws.cell(row=row, column=2, value="Total").font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3, value=note["total_cb"]).font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            ws.cell(row=row, column=3).border = TOP_BOTTOM_BORDER
            row += 1

        elif ntype == "cash":
            ws.cell(row=row, column=2, value="A. Cash in hand").font = FONT_NORMAL_BOLD
            row += 1
            cash_total = 0
            for name, ob, cb in note.get("cash_ledgers", []):
                ws.cell(row=row, column=2, value=name).font = FONT_NORMAL
                ws.cell(row=row, column=3, value=cb).font = FONT_NORMAL
                ws.cell(row=row, column=3).number_format = AMT_FMT
                ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
                ws.cell(row=row, column=4, value=ob).font = FONT_NORMAL
                ws.cell(row=row, column=4).number_format = AMT_FMT
                ws.cell(row=row, column=4).alignment = ALIGN_RIGHT
                cash_total += cb
                row += 1

            ws.cell(row=row, column=2, value="B. Balances with banks").font = FONT_NORMAL_BOLD
            row += 1
            bank_total = 0
            for name, ob, cb in note.get("bank_ledgers", []):
                ws.cell(row=row, column=2, value=name).font = FONT_NORMAL
                ws.cell(row=row, column=3, value=cb).font = FONT_NORMAL
                ws.cell(row=row, column=3).number_format = AMT_FMT
                ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
                ws.cell(row=row, column=4, value=ob).font = FONT_NORMAL
                ws.cell(row=row, column=4).number_format = AMT_FMT
                ws.cell(row=row, column=4).alignment = ALIGN_RIGHT
                bank_total += cb
                row += 1

            ws.cell(row=row, column=2, value="Total").font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3, value=note["total_cb"]).font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            ws.cell(row=row, column=3).border = TOP_BOTTOM_BORDER
            row += 1

        elif ntype in ("ledger_list", "ledger_list_abs"):
            for item in note.get("ledgers", []):
                name = item[0]
                cb = item[3] if len(item) > 3 else item[2]
                ob = item[2] if len(item) > 3 else (item[1] if len(item) > 2 else 0)
                ws.cell(row=row, column=2, value=name).font = FONT_NORMAL
                ws.cell(row=row, column=3, value=cb).font = FONT_NORMAL
                ws.cell(row=row, column=3).number_format = AMT_FMT
                ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
                row += 1
            ws.cell(row=row, column=2, value="Total").font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3, value=note.get("total_cb", 0)).font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            ws.cell(row=row, column=3).border = TOP_BOTTOM_BORDER
            row += 1

        elif ntype == "text":
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            ws.cell(row=row, column=2, value=note.get("content", "")).font = FONT_NORMAL
            ws.cell(row=row, column=2).alignment = ALIGN_LEFT_WRAP
            row += 1

        row += 2  # spacing between notes

    # ── Notes 17-23 (P&L notes) ──
    ws = wb.create_sheet("Notes 17-23 (P&L)")
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    row = 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value=company_info.get("name", "")).font = FONT_COMPANY
    ws.cell(row=row, column=1).alignment = ALIGN_CENTER
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1,
            value=f"Notes forming part of Financial Statements for the year ended {year_end}").font = FONT_SUBTITLE
    ws.cell(row=row, column=1).alignment = ALIGN_CENTER
    row += 2

    for note_num_str in ["17", "18", "19", "20", "21", "22", "23"]:
        note = notes_data.get(note_num_str)
        if not note:
            continue

        ws.cell(row=row, column=1, value="NOTE").font = FONT_NORMAL_BOLD
        ws.cell(row=row, column=2, value=f"{note_num_str}. {note['title']}").font = FONT_NORMAL_BOLD
        row += 1

        ws.cell(row=row, column=2, value="Particulars").font = FONT_HEADER
        ws.cell(row=row, column=2).border = HEADER_BORDER
        ws.cell(row=row, column=3, value=f"Year ended {year_end}").font = FONT_HEADER
        ws.cell(row=row, column=3).alignment = ALIGN_CENTER
        ws.cell(row=row, column=3).border = HEADER_BORDER
        ws.cell(row=row, column=4, value=f"Year ended {prev_year}").font = FONT_HEADER
        ws.cell(row=row, column=4).alignment = ALIGN_CENTER
        ws.cell(row=row, column=4).border = HEADER_BORDER
        row += 1

        ntype = note.get("type", "")

        if ntype == "inventory_change":
            ws.cell(row=row, column=2, value="Inventory at the end of the year").font = FONT_NORMAL_BOLD
            row += 1
            for name, ob, cb in note.get("items", []):
                ws.cell(row=row, column=2, value=f"  {name}").font = FONT_NORMAL
                ws.cell(row=row, column=3, value=cb).font = FONT_NORMAL
                ws.cell(row=row, column=3).number_format = AMT_FMT
                ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
                row += 1
            row += 1
            ws.cell(row=row, column=2, value="Inventory at the beginning of the year").font = FONT_NORMAL_BOLD
            row += 1
            for name, ob, cb in note.get("items", []):
                ws.cell(row=row, column=2, value=f"  {name}").font = FONT_NORMAL
                ws.cell(row=row, column=3, value=ob).font = FONT_NORMAL
                ws.cell(row=row, column=3).number_format = AMT_FMT
                ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
                row += 1
            row += 1
            ws.cell(row=row, column=2, value="(Increase)/Decrease in inventories").font = FONT_NORMAL_BOLD
            row += 1
            ws.cell(row=row, column=2, value="Total").font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3, value=note["change"]).font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            ws.cell(row=row, column=3).border = TOP_BOTTOM_BORDER
            row += 1

        elif ntype == "other_expenses":
            # Group by parent to organize
            ledgers = note.get("ledgers", [])
            total = 0
            for name, parent, ob, cb in sorted(ledgers, key=lambda x: (-x[3], x[0])):
                ws.cell(row=row, column=2, value=name).font = FONT_NORMAL
                ws.cell(row=row, column=3, value=cb).font = FONT_NORMAL
                ws.cell(row=row, column=3).number_format = AMT_FMT
                ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
                total += cb
                row += 1
            ws.cell(row=row, column=2, value="Total").font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3, value=note.get("total_cb", total)).font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            ws.cell(row=row, column=3).border = TOP_BOTTOM_BORDER
            row += 1

        elif ntype in ("ledger_list", "ledger_list_abs"):
            for item in note.get("ledgers", []):
                name = item[0]
                cb = item[3] if len(item) > 3 else item[2]
                ws.cell(row=row, column=2, value=name).font = FONT_NORMAL
                ws.cell(row=row, column=3, value=cb).font = FONT_NORMAL
                ws.cell(row=row, column=3).number_format = AMT_FMT
                ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
                row += 1
            ws.cell(row=row, column=2, value="Total").font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3, value=note.get("total_cb", 0)).font = FONT_NORMAL_BOLD
            ws.cell(row=row, column=3).number_format = AMT_FMT
            ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
            ws.cell(row=row, column=3).border = TOP_BOTTOM_BORDER
            row += 1

        row += 2

    # ── Notes 24-26 (Other Disclosures) ──
    ws = wb.create_sheet("Notes 24-26 (Other)")
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    row = 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value=company_info.get("name", "")).font = FONT_COMPANY
    ws.cell(row=row, column=1).alignment = ALIGN_CENTER
    row += 2

    for note_num_str in ["24", "25", "26"]:
        note = notes_data.get(note_num_str)
        if not note:
            continue

        ws.cell(row=row, column=1, value="NOTE").font = FONT_NORMAL_BOLD
        ws.cell(row=row, column=2, value=f"{note_num_str}. {note['title']}").font = FONT_NORMAL_BOLD
        row += 1

        ntype = note.get("type", "")

        if ntype == "related_party":
            ws.cell(row=row, column=2, value="(a) Key Managerial Personnel (KMP):").font = FONT_NORMAL_BOLD
            row += 1
            dir1 = company_info.get("director1_name", "")
            dir2 = company_info.get("director2_name", "")
            if dir1:
                ws.cell(row=row, column=2, value=f"    {dir1} - Director").font = FONT_NORMAL
                row += 1
            if dir2:
                ws.cell(row=row, column=2, value=f"    {dir2} - Director").font = FONT_NORMAL
                row += 1
            row += 1

            ws.cell(row=row, column=2, value="(b) Year-end balances with related parties:").font = FONT_NORMAL_BOLD
            row += 1
            ws.cell(row=row, column=2, value="Particulars").font = FONT_HEADER
            ws.cell(row=row, column=2).border = HEADER_BORDER
            ws.cell(row=row, column=3, value=f"As at {year_end}").font = FONT_HEADER
            ws.cell(row=row, column=3).alignment = ALIGN_CENTER
            ws.cell(row=row, column=3).border = HEADER_BORDER
            row += 1

            ws.cell(row=row, column=2, value="Loan Payable:").font = FONT_NORMAL_BOLD
            row += 1
            for name, ob, cb in note.get("loans", []):
                ws.cell(row=row, column=2, value=f"    {name}").font = FONT_NORMAL
                ws.cell(row=row, column=3, value=cb).font = FONT_NORMAL
                ws.cell(row=row, column=3).number_format = AMT_FMT
                ws.cell(row=row, column=3).alignment = ALIGN_RIGHT
                row += 1

        elif ntype == "text":
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            ws.cell(row=row, column=2, value=note.get("content", "")).font = FONT_NORMAL
            ws.cell(row=row, column=2).alignment = ALIGN_LEFT_WRAP
            ws.row_dimensions[row].height = 30
            row += 1

        row += 2

    # Signatory on last notes sheet
    _write_signatory_block(ws, row, company_info)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN GENERATOR
# ═══════════════════════════════════════════════════════════════════════════════

def generate_schedule_iii(db_path=None, company_info=None, output_path=None):
    """Generate complete Schedule III financial statements in Excel."""
    conn = _get_conn(db_path)
    metadata = _get_metadata(conn)

    if not company_info:
        company_info = {}
    if not company_info.get("name"):
        company_info["name"] = metadata.get("company_name", "Company")
    if not company_info.get("denomination"):
        company_info["denomination"] = "INR"

    # Extract data
    bs_data = extract_balance_sheet_data(conn)
    pl_data = extract_pl_data(conn)
    notes_data = extract_notes_data(conn, bs_data, pl_data)

    # Verify balance sheet balances
    diff = abs(bs_data["total_liabilities"] - bs_data["total_assets"])
    verification = {
        "bs_total_liabilities": bs_data["total_liabilities"],
        "bs_total_assets": bs_data["total_assets"],
        "bs_difference": diff,
        "bs_balanced": diff < 1.0,
        "pl_revenue": pl_data["revenue"],
        "pl_expenses": pl_data["total_expenses"],
        "pl_profit_before_tax": pl_data["profit_before_tax"],
        "pl_profit_after_tax": pl_data["profit_after_tax"],
    }

    # Build Excel
    wb = Workbook()
    _build_balance_sheet(wb, bs_data, company_info)
    _build_pl_statement(wb, pl_data, company_info)
    _build_notes_sheets(wb, notes_data, company_info)

    if output_path:
        wb.save(output_path)

    conn.close()

    return {
        "bs_data": bs_data,
        "pl_data": pl_data,
        "notes": notes_data,
        "verification": verification,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# PREVIEW HELPERS (for Streamlit page)
# ═══════════════════════════════════════════════════════════════════════════════

def get_bs_preview_data(db_path=None):
    """Get Balance Sheet data for Streamlit preview."""
    conn = _get_conn(db_path)
    data = extract_balance_sheet_data(conn)
    conn.close()
    return data


def get_pl_preview_data(db_path=None):
    """Get P&L data for Streamlit preview."""
    conn = _get_conn(db_path)
    data = extract_pl_data(conn)
    conn.close()
    return data
