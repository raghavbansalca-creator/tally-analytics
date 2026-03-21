"""
Seven Labs Vision -- Cash Flow Forecasting Engine
Pure-Python forecasting: trend extrapolation + seasonal adjustment + known payments.
No external ML libraries required.
"""

import sqlite3
import os
import math
from datetime import datetime, timedelta
from collections import defaultdict

DB_PATH = os.path.join(os.path.dirname(__file__), "tally_data.db")

# ---------------------------------------------------------------------------
# DEFAULTS
# ---------------------------------------------------------------------------

DEFAULT_ASSUMPTIONS = {
    "revenue_growth_pct": 0,
    "expense_growth_pct": 0,
    "planned_capex": [],
    "expected_receipts": [],
    "salary_increment_pct": 0,
    "new_hires": [],
    "loan_drawdown": [],
    "loan_repayment_emi": 0,
    "rent_increase_pct": 0,
    "minimum_cash_threshold": 500000,
    "advance_tax_rate_pct": 30,
}

# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

MONTH_NAMES = [
    "", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
]


def _month_label(ym):
    """Convert '202504' to 'Apr 2025'."""
    y = int(ym[:4])
    m = int(ym[4:6])
    return f"{MONTH_NAMES[m]} {y}"


def _next_month(ym):
    """Return the next month code. '202512' -> '202601'."""
    y = int(ym[:4])
    m = int(ym[4:6])
    m += 1
    if m > 12:
        m = 1
        y += 1
    return f"{y}{m:02d}"


def _months_between(start_ym, end_ym):
    """List of month codes from start to end inclusive."""
    months = []
    cur = start_ym
    while cur <= end_ym:
        months.append(cur)
        cur = _next_month(cur)
    return months


def _linear_trend(values):
    """Simple least-squares linear trend slope over a list of numbers.
    Returns slope per period.  Returns 0 if fewer than 2 points."""
    n = len(values)
    if n < 2:
        return 0.0
    x_mean = (n - 1) / 2.0
    y_mean = sum(values) / n
    num = sum((i - x_mean) * (v - y_mean) for i, v in enumerate(values))
    den = sum((i - x_mean) ** 2 for i in range(n))
    if den == 0:
        return 0.0
    return num / den


def _trend_direction(values):
    """Return 'increasing', 'decreasing', or 'stable'."""
    if len(values) < 3:
        return "stable"
    slope = _linear_trend(values)
    mean_val = sum(values) / len(values) if values else 1
    if mean_val == 0:
        return "stable"
    pct = slope / abs(mean_val) * 100
    if pct > 5:
        return "increasing"
    elif pct < -5:
        return "decreasing"
    return "stable"


def _safe_div(num, den):
    if den == 0:
        return 0.0
    return num / den


def _get_conn(db_path=None):
    return sqlite3.connect(db_path or DB_PATH)


# ---------------------------------------------------------------------------
# PART A: HISTORICAL ANALYSIS
# ---------------------------------------------------------------------------

def analyze_historical(db_path=None, months_back=12):
    """Extract and analyze historical cash flow patterns from Tally data."""
    conn = _get_conn(db_path)
    try:
        return _analyze_historical_impl(conn, months_back)
    finally:
        conn.close()


def _analyze_historical_impl(conn, months_back):
    # ---- Determine month range ----
    row = conn.execute("SELECT MIN(DATE), MAX(DATE) FROM trn_voucher").fetchone()
    if not row or not row[0]:
        return {"monthly_data": [], "patterns": {}, "current_position": {}}
    min_date, max_date = row
    max_ym = max_date[:6]
    min_ym = min_date[:6]

    # Go back months_back from max_ym
    y, m = int(max_ym[:4]), int(max_ym[4:6])
    for _ in range(months_back - 1):
        m -= 1
        if m < 1:
            m = 12
            y -= 1
    start_ym = f"{y}{m:02d}"
    if start_ym < min_ym:
        start_ym = min_ym

    all_months = _months_between(start_ym, max_ym)

    # ---- Monthly receipts (Receipt vouchers, deemed-positive side) ----
    receipt_rows = conn.execute("""
        SELECT SUBSTR(v.DATE,1,6) as month,
               SUM(ABS(CAST(a.AMOUNT AS REAL))) as amt
        FROM trn_voucher v
        JOIN trn_accounting a ON a.VOUCHER_GUID = v.GUID
        WHERE v.VOUCHERTYPENAME = 'Receipt' AND a.ISDEEMEDPOSITIVE = 'Yes'
        GROUP BY month
    """).fetchall()
    receipt_by_month = {r[0]: r[1] for r in receipt_rows}

    # ---- Monthly payments (Payment vouchers, deemed-positive side) ----
    payment_rows = conn.execute("""
        SELECT SUBSTR(v.DATE,1,6) as month,
               SUM(ABS(CAST(a.AMOUNT AS REAL))) as amt
        FROM trn_voucher v
        JOIN trn_accounting a ON a.VOUCHER_GUID = v.GUID
        WHERE v.VOUCHERTYPENAME = 'Payment' AND a.ISDEEMEDPOSITIVE = 'Yes'
        GROUP BY month
    """).fetchall()
    payment_by_month = {r[0]: r[1] for r in payment_rows}

    # ---- Sales receipts (from Sales Accounts ledgers) ----
    sales_rows = conn.execute("""
        SELECT SUBSTR(v.DATE,1,6) as month,
               SUM(ABS(CAST(a.AMOUNT AS REAL))) as amt
        FROM trn_voucher v
        JOIN trn_accounting a ON a.VOUCHER_GUID = v.GUID
        JOIN mst_ledger l ON l.NAME = a.LEDGERNAME
        WHERE l.PARENT = 'Sales Accounts'
        GROUP BY month
    """).fetchall()
    sales_by_month = {r[0]: r[1] for r in sales_rows}

    # ---- Purchase payments ----
    purchase_rows = conn.execute("""
        SELECT SUBSTR(v.DATE,1,6) as month,
               SUM(ABS(CAST(a.AMOUNT AS REAL))) as amt
        FROM trn_voucher v
        JOIN trn_accounting a ON a.VOUCHER_GUID = v.GUID
        JOIN mst_ledger l ON l.NAME = a.LEDGERNAME
        WHERE l.PARENT = 'Purchase Accounts'
        GROUP BY month
    """).fetchall()
    purchase_by_month = {r[0]: r[1] for r in purchase_rows}

    # ---- Salary payments ----
    salary_rows = conn.execute("""
        SELECT SUBSTR(v.DATE,1,6) as month,
               SUM(ABS(CAST(a.AMOUNT AS REAL))) as amt
        FROM trn_voucher v
        JOIN trn_accounting a ON a.VOUCHER_GUID = v.GUID
        JOIN mst_ledger l ON l.NAME = a.LEDGERNAME
        WHERE l.PARENT = 'Salary Expenses'
        GROUP BY month
    """).fetchall()
    salary_by_month = {r[0]: r[1] for r in salary_rows}

    # ---- GST payments (Duties & Taxes ledgers with GST/CGST/SGST/IGST) ----
    gst_rows = conn.execute("""
        SELECT SUBSTR(v.DATE,1,6) as month,
               SUM(ABS(CAST(a.AMOUNT AS REAL))) as amt
        FROM trn_voucher v
        JOIN trn_accounting a ON a.VOUCHER_GUID = v.GUID
        JOIN mst_ledger l ON l.NAME = a.LEDGERNAME
        WHERE l.PARENT = 'Duties & Taxes'
          AND (l.NAME LIKE '%GST%' OR l.NAME LIKE '%CGST%'
               OR l.NAME LIKE '%SGST%' OR l.NAME LIKE '%IGST%')
          AND v.VOUCHERTYPENAME = 'Payment'
        GROUP BY month
    """).fetchall()
    gst_by_month = {r[0]: r[1] for r in gst_rows}

    # ---- TDS payments ----
    tds_rows = conn.execute("""
        SELECT SUBSTR(v.DATE,1,6) as month,
               SUM(ABS(CAST(a.AMOUNT AS REAL))) as amt
        FROM trn_voucher v
        JOIN trn_accounting a ON a.VOUCHER_GUID = v.GUID
        JOIN mst_ledger l ON l.NAME = a.LEDGERNAME
        WHERE l.PARENT = 'Duties & Taxes'
          AND l.NAME LIKE '%TDS%'
          AND v.VOUCHERTYPENAME = 'Payment'
        GROUP BY month
    """).fetchall()
    tds_by_month = {r[0]: r[1] for r in tds_rows}

    # ---- Rent payments ----
    rent_rows = conn.execute("""
        SELECT SUBSTR(v.DATE,1,6) as month,
               SUM(ABS(CAST(a.AMOUNT AS REAL))) as amt
        FROM trn_voucher v
        JOIN trn_accounting a ON a.VOUCHER_GUID = v.GUID
        JOIN mst_ledger l ON l.NAME = a.LEDGERNAME
        WHERE l.PARENT = 'Rent Expenses'
        GROUP BY month
    """).fetchall()
    rent_by_month = {r[0]: r[1] for r in rent_rows}

    # ---- Loan payments (EMI detection) ----
    loan_rows = conn.execute("""
        SELECT SUBSTR(v.DATE,1,6) as month,
               SUM(ABS(CAST(a.AMOUNT AS REAL))) as amt
        FROM trn_voucher v
        JOIN trn_accounting a ON a.VOUCHER_GUID = v.GUID
        JOIN mst_ledger l ON l.NAME = a.LEDGERNAME
        WHERE l.PARENT IN ('Secured Loans', 'Unsecured Loans', 'Loans (Liability)')
          AND v.VOUCHERTYPENAME = 'Payment'
        GROUP BY month
    """).fetchall()
    loan_by_month = {r[0]: r[1] for r in loan_rows}

    # ---- Build monthly data ----
    monthly_data = []
    running_bank = None
    for ym in all_months:
        receipts = receipt_by_month.get(ym, 0) or 0
        payments = payment_by_month.get(ym, 0) or 0
        net_cash = receipts - payments
        sales_rec = sales_by_month.get(ym, 0) or 0
        purchase_pay = purchase_by_month.get(ym, 0) or 0
        salary_pay = salary_by_month.get(ym, 0) or 0
        gst_pay = gst_by_month.get(ym, 0) or 0
        tds_pay = tds_by_month.get(ym, 0) or 0
        rent_pay = rent_by_month.get(ym, 0) or 0
        loan_pay = loan_by_month.get(ym, 0) or 0
        other_receipts = max(0, receipts - sales_rec)
        other_payments = max(0, payments - purchase_pay - salary_pay - gst_pay - tds_pay - rent_pay - loan_pay)

        monthly_data.append({
            "month": ym,
            "label": _month_label(ym),
            "receipts": receipts,
            "payments": payments,
            "net_cash": net_cash,
            "opening_bank": 0,
            "closing_bank": 0,
            "sales_receipts": sales_rec,
            "purchase_payments": purchase_pay,
            "salary_payments": salary_pay,
            "gst_payments": gst_pay,
            "tds_payments": tds_pay,
            "rent_payments": rent_pay,
            "loan_payments": loan_pay,
            "other_receipts": other_receipts,
            "other_payments": other_payments,
        })

    # ---- Patterns ----
    receipt_vals = [d["receipts"] for d in monthly_data if d["receipts"] > 0]
    payment_vals = [d["payments"] for d in monthly_data if d["payments"] > 0]
    net_vals = [d["net_cash"] for d in monthly_data]

    avg_receipts = sum(receipt_vals) / len(receipt_vals) if receipt_vals else 0
    avg_payments = sum(payment_vals) / len(payment_vals) if payment_vals else 0
    avg_net = sum(net_vals) / len(net_vals) if net_vals else 0

    # Seasonal detection: find high/low months by comparing to average
    month_avg = defaultdict(list)
    for d in monthly_data:
        m_num = int(d["month"][4:6])
        month_avg[m_num].append(d["receipts"])
    seasonal_high = []
    seasonal_low = []
    for m_num, vals in month_avg.items():
        m_avg = sum(vals) / len(vals) if vals else 0
        if avg_receipts > 0:
            ratio = m_avg / avg_receipts
            if ratio > 1.2:
                seasonal_high.append(MONTH_NAMES[m_num])
            elif ratio < 0.8:
                seasonal_low.append(MONTH_NAMES[m_num])

    # DSO / DPO
    total_sales = sum(sales_by_month.get(ym, 0) or 0 for ym in all_months)
    total_purchases = sum(purchase_by_month.get(ym, 0) or 0 for ym in all_months)

    # Receivables and payables
    debtor_row = conn.execute("""
        SELECT COALESCE(SUM(ABS(CAST(CLOSINGBALANCE AS REAL))), 0)
        FROM mst_ledger WHERE PARENT = 'Sundry Debtors'
    """).fetchone()
    total_receivables = debtor_row[0] if debtor_row else 0

    creditor_row = conn.execute("""
        SELECT COALESCE(SUM(ABS(CAST(CLOSINGBALANCE AS REAL))), 0)
        FROM mst_ledger WHERE PARENT = 'Sundry Creditors'
    """).fetchone()
    total_payables = creditor_row[0] if creditor_row else 0

    months_count = len(all_months)
    annualized_sales = total_sales / months_count * 12 if months_count > 0 else total_sales
    annualized_purchases = total_purchases / months_count * 12 if months_count > 0 else total_purchases

    dso = _safe_div(total_receivables, annualized_sales) * 365
    dpo = _safe_div(total_payables, annualized_purchases) * 365

    patterns = {
        "avg_monthly_receipts": avg_receipts,
        "avg_monthly_payments": avg_payments,
        "avg_net_cash": avg_net,
        "receipt_trend": _trend_direction(receipt_vals),
        "payment_trend": _trend_direction(payment_vals),
        "seasonal_months": {"high": seasonal_high, "low": seasonal_low},
        "dso": round(dso, 1),
        "dpo": round(dpo, 1),
        "working_capital_cycle": round(dso - dpo, 1),
    }

    # ---- Current position ----
    bank_rows = conn.execute("""
        SELECT NAME, PARENT, CAST(CLOSINGBALANCE AS REAL)
        FROM mst_ledger
        WHERE PARENT IN ('Bank Accounts', 'Bank OD A/c')
    """).fetchall()
    # In Tally, bank debit balance (asset) is negative CLOSINGBALANCE
    bank_balance = sum(abs(r[2]) for r in bank_rows if r[2] and r[2] < 0)

    cash_rows = conn.execute("""
        SELECT CAST(CLOSINGBALANCE AS REAL)
        FROM mst_ledger WHERE PARENT = 'Cash-in-Hand'
    """).fetchall()
    # Cash debit balance is negative in Tally
    cash_balance = sum(abs(r[0]) for r in cash_rows if r[0] and r[0] < 0)
    # Also check positive (some Tally versions)
    if cash_balance == 0:
        cash_balance = sum(abs(r[0]) for r in cash_rows if r[0] and r[0] != 0)

    current_position = {
        "bank_balance": bank_balance,
        "cash_balance": cash_balance,
        "total_liquid": bank_balance + cash_balance,
        "receivables": total_receivables,
        "payables": total_payables,
        "net_working_capital": total_receivables - total_payables + bank_balance + cash_balance,
    }

    return {
        "monthly_data": monthly_data,
        "patterns": patterns,
        "current_position": current_position,
    }


# ---------------------------------------------------------------------------
# PART B: FORECAST ENGINE
# ---------------------------------------------------------------------------

def forecast_cashflow(historical, assumptions=None, months_ahead=6, scenario="base"):
    """Generate cash flow forecast based on historical data and assumptions."""
    if assumptions is None:
        assumptions = dict(DEFAULT_ASSUMPTIONS)
    else:
        merged = dict(DEFAULT_ASSUMPTIONS)
        merged.update(assumptions)
        assumptions = merged

    monthly_data = historical.get("monthly_data", [])
    patterns = historical.get("patterns", {})
    current_pos = historical.get("current_position", {})

    if not monthly_data:
        return {
            "scenario": scenario,
            "forecast_months": [],
            "runway_months": 0,
            "min_cash_point": {"month": "", "amount": 0},
            "alerts": [],
        }

    # --- Compute base rates from historical ---
    # Use last 6 months (or all available) with recency weighting
    recent_n = min(6, len(monthly_data))
    recent = monthly_data[-recent_n:]

    # Weighted average: most recent gets highest weight
    weights = list(range(1, recent_n + 1))
    total_weight = sum(weights)

    def weighted_avg(key):
        vals = [d.get(key, 0) or 0 for d in recent]
        return sum(v * w for v, w in zip(vals, weights)) / total_weight

    base_receipts = weighted_avg("receipts")
    base_payments = weighted_avg("payments")
    base_salary = weighted_avg("salary_payments")
    base_rent = weighted_avg("rent_payments")
    base_gst = weighted_avg("gst_payments")
    base_tds = weighted_avg("tds_payments")
    base_loan = weighted_avg("loan_payments")

    # Trend: growth rate per month from linear regression
    receipt_vals = [d["receipts"] for d in monthly_data if d["receipts"] > 0]
    payment_vals = [d["payments"] for d in monthly_data if d["payments"] > 0]
    receipt_slope = _linear_trend(receipt_vals) if len(receipt_vals) >= 3 else 0
    payment_slope = _linear_trend(payment_vals) if len(payment_vals) >= 3 else 0

    # Monthly growth rates
    receipt_growth_pct = _safe_div(receipt_slope, base_receipts) * 100
    payment_growth_pct = _safe_div(payment_slope, base_payments) * 100

    # Cap growth rates at reasonable bounds
    receipt_growth_pct = max(-10, min(15, receipt_growth_pct))
    payment_growth_pct = max(-10, min(15, payment_growth_pct))

    # Seasonal factors: ratio of each calendar month to average
    seasonal_factors = {}
    month_totals = defaultdict(list)
    for d in monthly_data:
        m_num = int(d["month"][4:6])
        month_totals[m_num].append(d["receipts"])
    avg_receipt = patterns.get("avg_monthly_receipts", base_receipts) or base_receipts
    for m_num in range(1, 13):
        vals = month_totals.get(m_num, [])
        if vals and avg_receipt > 0:
            seasonal_factors[m_num] = (sum(vals) / len(vals)) / avg_receipt
        else:
            seasonal_factors[m_num] = 1.0

    # Payment seasonal factors
    payment_seasonal = {}
    pmt_month_totals = defaultdict(list)
    for d in monthly_data:
        m_num = int(d["month"][4:6])
        pmt_month_totals[m_num].append(d["payments"])
    avg_payment = patterns.get("avg_monthly_payments", base_payments) or base_payments
    for m_num in range(1, 13):
        vals = pmt_month_totals.get(m_num, [])
        if vals and avg_payment > 0:
            payment_seasonal[m_num] = (sum(vals) / len(vals)) / avg_payment
        else:
            payment_seasonal[m_num] = 1.0

    # --- Scenario adjustments ---
    receipt_multiplier = 1.0
    payment_multiplier = 1.0
    dso_adj_days = 0

    if scenario == "optimistic":
        receipt_multiplier = 1.15
        payment_multiplier = 0.95
    elif scenario == "pessimistic":
        receipt_multiplier = 0.80
        payment_multiplier = 1.10
        dso_adj_days = 15

    # Manual overrides
    rev_growth_override = assumptions.get("revenue_growth_pct", 0)
    exp_growth_override = assumptions.get("expense_growth_pct", 0)
    salary_increment = assumptions.get("salary_increment_pct", 0) / 100.0
    rent_increase = assumptions.get("rent_increase_pct", 0) / 100.0
    emi_override = assumptions.get("loan_repayment_emi", 0)
    advance_tax_rate = assumptions.get("advance_tax_rate_pct", 30) / 100.0

    # Build planned items lookup
    planned_capex = {item["month"]: item for item in assumptions.get("planned_capex", [])}
    expected_receipts = {item["month"]: item for item in assumptions.get("expected_receipts", [])}
    new_hires_map = {}
    for h in assumptions.get("new_hires", []):
        new_hires_map[h["month"]] = h

    loan_drawdowns = {item["month"]: item["amount"] for item in assumptions.get("loan_drawdown", [])}

    # --- Generate forecast ---
    last_month = monthly_data[-1]["month"]
    current_bank = current_pos.get("total_liquid", 0)

    forecast_months = []
    cumulative_new_hire_cost = 0

    for i in range(months_ahead):
        forecast_ym = last_month
        for _ in range(i + 1):
            forecast_ym = _next_month(forecast_ym)

        cal_month = int(forecast_ym[4:6])
        cal_year = int(forecast_ym[:4])

        # --- Projected receipts ---
        trend_factor = 1.0 + (receipt_growth_pct / 100.0) * (i + 1)
        if rev_growth_override:
            trend_factor = 1.0 + (rev_growth_override / 100.0) * (i + 1)
        seasonal_f = seasonal_factors.get(cal_month, 1.0)
        proj_receipts = base_receipts * trend_factor * seasonal_f * receipt_multiplier

        # Add expected large receipts
        if forecast_ym in expected_receipts:
            proj_receipts += expected_receipts[forecast_ym].get("amount", 0)

        # Add loan drawdowns
        if forecast_ym in loan_drawdowns:
            proj_receipts += loan_drawdowns[forecast_ym]

        # --- Projected payments ---
        pmt_trend_factor = 1.0 + (payment_growth_pct / 100.0) * (i + 1)
        if exp_growth_override:
            pmt_trend_factor = 1.0 + (exp_growth_override / 100.0) * (i + 1)
        pmt_seasonal_f = payment_seasonal.get(cal_month, 1.0)
        proj_payments_base = base_payments * pmt_trend_factor * pmt_seasonal_f * payment_multiplier

        # Salary with increment (apply in April or proportionally)
        proj_salary = base_salary
        if salary_increment > 0 and cal_month == 4:
            proj_salary = base_salary * (1 + salary_increment)
        elif salary_increment > 0 and cal_month > 4:
            proj_salary = base_salary * (1 + salary_increment)

        # New hires cumulative cost
        if forecast_ym in new_hires_map:
            hire = new_hires_map[forecast_ym]
            cumulative_new_hire_cost += hire.get("monthly_cost", 0) * hire.get("count", 1)
        proj_salary += cumulative_new_hire_cost

        # Rent with annual increase
        proj_rent = base_rent
        if rent_increase > 0:
            years_ahead = (cal_year - int(last_month[:4]))
            if cal_month >= 4:
                years_ahead = max(0, years_ahead)
            proj_rent = base_rent * (1 + rent_increase * max(0, years_ahead))

        # GST payment (20th of every month)
        proj_gst = base_gst * pmt_trend_factor

        # TDS payment (7th of every month)
        proj_tds = base_tds * pmt_trend_factor

        # EMI / loan repayment
        proj_emi = emi_override if emi_override > 0 else base_loan

        # Advance tax (15 Jun = 15%, 15 Sep = 45%, 15 Dec = 75%, 15 Mar = 100%)
        proj_advance_tax = 0
        # Estimate annual profit from receipts - payments
        est_annual_profit = (base_receipts - base_payments) * 12
        if est_annual_profit > 0:
            annual_tax = est_annual_profit * advance_tax_rate
            if cal_month == 6:
                proj_advance_tax = annual_tax * 0.15
            elif cal_month == 9:
                proj_advance_tax = annual_tax * 0.30  # cumulative 45% - 15% already paid
            elif cal_month == 12:
                proj_advance_tax = annual_tax * 0.30  # cumulative 75% - 45%
            elif cal_month == 3:
                proj_advance_tax = annual_tax * 0.25  # remaining 25%

        # CapEx
        proj_capex = 0
        if forecast_ym in planned_capex:
            proj_capex = planned_capex[forecast_ym].get("amount", 0)

        # Total projected payments
        proj_payments = (proj_payments_base + proj_capex + proj_advance_tax)
        # (salary, rent, gst, tds, emi are already part of base_payments from historical;
        #  only add incremental differences)
        salary_diff = max(0, proj_salary - base_salary)
        rent_diff = max(0, proj_rent - base_rent)
        proj_payments += salary_diff + rent_diff

        proj_net = proj_receipts - proj_payments
        current_bank += proj_net

        # Confidence decreases over time
        confidence = max(0.40, 0.90 - i * 0.08)

        forecast_months.append({
            "month": forecast_ym,
            "label": _month_label(forecast_ym),
            "projected_receipts": round(proj_receipts, 2),
            "projected_payments": round(proj_payments, 2),
            "projected_net": round(proj_net, 2),
            "projected_closing_bank": round(current_bank, 2),
            "gst_payment": round(proj_gst, 2),
            "tds_payment": round(proj_tds, 2),
            "advance_tax": round(proj_advance_tax, 2),
            "salary": round(proj_salary, 2),
            "rent": round(proj_rent, 2),
            "emi": round(proj_emi, 2),
            "capex": round(proj_capex, 2),
            "confidence": round(confidence, 2),
        })

    # --- Runway ---
    runway = calculate_runway(forecast_months, current_pos.get("total_liquid", 0))

    # --- Min cash point ---
    min_cash = {"month": "", "amount": current_pos.get("total_liquid", 0)}
    for fm in forecast_months:
        if fm["projected_closing_bank"] < min_cash["amount"]:
            min_cash = {"month": fm["label"], "amount": fm["projected_closing_bank"]}

    # --- Alerts ---
    alerts = generate_alerts(forecast_months, assumptions)

    return {
        "scenario": scenario,
        "forecast_months": forecast_months,
        "runway_months": runway,
        "min_cash_point": min_cash,
        "alerts": alerts,
    }


# ---------------------------------------------------------------------------
# PART D: RUNWAY & ALERTS
# ---------------------------------------------------------------------------

def calculate_runway(forecast_months, starting_cash=0):
    """Calculate months of cash remaining at current burn rate."""
    if not forecast_months:
        return 0

    # Find when cash goes negative
    for i, fm in enumerate(forecast_months):
        if fm["projected_closing_bank"] < 0:
            return i  # goes negative in month i (0-indexed)

    # If cash never goes negative, return full forecast period
    # But also check average burn rate
    avg_net = sum(fm["projected_net"] for fm in forecast_months) / len(forecast_months)
    if avg_net < 0 and starting_cash > 0:
        return min(len(forecast_months), int(starting_cash / abs(avg_net)))

    return len(forecast_months)


def generate_alerts(forecast_months, assumptions=None):
    """Generate actionable alerts."""
    if assumptions is None:
        assumptions = dict(DEFAULT_ASSUMPTIONS)
    else:
        merged = dict(DEFAULT_ASSUMPTIONS)
        merged.update(assumptions)
        assumptions = merged

    threshold = assumptions.get("minimum_cash_threshold", 500000)
    alerts = []

    for fm in forecast_months:
        ym = fm["month"]
        label = fm["label"]
        cal_month = int(ym[4:6])

        # LOW_CASH / NEGATIVE_CASH
        if fm["projected_closing_bank"] < 0:
            alerts.append({
                "month": label,
                "type": "NEGATIVE_CASH",
                "severity": "critical",
                "message": f"Cash balance projected to go negative: Rs {fm['projected_closing_bank']:,.0f}",
            })
        elif fm["projected_closing_bank"] < threshold:
            alerts.append({
                "month": label,
                "type": "LOW_CASH",
                "severity": "warning",
                "message": f"Cash drops below Rs {threshold/100000:.1f}L threshold: Rs {fm['projected_closing_bank']:,.0f}",
            })

        # GST_DUE (20th of every month)
        if fm.get("gst_payment", 0) > 0:
            alerts.append({
                "month": label,
                "type": "GST_DUE",
                "severity": "info",
                "message": f"GST payment of Rs {fm['gst_payment']:,.0f} due by 20th",
            })

        # TDS_DUE (7th of every month)
        if fm.get("tds_payment", 0) > 0:
            alerts.append({
                "month": label,
                "type": "TDS_DUE",
                "severity": "info",
                "message": f"TDS payment of Rs {fm['tds_payment']:,.0f} due by 7th",
            })

        # ADVANCE_TAX
        if fm.get("advance_tax", 0) > 0:
            due_date_map = {6: "15 Jun", 9: "15 Sep", 12: "15 Dec", 3: "15 Mar"}
            due_str = due_date_map.get(cal_month, f"15th {label}")
            alerts.append({
                "month": label,
                "type": "ADVANCE_TAX",
                "severity": "warning",
                "message": f"Advance tax installment of Rs {fm['advance_tax']:,.0f} due by {due_str}",
            })

        # CAPEX_PLANNED
        if fm.get("capex", 0) > 0:
            alerts.append({
                "month": label,
                "type": "CAPEX_PLANNED",
                "severity": "info",
                "message": f"Planned capital expenditure of Rs {fm['capex']:,.0f}",
            })

        # EMI_DUE
        if fm.get("emi", 0) > 0:
            alerts.append({
                "month": label,
                "type": "EMI_DUE",
                "severity": "info",
                "message": f"Loan EMI of Rs {fm['emi']:,.0f} due",
            })

    return alerts


# ---------------------------------------------------------------------------
# PART E: EXCEL EXPORT
# ---------------------------------------------------------------------------

def export_forecast_excel(historical, forecasts, output_path):
    """Export to Excel with multiple sheets.
    forecasts: dict of {scenario_name: forecast_result}
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    except ImportError:
        # Fallback: try xlsxwriter or just return None
        return _export_forecast_csv(historical, forecasts, output_path)

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    num_fmt = '#,##0'
    pct_fmt = '0.0%'
    thin_border = Border(
        bottom=Side(style='thin', color='E2E8F0')
    )
    bold_font = Font(name="Calibri", bold=True, size=11)
    green_font = Font(name="Calibri", bold=True, color="059669")
    red_font = Font(name="Calibri", bold=True, color="DC2626")

    def write_header_row(ws, row, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 20)

    # ---- Sheet 1: Summary ----
    ws = wb.active
    ws.title = "Summary"
    current_pos = historical.get("current_position", {})
    patterns = historical.get("patterns", {})

    ws.cell(row=1, column=1, value="CASH FLOW FORECAST SUMMARY").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}")
    ws.cell(row=4, column=1, value="CURRENT POSITION").font = bold_font
    summary_items = [
        ("Bank Balance", current_pos.get("bank_balance", 0)),
        ("Cash Balance", current_pos.get("cash_balance", 0)),
        ("Total Liquid Funds", current_pos.get("total_liquid", 0)),
        ("Receivables", current_pos.get("receivables", 0)),
        ("Payables", current_pos.get("payables", 0)),
        ("Net Working Capital", current_pos.get("net_working_capital", 0)),
    ]
    for i, (label, val) in enumerate(summary_items):
        ws.cell(row=5 + i, column=1, value=label)
        cell = ws.cell(row=5 + i, column=2, value=round(val, 0))
        cell.number_format = num_fmt

    row_offset = 5 + len(summary_items) + 1
    ws.cell(row=row_offset, column=1, value="KEY METRICS").font = bold_font
    ws.cell(row=row_offset + 1, column=1, value="DSO (Days Sales Outstanding)")
    ws.cell(row=row_offset + 1, column=2, value=patterns.get("dso", 0))
    ws.cell(row=row_offset + 2, column=1, value="DPO (Days Payable Outstanding)")
    ws.cell(row=row_offset + 2, column=2, value=patterns.get("dpo", 0))
    ws.cell(row=row_offset + 3, column=1, value="Working Capital Cycle (days)")
    ws.cell(row=row_offset + 3, column=2, value=patterns.get("working_capital_cycle", 0))
    ws.cell(row=row_offset + 4, column=1, value="Receipt Trend")
    ws.cell(row=row_offset + 4, column=2, value=patterns.get("receipt_trend", ""))
    ws.cell(row=row_offset + 5, column=1, value="Payment Trend")
    ws.cell(row=row_offset + 5, column=2, value=patterns.get("payment_trend", ""))

    auto_width(ws)

    # ---- Sheet 2: Historical ----
    ws_hist = wb.create_sheet("Historical")
    hist_headers = [
        "Month", "Receipts", "Payments", "Net Cash Flow",
        "Sales Receipts", "Purchase Payments", "Salary", "GST", "TDS",
        "Rent", "Loan Payments", "Other Receipts", "Other Payments",
    ]
    write_header_row(ws_hist, 1, hist_headers)
    for i, d in enumerate(historical.get("monthly_data", []), 2):
        ws_hist.cell(row=i, column=1, value=d["label"])
        ws_hist.cell(row=i, column=2, value=round(d["receipts"], 0)).number_format = num_fmt
        ws_hist.cell(row=i, column=3, value=round(d["payments"], 0)).number_format = num_fmt
        ws_hist.cell(row=i, column=4, value=round(d["net_cash"], 0)).number_format = num_fmt
        ws_hist.cell(row=i, column=5, value=round(d.get("sales_receipts", 0), 0)).number_format = num_fmt
        ws_hist.cell(row=i, column=6, value=round(d.get("purchase_payments", 0), 0)).number_format = num_fmt
        ws_hist.cell(row=i, column=7, value=round(d.get("salary_payments", 0), 0)).number_format = num_fmt
        ws_hist.cell(row=i, column=8, value=round(d.get("gst_payments", 0), 0)).number_format = num_fmt
        ws_hist.cell(row=i, column=9, value=round(d.get("tds_payments", 0), 0)).number_format = num_fmt
        ws_hist.cell(row=i, column=10, value=round(d.get("rent_payments", 0), 0)).number_format = num_fmt
        ws_hist.cell(row=i, column=11, value=round(d.get("loan_payments", 0), 0)).number_format = num_fmt
        ws_hist.cell(row=i, column=12, value=round(d.get("other_receipts", 0), 0)).number_format = num_fmt
        ws_hist.cell(row=i, column=13, value=round(d.get("other_payments", 0), 0)).number_format = num_fmt
    auto_width(ws_hist)

    # ---- Forecast sheets ----
    for scenario_name, forecast in forecasts.items():
        sheet_name = f"{scenario_name.title()} Forecast"
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        ws_fc = wb.create_sheet(sheet_name)
        fc_headers = [
            "Month", "Projected Receipts", "Projected Payments", "Net Cash Flow",
            "Closing Bank", "GST", "TDS", "Advance Tax", "Salary", "Rent", "EMI",
            "CapEx", "Confidence",
        ]
        write_header_row(ws_fc, 1, fc_headers)
        for i, fm in enumerate(forecast.get("forecast_months", []), 2):
            ws_fc.cell(row=i, column=1, value=fm["label"])
            ws_fc.cell(row=i, column=2, value=round(fm["projected_receipts"], 0)).number_format = num_fmt
            ws_fc.cell(row=i, column=3, value=round(fm["projected_payments"], 0)).number_format = num_fmt
            ws_fc.cell(row=i, column=4, value=round(fm["projected_net"], 0)).number_format = num_fmt
            cell_bank = ws_fc.cell(row=i, column=5, value=round(fm["projected_closing_bank"], 0))
            cell_bank.number_format = num_fmt
            if fm["projected_closing_bank"] < 0:
                cell_bank.font = red_font
            else:
                cell_bank.font = green_font
            ws_fc.cell(row=i, column=6, value=round(fm.get("gst_payment", 0), 0)).number_format = num_fmt
            ws_fc.cell(row=i, column=7, value=round(fm.get("tds_payment", 0), 0)).number_format = num_fmt
            ws_fc.cell(row=i, column=8, value=round(fm.get("advance_tax", 0), 0)).number_format = num_fmt
            ws_fc.cell(row=i, column=9, value=round(fm.get("salary", 0), 0)).number_format = num_fmt
            ws_fc.cell(row=i, column=10, value=round(fm.get("rent", 0), 0)).number_format = num_fmt
            ws_fc.cell(row=i, column=11, value=round(fm.get("emi", 0), 0)).number_format = num_fmt
            ws_fc.cell(row=i, column=12, value=round(fm.get("capex", 0), 0)).number_format = num_fmt
            ws_fc.cell(row=i, column=13, value=fm.get("confidence", 0))
        auto_width(ws_fc)

    # ---- Alerts sheet ----
    ws_alerts = wb.create_sheet("Alerts")
    alert_headers = ["Month", "Type", "Severity", "Message"]
    write_header_row(ws_alerts, 1, alert_headers)
    row_num = 2
    for scenario_name, forecast in forecasts.items():
        for a in forecast.get("alerts", []):
            ws_alerts.cell(row=row_num, column=1, value=a.get("month", ""))
            ws_alerts.cell(row=row_num, column=2, value=a.get("type", ""))
            ws_alerts.cell(row=row_num, column=3, value=a.get("severity", ""))
            ws_alerts.cell(row=row_num, column=4, value=a.get("message", ""))
            if a.get("severity") == "critical":
                for c in range(1, 5):
                    ws_alerts.cell(row=row_num, column=c).font = red_font
            row_num += 1
    auto_width(ws_alerts)

    # ---- Assumptions sheet ----
    ws_assum = wb.create_sheet("Assumptions")
    ws_assum.cell(row=1, column=1, value="FORECAST ASSUMPTIONS").font = Font(bold=True, size=14)
    assumption_items = [
        ("Revenue Growth %", assumptions.get("revenue_growth_pct", 0)),
        ("Expense Growth %", assumptions.get("expense_growth_pct", 0)),
        ("Salary Increment %", assumptions.get("salary_increment_pct", 0)),
        ("Rent Increase %", assumptions.get("rent_increase_pct", 0)),
        ("Loan EMI (monthly)", assumptions.get("loan_repayment_emi", 0)),
        ("Minimum Cash Threshold", assumptions.get("minimum_cash_threshold", 500000)),
        ("Advance Tax Rate %", assumptions.get("advance_tax_rate_pct", 30)),
    ]
    for i, (label, val) in enumerate(assumption_items, 3):
        ws_assum.cell(row=i, column=1, value=label)
        ws_assum.cell(row=i, column=2, value=val)

    capex_start = 3 + len(assumption_items) + 1
    ws_assum.cell(row=capex_start, column=1, value="PLANNED CAPEX").font = bold_font
    for i, item in enumerate(assumptions.get("planned_capex", []), capex_start + 1):
        ws_assum.cell(row=i, column=1, value=item.get("month", ""))
        ws_assum.cell(row=i, column=2, value=item.get("amount", 0))
        ws_assum.cell(row=i, column=3, value=item.get("description", ""))
    auto_width(ws_assum)

    wb.save(output_path)
    return output_path


def _export_forecast_csv(historical, forecasts, output_path):
    """Fallback CSV export if openpyxl is not available."""
    import csv
    csv_path = output_path.replace(".xlsx", ".csv")
    with open(csv_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Month", "Receipts", "Payments", "Net Cash Flow"])
        for d in historical.get("monthly_data", []):
            writer.writerow([d["label"], d["receipts"], d["payments"], d["net_cash"]])
        writer.writerow([])
        for scenario_name, forecast in forecasts.items():
            writer.writerow([f"--- {scenario_name.title()} Forecast ---"])
            writer.writerow(["Month", "Projected Receipts", "Projected Payments",
                             "Net Cash Flow", "Closing Bank"])
            for fm in forecast.get("forecast_months", []):
                writer.writerow([
                    fm["label"], fm["projected_receipts"],
                    fm["projected_payments"], fm["projected_net"],
                    fm["projected_closing_bank"],
                ])
            writer.writerow([])
    return csv_path


# ---------------------------------------------------------------------------
# QUICK TEST
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    print("Analyzing historical cash flow...")
    hist = analyze_historical()
    print(f"  Months of data: {len(hist['monthly_data'])}")
    print(f"  Avg monthly receipts: {hist['patterns'].get('avg_monthly_receipts', 0):,.0f}")
    print(f"  Avg monthly payments: {hist['patterns'].get('avg_monthly_payments', 0):,.0f}")
    print(f"  Receipt trend: {hist['patterns'].get('receipt_trend', '')}")
    print(f"  DSO: {hist['patterns'].get('dso', 0)} days")
    print(f"  DPO: {hist['patterns'].get('dpo', 0)} days")
    print(f"  Current bank: {hist['current_position'].get('bank_balance', 0):,.0f}")
    print(f"  Current cash: {hist['current_position'].get('cash_balance', 0):,.0f}")
    print(f"  Receivables: {hist['current_position'].get('receivables', 0):,.0f}")
    print(f"  Payables: {hist['current_position'].get('payables', 0):,.0f}")

    print("\nForecasting (base scenario, 6 months)...")
    fc = forecast_cashflow(hist, months_ahead=6, scenario="base")
    print(f"  Runway: {fc['runway_months']} months")
    print(f"  Min cash point: {fc['min_cash_point']}")
    for fm in fc["forecast_months"]:
        print(f"    {fm['label']}: Receipts={fm['projected_receipts']:,.0f}  "
              f"Payments={fm['projected_payments']:,.0f}  "
              f"Closing={fm['projected_closing_bank']:,.0f}  "
              f"Conf={fm['confidence']}")
    print(f"  Alerts: {len(fc['alerts'])}")
    for a in fc["alerts"][:5]:
        print(f"    [{a['type']}] {a['month']}: {a['message']}")
