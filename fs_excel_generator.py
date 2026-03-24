"""
Financial Statement Excel Generator

Generates professional Schedule III (IGAAP) financial statements in Excel format.
Uses SUMPRODUCT formulas to link all B/S and P&L amounts to the source trial balance.

Author: Seven Labs Vision
Date: 2026-03-22
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from decimal import Decimal
from datetime import datetime
import logging
import sqlite3

from fs_engine import (
    TallyDataExtractor,
    ScheduleIIIClassifier,
    YearEndAdjustments,
    ReconciliationEngine,
    ScheduleIIISection,
    ClassifiedLedger,
    PriorYearParser,
)

logger = logging.getLogger(__name__)


class FinancialStatementsExcelGenerator:
    """
    Generates a professional Excel workbook with:
    - Trial Balance (TB_Tally) sheet with all ledgers and Schedule III mapping
    - Balance Sheet with SUMPRODUCT formulas
    - P&L Statement with SUMPRODUCT formulas
    - Reconciliation sheet with 8 verification checks
    """

    def __init__(self, classifier: ScheduleIIIClassifier, adjustments: YearEndAdjustments,
                 company_name: str = "ROHIT PHARMA", as_on_date: str = "31-03-2026",
                 extractor: TallyDataExtractor = None, db_path: str = None,
                 prior_year_data: dict = None, prior_year_date: str = None):
        """
        Initialize Excel generator.

        Args:
            classifier: ScheduleIIIClassifier with classified ledgers
            adjustments: YearEndAdjustments with adjustment entries
            company_name: Company name for headers
            as_on_date: Balance sheet date (DD-MM-YYYY format)
            extractor: Optional TallyDataExtractor for database access
            db_path: Optional database path for direct access
            prior_year_data: Optional dict with {"bs": {...}, "pl": {...}} prior year figures
            prior_year_date: Optional prior year date for column header
        """
        self.classifier = classifier
        self.adjustments = adjustments
        self.company_name = company_name
        self.as_on_date = as_on_date
        self.workbook = None
        self.classified_ledgers = classifier.classified_ledgers
        self.extractor = extractor
        self.db_path = db_path
        self.prior_year_data = prior_year_data or {}
        self.prior_year_date = prior_year_date or "31-Mar-2025"

        # Style definitions
        self.header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.subheader_font = Font(name="Arial", size=10, bold=True)
        self.subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        self.data_font = Font(name="Arial", size=10)
        self.number_format = "#,##0.00"
        self.thousands_format = "#,##0"
        self.input_font = Font(name="Arial", size=10, color="0070C0")
        self.formula_font = Font(name="Arial", size=10, color="000000")
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def generate(self, output_path: str) -> str:
        """
        Generate complete Excel workbook.

        Args:
            output_path: Path to save output Excel file

        Returns:
            Path to generated file
        """
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

        # Create sheets in order
        self._create_tb_tally_sheet()
        self._create_bs_sheet()
        self._create_pl_sheet()
        self._create_notes_sheet()
        self._create_depreciation_sheet()
        self._create_ratios_sheet()
        self._create_tr_ageing_sheet()
        self._create_tp_ageing_sheet()
        self._create_cfs_sheet()
        self._create_recon_sheet()

        # Save workbook
        self.workbook.save(output_path)
        logger.info(f"Generated financial statements: {output_path}")
        return output_path

    def _get_prior_year_value(self, section: str, key: str, default: Decimal = Decimal(0)) -> Decimal:
        """Get prior year value from parsed data, convert to thousands."""
        if not self.prior_year_data or section not in self.prior_year_data:
            return default
        val = self.prior_year_data[section].get(key, default)
        if val:
            return val / Decimal(1000)  # Convert to thousands
        return default

    def _has_prior_year_data(self) -> bool:
        """Check if prior year data is available."""
        return bool(self.prior_year_data and (self.prior_year_data.get("bs") or self.prior_year_data.get("pl")))

    def _create_tb_tally_sheet(self):
        """
        Create TB_Tally sheet with all ledgers and Schedule III mapping.

        Columns:
        A: Ledger Name
        B: Tally Group
        C: Schedule III Section
        D: Schedule III Line
        E: Schedule III Sub
        F: Opening Balance
        G: Closing Balance
        H: Display Amount
        I: Is Reclassified
        J: Reclassification Note
        """
        ws = self.workbook.create_sheet("TB_Tally", 0)

        # Headers
        headers = [
            "Ledger Name",
            "Tally Group",
            "Schedule III Section",
            "Schedule III Line",
            "Schedule III Sub",
            "Opening Balance",
            "Closing Balance",
            "Display Amount",
            "Is Reclassified",
            "Reclassification Note",
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self.border

        # Data rows
        row_num = 2
        for ledger in sorted(self.classified_ledgers, key=lambda x: x.name):
            ws.cell(row=row_num, column=1).value = ledger.name
            ws.cell(row=row_num, column=2).value = ledger.tally_group
            ws.cell(row=row_num, column=3).value = ledger.schedule_iii_section.value
            ws.cell(row=row_num, column=4).value = ledger.schedule_iii_line
            ws.cell(row=row_num, column=5).value = ledger.schedule_iii_sub
            ws.cell(row=row_num, column=6).value = float(ledger.opening_balance) / 1000  # in thousands
            ws.cell(row=row_num, column=7).value = float(ledger.closing_balance) / 1000  # in thousands
            ws.cell(row=row_num, column=8).value = float(ledger.display_amount) / 1000  # in thousands
            ws.cell(row=row_num, column=9).value = "Yes" if ledger.is_reclassified else "No"
            ws.cell(row=row_num, column=10).value = ledger.reclassification_note

            # Format row
            for col_num in range(1, 11):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = self.data_font
                cell.border = self.border
                if col_num in (6, 7, 8):  # Number columns
                    cell.number_format = self.thousands_format
                cell.alignment = Alignment(horizontal="left" if col_num <= 5 else "right", vertical="center")

            row_num += 1

        # Set column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 12
        ws.column_dimensions["J"].width = 35

    def _create_bs_sheet(self):
        """
        Create Balance Sheet sheet with EXACT Schedule III (Division I) format.

        Structure:
        - Company name and CIN headers
        - Balance Sheet as at [date]
        - Amount in Thousands INR note
        - Column headers: Particulars | Note | Current Date | Prior Year
        - All amounts are SUMIF formulas referencing TB_Tally sheet
        - Every SUMIF divided by 1000 for thousands display
        - Note references in blue (0070C0)
        - Section headers with dark blue background
        - Total lines with borders and bold font
        """
        ws = self.workbook.create_sheet("Balance Sheet", 1)

        # Set column widths
        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20

        # Currency format for amounts (thousands, no decimals)
        currency_fmt = "#,##0;(#,##0);-"

        # Row counter
        row = 1

        # HEADER BLOCK
        ws[f"A{row}"] = self.company_name
        ws[f"A{row}"].font = Font(name="Arial", size=12, bold=True)
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        ws[f"A{row}"] = "CIN: [CIN Number]"
        ws[f"A{row}"].font = Font(name="Arial", size=10)
        row += 1

        ws[f"A{row}"] = "Registered Office: [Address]"
        ws[f"A{row}"].font = Font(name="Arial", size=10)
        row += 1

        # Title and date
        ws[f"A{row}"] = f"Balance Sheet as at {self.as_on_date}"
        ws[f"A{row}"].font = Font(name="Arial", size=11, bold=True)
        row += 1

        ws[f"A{row}"] = "(Amount in Thousands INR)"
        ws[f"A{row}"].font = Font(name="Arial", size=9, italic=True)
        row += 2

        # COLUMN HEADERS
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

        ws[f"A{row}"] = "Particulars"
        ws[f"B{row}"] = "Note"
        ws[f"C{row}"] = f"As at {self.as_on_date}"
        ws[f"D{row}"] = f"As at {self.prior_year_date}"

        for col in ["A", "B", "C", "D"]:
            cell = ws[f"{col}{row}"]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        row += 1

        # SECTION I: EQUITY AND LIABILITIES
        section_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        section_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

        ws[f"A{row}"] = "I. EQUITY AND LIABILITIES"
        ws[f"A{row}"].font = section_font
        ws[f"A{row}"].fill = section_fill
        row += 1

        # (1) Shareholders' Funds
        subsection_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subsection_font = Font(name="Arial", size=10, bold=True)

        ws[f"A{row}"] = "(1) Shareholders' Funds"
        ws[f"A{row}"].font = subsection_font
        ws[f"A{row}"].fill = subsection_fill
        row += 1

        sf_start = row

        # Share Capital
        ws[f"A{row}"] = "    (a) Share Capital"
        ws[f"B{row}"] = "1"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"share_capital",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "share_capital")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Reserves and Surplus
        ws[f"A{row}"] = "    (b) Reserves and Surplus"
        ws[f"B{row}"] = "2"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"other_equity",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "other_equity")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Money against share warrants
        ws[f"A{row}"] = "    (c) Money received against share warrants"
        ws[f"B{row}"] = ""
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"money_against_warrants",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = 0
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # (2) Share application money
        ws[f"A{row}"] = "(2) Share application money pending allotment"
        ws[f"B{row}"] = ""
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"share_app_money",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = 0
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # (3) Non-Current Liabilities
        ws[f"A{row}"] = "(3) Non-Current Liabilities"
        ws[f"A{row}"].font = subsection_font
        ws[f"A{row}"].fill = subsection_fill
        row += 1

        ncl_start = row

        # Long-term borrowings
        ws[f"A{row}"] = "    (a) Long-term borrowings"
        ws[f"B{row}"] = "3"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"borrowings",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "borrowings")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Deferred tax liabilities
        ws[f"A{row}"] = "    (b) Deferred tax liabilities (Net)"
        ws[f"B{row}"] = "4"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"deferred_tax_liabilities",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "deferred_tax_liabilities")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Other long-term liabilities
        ws[f"A{row}"] = "    (c) Other Long term liabilities"
        ws[f"B{row}"] = ""
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"other_long_term_liabilities",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = 0
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Long-term provisions
        ws[f"A{row}"] = "    (d) Long term provisions"
        ws[f"B{row}"] = ""
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"long_term_provisions",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = 0
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # (4) Current Liabilities
        ws[f"A{row}"] = "(4) Current Liabilities"
        ws[f"A{row}"].font = subsection_font
        ws[f"A{row}"].fill = subsection_fill
        row += 1

        cl_start = row

        # Short-term borrowings
        ws[f"A{row}"] = "    (a) Short-term borrowings"
        ws[f"B{row}"] = "5"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"borrowings_cl",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "borrowings_cl")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Trade payables
        ws[f"A{row}"] = "    (b) Trade payables"
        ws[f"B{row}"] = "6"
        row += 1

        tp_start = row

        # MSME trade payables
        ws[f"A{row}"] = "        (A) total outstanding dues of MSME"
        ws[f"C{row}"] = '=SUMIFS(TB_Tally!H:H,TB_Tally!D:D,"trade_payables",TB_Tally!E:E,"msme")/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "trade_payables_msme")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Others trade payables
        ws[f"A{row}"] = "        (B) total outstanding dues of others"
        ws[f"C{row}"] = '=SUMIFS(TB_Tally!H:H,TB_Tally!D:D,"trade_payables",TB_Tally!E:E,"others")/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "trade_payables_others")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Other current liabilities
        ws[f"A{row}"] = "    (c) Other current liabilities"
        ws[f"B{row}"] = "7"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"other_current_liabilities",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "other_current_liabilities")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Short-term provisions
        ws[f"A{row}"] = "    (d) Short-term provisions"
        ws[f"B{row}"] = "8"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"provisions",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "provisions")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # TOTAL (Column is dynamic)
        total_row = row
        ws[f"A{row}"] = "TOTAL EQUITY AND LIABILITIES"
        ws[f"A{row}"].font = Font(name="Arial", size=10, bold=True)
        ws[f"C{row}"] = f"=SUM(C{sf_start}:C{row-1})"
        ws[f"C{row}"].font = Font(name="Arial", size=10, bold=True)
        ws[f"C{row}"].number_format = currency_fmt
        ws[f"C{row}"].border = Border(
            top=Side(style="thin"),
            bottom=Side(style="double"),
            left=Side(style="thin"),
            right=Side(style="thin"),
        )
        if self._has_prior_year_data():
            ws[f"D{row}"] = f"=SUM(D{sf_start}:D{row-1})"
            ws[f"D{row}"].font = Font(name="Arial", size=10, bold=True)
            ws[f"D{row}"].number_format = currency_fmt
            ws[f"D{row}"].border = Border(
                top=Side(style="thin"),
                bottom=Side(style="double"),
                left=Side(style="thin"),
                right=Side(style="thin"),
            )
        row += 2

        # SECTION II: ASSETS
        ws[f"A{row}"] = "II. ASSETS"
        ws[f"A{row}"].font = section_font
        ws[f"A{row}"].fill = section_fill
        row += 1

        # (1) Non-Current Assets
        ws[f"A{row}"] = "(1) Non-current assets"
        ws[f"A{row}"].font = subsection_font
        ws[f"A{row}"].fill = subsection_fill
        row += 1

        noca_start = row

        # Property, Plant and Equipment
        ws[f"A{row}"] = "    (a) Property, Plant and Equipment and Intangible assets"
        ws[f"B{row}"] = "9"
        row += 1

        ws[f"A{row}"] = "        (i) Property, Plant and Equipment"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"property_plant_equipment",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "property_plant_equipment")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        ws[f"A{row}"] = "        (ii) Intangible assets"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"intangible_assets",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "intangible_assets")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        ws[f"A{row}"] = "        (iii) Capital work-in-progress"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"capital_wip",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = 0
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        ws[f"A{row}"] = "        (iv) Intangible assets under development"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"intangible_wip",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = 0
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Non-current investments
        ws[f"A{row}"] = "    (b) Non-current investments"
        ws[f"B{row}"] = "10"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"investments",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "investments")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Deferred tax assets
        ws[f"A{row}"] = "    (c) Deferred tax assets (net)"
        ws[f"B{row}"] = "4"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"deferred_tax_assets",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "deferred_tax_assets")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Long-term loans and advances
        ws[f"A{row}"] = "    (d) Long term loans and advances"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"long_term_loans",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = 0
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Other non-current assets
        ws[f"A{row}"] = "    (e) Other non-current assets"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"other_noncurrent_assets",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = 0
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # (2) Current Assets
        ws[f"A{row}"] = "(2) Current assets"
        ws[f"A{row}"].font = subsection_font
        ws[f"A{row}"].fill = subsection_fill
        row += 1

        ca_start = row

        # Current investments
        ws[f"A{row}"] = "    (a) Current investments"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"current_investments",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = 0
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Inventories
        ws[f"A{row}"] = "    (b) Inventories"
        ws[f"B{row}"] = "11"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"inventories",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "inventories")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Trade receivables
        ws[f"A{row}"] = "    (c) Trade receivables"
        ws[f"B{row}"] = "12"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"trade_receivables",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "trade_receivables")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Cash and cash equivalents
        ws[f"A{row}"] = "    (d) Cash and cash equivalents"
        ws[f"B{row}"] = "13"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"cash",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "cash")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Short-term loans and advances
        ws[f"A{row}"] = "    (e) Short-term loans and advances"
        ws[f"B{row}"] = "14"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"short_term_loans",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "short_term_loans")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Other current assets
        ws[f"A{row}"] = "    (f) Other current assets"
        ws[f"B{row}"] = "15"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"other_current_assets",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("bs", "other_current_assets")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # TOTAL ASSETS
        ws[f"A{row}"] = "TOTAL ASSETS"
        ws[f"A{row}"].font = Font(name="Arial", size=10, bold=True)
        ws[f"C{row}"] = f"=SUM(C{noca_start}:C{row-1})"
        ws[f"C{row}"].font = Font(name="Arial", size=10, bold=True)
        ws[f"C{row}"].number_format = currency_fmt
        ws[f"C{row}"].border = Border(
            top=Side(style="thin"),
            bottom=Side(style="double"),
            left=Side(style="thin"),
            right=Side(style="thin"),
        )
        if self._has_prior_year_data():
            ws[f"D{row}"] = f"=SUM(D{noca_start}:D{row-1})"
            ws[f"D{row}"].font = Font(name="Arial", size=10, bold=True)
            ws[f"D{row}"].number_format = currency_fmt
            ws[f"D{row}"].border = Border(
                top=Side(style="thin"),
                bottom=Side(style="double"),
                left=Side(style="thin"),
                right=Side(style="thin"),
            )
        row += 2

        # SIGN-OFF BLOCK
        row += 1
        ws[f"A{row}"] = "As per our Report of even date attached"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        row += 1

        ws[f"A{row}"] = "For [Firm Name]"
        ws[f"C{row}"] = "[Director Name]"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        ws[f"C{row}"].font = Font(name="Arial", size=9)
        row += 1

        ws[f"A{row}"] = "Chartered Accountants"
        ws[f"C{row}"] = "Director"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        ws[f"C{row}"].font = Font(name="Arial", size=9)
        row += 1

        ws[f"A{row}"] = "FRN: [FRN Number]"
        ws[f"C{row}"] = "DIN: [DIN Number]"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        ws[f"C{row}"].font = Font(name="Arial", size=9)
        row += 2

        ws[f"A{row}"] = "[Partner Name]"
        ws[f"C{row}"] = "[Director Name]"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        ws[f"C{row}"].font = Font(name="Arial", size=9)
        row += 1

        ws[f"A{row}"] = "Partner"
        ws[f"C{row}"] = "Director"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        ws[f"C{row}"].font = Font(name="Arial", size=9)
        row += 1

        ws[f"A{row}"] = "(Membership No: [Membership Number])"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        row += 1

        ws[f"A{row}"] = "UDIN: [UDIN]"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        row += 1

        ws[f"A{row}"] = "Place: [Place]"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        row += 1

        ws[f"A{row}"] = "Date: [Date]"
        ws[f"A{row}"].font = Font(name="Arial", size=9)

    def _create_pl_sheet(self):
        """
        Create Statement of Profit and Loss sheet with EXACT Schedule III (Division I) format.

        Structure:
        - Company name and CIN headers
        - Statement of Profit and Loss for the year ended [date]
        - Amount in Thousands INR note
        - Column headers: Particulars | Note | Current Year | Prior Year
        - All amounts are SUMIF formulas referencing TB_Tally sheet
        - Every SUMIF divided by 1000 for thousands display
        - Section headers with dark blue background
        - Total lines with borders and bold font
        """
        ws = self.workbook.create_sheet("P_L", 2)

        # Set column widths
        ws.column_dimensions["A"].width = 60
        ws.column_dimensions["B"].width = 8
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20

        # Currency format for amounts (thousands, no decimals)
        currency_fmt = "#,##0;(#,##0);-"

        # Row counter
        row = 1

        # HEADER BLOCK
        ws[f"A{row}"] = self.company_name
        ws[f"A{row}"].font = Font(name="Arial", size=12, bold=True)
        ws[f"A{row}"].alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        ws[f"A{row}"] = "CIN: [CIN Number]"
        ws[f"A{row}"].font = Font(name="Arial", size=10)
        row += 1

        ws[f"A{row}"] = "Registered Office: [Address]"
        ws[f"A{row}"].font = Font(name="Arial", size=10)
        row += 1

        # Title and date
        ws[f"A{row}"] = f"Statement of Profit and Loss for the year ended {self.as_on_date}"
        ws[f"A{row}"].font = Font(name="Arial", size=11, bold=True)
        row += 1

        ws[f"A{row}"] = "(Amount in Thousands INR)"
        ws[f"A{row}"].font = Font(name="Arial", size=9, italic=True)
        row += 2

        # COLUMN HEADERS
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

        ws[f"A{row}"] = "Particulars"
        ws[f"B{row}"] = "Note"
        ws[f"C{row}"] = f"Year ended {self.as_on_date}"
        ws[f"D{row}"] = f"Year ended {self.prior_year_date}"

        for col in ["A", "B", "C", "D"]:
            cell = ws[f"{col}{row}"]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        row += 1

        # REVENUE FROM OPERATIONS
        section_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
        section_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

        ws[f"A{row}"] = "I. Revenue from operations"
        ws[f"B{row}"] = "16"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"revenue_operations",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("pl", "revenue_operations")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # OTHER INCOME
        ws[f"A{row}"] = "II. Other Income"
        ws[f"B{row}"] = "17"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"other_income",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("pl", "other_income")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # TOTAL INCOME
        total_income_row = row
        ws[f"A{row}"] = "III. Total Income (I + II)"
        ws[f"C{row}"] = f"=C{row-2}+C{row-1}"
        ws[f"C{row}"].font = Font(name="Arial", size=10, bold=True)
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = f"=D{row-2}+D{row-1}"
            ws[f"D{row}"].font = Font(name="Arial", size=10, bold=True)
            ws[f"D{row}"].number_format = currency_fmt
        row += 2

        # EXPENSES SECTION
        ws[f"A{row}"] = "IV. Expenses:"
        ws[f"A{row}"].font = Font(name="Arial", size=10, bold=True)
        row += 1

        expenses_start = row

        # Cost of materials consumed
        ws[f"A{row}"] = "    Cost of materials consumed"
        ws[f"B{row}"] = "18"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"cost_materials",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("pl", "cost_materials")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Purchases of Stock-in-Trade
        ws[f"A{row}"] = "    Purchases of Stock-in-Trade"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"purchases_stock",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("pl", "purchases_stock")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Changes in inventories
        ws[f"A{row}"] = "    Changes in inventories of FG, WIP & SiT"
        ws[f"B{row}"] = "19"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"changes_inventory",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("pl", "changes_inventory")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Employee benefit expense
        ws[f"A{row}"] = "    Employee benefit expense"
        ws[f"B{row}"] = "20"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"employee_benefits",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("pl", "employee_benefits")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Finance costs
        ws[f"A{row}"] = "    Finance costs"
        ws[f"B{row}"] = "21"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"finance_costs",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("pl", "finance_costs")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Depreciation and amortization expense
        ws[f"A{row}"] = "    Depreciation and amortization expense"
        ws[f"B{row}"] = "9"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"depreciation",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("pl", "depreciation")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Other expenses
        ws[f"A{row}"] = "    Other expenses"
        ws[f"B{row}"] = "22"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"other_expenses",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("pl", "other_expenses")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # TOTAL EXPENSES
        total_expenses_row = row
        ws[f"A{row}"] = "    Total Expenses"
        ws[f"C{row}"] = f"=SUM(C{expenses_start}:C{row-1})"
        ws[f"C{row}"].font = Font(name="Arial", size=10, bold=True)
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = f"=SUM(D{expenses_start}:D{row-1})"
            ws[f"D{row}"].font = Font(name="Arial", size=10, bold=True)
            ws[f"D{row}"].number_format = currency_fmt
        row += 2

        # PROFIT BEFORE EXCEPTIONAL ITEMS AND TAX
        ws[f"A{row}"] = "V. Profit before exceptional and extraordinary items and tax (III - IV)"
        ws[f"C{row}"] = f"=C{total_income_row}-C{total_expenses_row}"
        ws[f"C{row}"].font = Font(name="Arial", size=10, bold=True)
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = f"=D{total_income_row}-D{total_expenses_row}"
            ws[f"D{row}"].font = Font(name="Arial", size=10, bold=True)
            ws[f"D{row}"].number_format = currency_fmt
        profit_before_exc_row = row
        row += 1

        # EXCEPTIONAL ITEMS
        ws[f"A{row}"] = "VI. Exceptional Items"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"exceptional_items",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = 0
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # PROFIT BEFORE EXTRAORDINARY ITEMS AND TAX
        profit_before_extra_row = row
        ws[f"A{row}"] = "VII. Profit before extraordinary items and tax (V - VI)"
        ws[f"C{row}"] = f"=C{profit_before_exc_row}-C{row-1}"
        ws[f"C{row}"].font = Font(name="Arial", size=10, bold=True)
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = f"=D{profit_before_exc_row}-D{row-1}"
            ws[f"D{row}"].font = Font(name="Arial", size=10, bold=True)
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # EXTRAORDINARY ITEMS
        ws[f"A{row}"] = "VIII. Extraordinary Items"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"extraordinary_items",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = 0
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # PROFIT BEFORE TAX
        profit_before_tax_row = row
        ws[f"A{row}"] = "IX. Profit before tax (VII - VIII)"
        ws[f"C{row}"] = f"=C{profit_before_extra_row}-C{row-1}"
        ws[f"C{row}"].font = Font(name="Arial", size=10, bold=True)
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = f"=D{profit_before_extra_row}-D{row-1}"
            ws[f"D{row}"].font = Font(name="Arial", size=10, bold=True)
            ws[f"D{row}"].number_format = currency_fmt
        row += 2

        # TAX EXPENSE
        ws[f"A{row}"] = "X. Tax expense:"
        row += 1

        tax_start = row

        # Current tax
        ws[f"A{row}"] = "    (1) Current tax"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"current_tax",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("pl", "current_tax")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # Deferred tax
        ws[f"A{row}"] = "    (2) Deferred tax"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"deferred_tax",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("pl", "deferred_tax")
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # PROFIT FROM CONTINUING OPERATIONS
        profit_cont_row = row
        ws[f"A{row}"] = "XI. Profit/(Loss) for the period from continuing operations (IX - X)"
        ws[f"C{row}"] = f"=C{profit_before_tax_row}-SUM(C{tax_start}:C{row-1})"
        ws[f"C{row}"].font = Font(name="Arial", size=10, bold=True)
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = f"=D{profit_before_tax_row}-SUM(D{tax_start}:D{row-1})"
            ws[f"D{row}"].font = Font(name="Arial", size=10, bold=True)
            ws[f"D{row}"].number_format = currency_fmt
        row += 1

        # PROFIT FROM DISCONTINUING OPERATIONS
        ws[f"A{row}"] = "XII. Profit/(Loss) from discontinuing operations"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"discontinued_ops",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = 0
            ws[f"D{row}"].number_format = currency_fmt
        disc_ops_row = row
        row += 1

        # TAX EXPENSE OF DISCONTINUING OPERATIONS
        ws[f"A{row}"] = "XIII. Tax expense of discontinuing operations"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"discontinued_tax",TB_Tally!H:H)/1000'
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = 0
            ws[f"D{row}"].number_format = currency_fmt
        disc_tax_row = row
        row += 1

        # PROFIT FROM DISCONTINUING OPERATIONS (NET)
        ws[f"A{row}"] = "XIV. Profit/(Loss) from Discontinuing operations (XII - XIII)"
        ws[f"C{row}"] = f"=C{disc_ops_row}-C{disc_tax_row}"
        ws[f"C{row}"].number_format = currency_fmt
        if self._has_prior_year_data():
            ws[f"D{row}"] = f"=D{disc_ops_row}-D{disc_tax_row}"
            ws[f"D{row}"].number_format = currency_fmt
        disc_net_row = row
        row += 1

        # NET PROFIT FOR THE PERIOD
        ws[f"A{row}"] = "XV. Profit/(Loss) for the period (XI + XIV)"
        ws[f"C{row}"] = f"=C{profit_cont_row}+C{disc_net_row}"
        ws[f"C{row}"].font = Font(name="Arial", size=10, bold=True)
        ws[f"C{row}"].number_format = currency_fmt
        ws[f"C{row}"].border = Border(
            top=Side(style="thin"),
            bottom=Side(style="double"),
            left=Side(style="thin"),
            right=Side(style="thin"),
        )
        if self._has_prior_year_data():
            ws[f"D{row}"] = f"=D{profit_cont_row}+D{disc_net_row}"
            ws[f"D{row}"].font = Font(name="Arial", size=10, bold=True)
            ws[f"D{row}"].number_format = currency_fmt
            ws[f"D{row}"].border = Border(
                top=Side(style="thin"),
                bottom=Side(style="double"),
                left=Side(style="thin"),
                right=Side(style="thin"),
            )
        row += 2

        # EARNINGS PER SHARE
        ws[f"A{row}"] = "XVI. Earnings per equity share:"
        row += 1

        # Basic EPS
        ws[f"A{row}"] = "    (1) Basic"
        ws[f"B{row}"] = "23"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"eps_basic",TB_Tally!H:H)'
        ws[f"C{row}"].number_format = "#,##0.00"
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("pl", "eps_basic")
            ws[f"D{row}"].number_format = "#,##0.00"
        row += 1

        # Diluted EPS
        ws[f"A{row}"] = "    (2) Diluted"
        ws[f"C{row}"] = '=SUMIF(TB_Tally!D:D,"eps_diluted",TB_Tally!H:H)'
        ws[f"C{row}"].number_format = "#,##0.00"
        if self._has_prior_year_data():
            ws[f"D{row}"] = self._get_prior_year_value("pl", "eps_diluted")
            ws[f"D{row}"].number_format = "#,##0.00"
        row += 2

        # SIGN-OFF BLOCK
        row += 1
        ws[f"A{row}"] = "As per our Report of even date attached"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        row += 1

        ws[f"A{row}"] = "For [Firm Name]"
        ws[f"C{row}"] = "[Director Name]"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        ws[f"C{row}"].font = Font(name="Arial", size=9)
        row += 1

        ws[f"A{row}"] = "Chartered Accountants"
        ws[f"C{row}"] = "Director"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        ws[f"C{row}"].font = Font(name="Arial", size=9)
        row += 1

        ws[f"A{row}"] = "FRN: [FRN Number]"
        ws[f"C{row}"] = "DIN: [DIN Number]"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        ws[f"C{row}"].font = Font(name="Arial", size=9)
        row += 2

        ws[f"A{row}"] = "[Partner Name]"
        ws[f"C{row}"] = "[Director Name]"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        ws[f"C{row}"].font = Font(name="Arial", size=9)
        row += 1

        ws[f"A{row}"] = "Partner"
        ws[f"C{row}"] = "Director"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        ws[f"C{row}"].font = Font(name="Arial", size=9)
        row += 1

        ws[f"A{row}"] = "(Membership No: [Membership Number])"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        row += 1

        ws[f"A{row}"] = "UDIN: [UDIN]"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        row += 1

        ws[f"A{row}"] = "Place: [Place]"
        ws[f"A{row}"].font = Font(name="Arial", size=9)
        row += 1

        ws[f"A{row}"] = "Date: [Date]"
        ws[f"A{row}"].font = Font(name="Arial", size=9)

    def _create_recon_sheet(self):
        """
        Create Reconciliation sheet with all 8 verification checks.
        """
        ws = self.workbook.create_sheet("Recon", 3)

        # Title
        ws["A1"] = "RECONCILIATION & VERIFICATION CHECKS"
        ws["A1"].font = Font(name="Arial", size=12, bold=True)

        # Run reconciliation
        adjustments = YearEndAdjustments(self.classifier)
        reconciler = ReconciliationEngine(self.classifier.extractor, self.classifier, adjustments)
        checks = reconciler.run_all_checks()

        # Headers
        headers = ["Check #", "Check Name", "Expected", "Actual", "Difference", "Status", "Details"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.border

        # Data rows
        row_num = 4
        for idx, check in enumerate(checks, 1):
            ws.cell(row=row_num, column=1).value = idx
            ws.cell(row=row_num, column=2).value = check.check_name
            ws.cell(row=row_num, column=3).value = float(check.expected) / 1000
            ws.cell(row=row_num, column=4).value = float(check.actual) / 1000
            ws.cell(row=row_num, column=5).value = float(check.difference) / 1000
            ws.cell(row=row_num, column=6).value = check.status
            ws.cell(row=row_num, column=7).value = check.details

            # Format row
            for col_num in range(1, 8):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = self.data_font
                cell.border = self.border
                if col_num in (3, 4, 5):  # Number columns
                    cell.number_format = "#,##0.00"
                # Status color coding
                if col_num == 6:
                    if check.status == "PASS":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif check.status == "FAIL":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    else:  # WARNING
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            row_num += 1

        # Set column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 40

    def _create_notes_sheet(self):
        """Create Notes to Accounts sheet with Notes 1-24."""
        ws = self.workbook.create_sheet("Notes", 4)

        # Title
        ws["A1"] = "NOTES TO ACCOUNTS"
        ws["A1"].font = Font(name="Arial", size=12, bold=True)
        ws["A2"] = f"for the year ended {self.as_on_date}"
        ws["A2"].font = Font(name="Arial", size=10, italic=True)

        row = 4

        # Note 1: Significant Accounting Policies
        ws[f"A{row}"] = "Note 1: SIGNIFICANT ACCOUNTING POLICIES"
        ws[f"A{row}"].font = self.subheader_font
        row += 1
        ws[f"A{row}"] = "a) Basis of Preparation: The financial statements are prepared in accordance with Indian GAAP (IGAAP), using the historical cost convention on an accrual basis. Revenue is recognized as per AS-9."
        row += 1
        ws[f"A{row}"] = "b) Inventories: Valued at lower of cost and net realizable value as per AS-2."
        row += 1
        ws[f"A{row}"] = "c) Fixed Assets: Stated at cost less accumulated depreciation, computed using written down value method as per Schedule II of the Companies Act."
        row += 1
        ws[f"A{row}"] = "d) Employee Benefits: As per AS-15. Provision made for gratuity and other benefits."
        row += 1
        ws[f"A{row}"] = "e) Taxation: Current tax provided at applicable rates. Deferred tax recognized on timing differences as per AS-22."
        row += 1
        ws[f"A{row}"] = "f) Borrowing Costs: As per AS-16, directly attributable borrowing costs capitalized. Others expensed."
        row += 2

        # Note 2: Share Capital
        ws[f"A{row}"] = "Note 2: SHARE CAPITAL"
        ws[f"A{row}"].font = self.subheader_font
        row += 1
        ws[f"A{row}"] = "Authorized Capital"
        ws[f"B{row}"] = "Amount (₹)"
        row += 1
        ws[f"A{row}"] = "Issued, Subscribed & Paid-up"
        ws[f"B{row}"] = self._build_sumproduct_formula("share_capital")
        ws[f"B{row}"].number_format = self.thousands_format
        row += 2

        # Note 3: Reserves & Surplus
        ws[f"A{row}"] = "Note 3: RESERVES & SURPLUS"
        ws[f"A{row}"].font = self.subheader_font
        row += 1
        ws[f"A{row}"] = "Opening Balance"
        row += 1
        ws[f"A{row}"] = "Add: Profit for the Year"
        row += 1
        ws[f"A{row}"] = "Closing Balance"
        ws[f"B{row}"] = self._build_sumproduct_formula("other_equity")
        ws[f"B{row}"].number_format = self.thousands_format
        row += 2

        # Note 4: Long-term Borrowings
        ws[f"A{row}"] = "Note 4: LONG-TERM BORROWINGS"
        ws[f"A{row}"].font = self.subheader_font
        row += 1
        # List borrowing ledgers from classified data
        borrowings = [c for c in self.classified_ledgers if c.schedule_iii_line == "borrowings"]
        for ledger in borrowings[:10]:  # Limit to 10
            ws[f"A{row}"] = ledger.name
            ws[f"B{row}"] = float(ledger.display_amount) / 1000
            ws[f"B{row}"].number_format = self.thousands_format
            row += 1
        row += 1

        # Note 5-9: Placeholders
        for note_num in range(5, 10):
            ws[f"A{row}"] = f"Note {note_num}: [Details to be populated]"
            ws[f"A{row}"].font = self.subheader_font
            row += 2

        # Note 10: PPE
        ws[f"A{row}"] = "Note 10: PROPERTY, PLANT & EQUIPMENT"
        ws[f"A{row}"].font = self.subheader_font
        row += 1
        ws[f"A{row}"] = "Asset Name"
        ws[f"B{row}"] = "Opening WDV"
        ws[f"C{row}"] = "Additions"
        ws[f"D{row}"] = "Depreciation"
        ws[f"E{row}"] = "Closing WDV"
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}{row}"].font = self.subheader_font
        row += 1

        ppe_items = [c for c in self.classified_ledgers if c.schedule_iii_line == "property_plant_equipment"]
        for ledger in ppe_items:
            ws[f"A{row}"] = ledger.name
            ws[f"B{row}"] = float(ledger.opening_balance) / 1000
            ws[f"E{row}"] = float(ledger.closing_balance) / 1000
            for col in ["B", "C", "D", "E"]:
                ws[f"{col}{row}"].number_format = self.thousands_format
            row += 1
        row += 2

        # Notes 11-23: More placeholders
        for note_num in range(11, 24):
            ws[f"A{row}"] = f"Note {note_num}: [Details to be populated]"
            ws[f"A{row}"].font = self.subheader_font
            row += 2

        # Note 24: Related Party
        ws[f"A{row}"] = "Note 24: RELATED PARTY DISCLOSURES"
        ws[f"A{row}"].font = self.subheader_font
        row += 1
        ws[f"A{row}"] = "Related Party"
        ws[f"B{row}"] = "Nature of Transaction"
        ws[f"C{row}"] = "Amount (₹)"
        row += 1
        ws[f"A{row}"] = "[To be populated with related party transactions]"

        # Set column widths
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15

    def _create_depreciation_sheet(self):
        """Create Depreciation Schedule sheet."""
        ws = self.workbook.create_sheet("Dep_Schedule", 5)

        ws["A1"] = "DEPRECIATION SCHEDULE"
        ws["A1"].font = Font(name="Arial", size=12, bold=True)
        ws["A2"] = f"as at {self.as_on_date}"
        ws["A2"].font = Font(name="Arial", size=10, italic=True)

        row = 4

        # Headers
        headers = ["Asset Name", "Opening WDV", "Additions", "Disposals", "Total", "Depreciation Rate %", "Depreciation", "Closing WDV"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

        # Standard depreciation rates
        dep_rates = {
            "building": 1.58, "plant": 4.75, "machinery": 4.75, "furniture": 9.5,
            "computer": 31.67, "vehicle": 11.88, "motor": 9.5
        }

        ppe_items = [c for c in self.classified_ledgers if c.schedule_iii_line == "property_plant_equipment"]
        for ledger in sorted(ppe_items, key=lambda x: x.name):
            ws.cell(row=row, column=1).value = ledger.name
            ws.cell(row=row, column=2).value = float(ledger.opening_balance) / 1000
            ws.cell(row=row, column=3).value = 0  # Additions placeholder
            ws.cell(row=row, column=4).value = 0  # Disposals placeholder
            ws.cell(row=row, column=5).value = f"=B{row}+C{row}-D{row}"

            # Determine rate
            rate = 4.75  # Default
            for key, val in dep_rates.items():
                if key.lower() in ledger.name.lower():
                    rate = val
                    break
            ws.cell(row=row, column=6).value = rate

            ws.cell(row=row, column=7).value = float(ledger.opening_balance - ledger.closing_balance) / 1000
            ws.cell(row=row, column=8).value = float(ledger.closing_balance) / 1000

            for col in range(2, 9):
                ws.cell(row=row, column=col).number_format = self.thousands_format

            row += 1

        # Set column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_ratios_sheet(self):
        """Create Financial Ratios sheet."""
        ws = self.workbook.create_sheet("Ratios", 6)

        ws["A1"] = "FINANCIAL RATIOS"
        ws["A1"].font = Font(name="Arial", size=12, bold=True)
        ws["A2"] = f"for the year ended {self.as_on_date}"

        row = 4

        # Headers
        headers = ["Ratio", "Numerator", "Denominator", "Current Year", "Previous Year", "Variance %"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = self.subheader_font
            cell.fill = self.subheader_fill

        row += 1

        # Ratios (with formula references to B/S and P&L sheets)
        ratios = [
            ("Current Ratio", "Current Assets", "Current Liabilities", "='Balance Sheet'!B40/'Balance Sheet'!B32"),
            ("Debt-Equity Ratio", "Total Liabilities", "Total Equity", "=('Balance Sheet'!B26+'Balance Sheet'!B32)/'Balance Sheet'!B16"),
            ("Gross Profit Ratio", "Gross Profit", "Revenue", "=(('P_L'!B7-'P_L'!B12)/('P_L'!B7))*100"),
            ("Net Profit Ratio", "Net Profit", "Revenue", "=('P_L'!B23/'P_L'!B7)*100"),
            ("ROE", "Net Profit", "Average Equity", "='P_L'!B23/'Balance Sheet'!B16"),
            ("ROA", "Net Profit", "Average Assets", "='P_L'!B23/('Balance Sheet'!B40)"),
            ("Inventory Turnover", "Cost of Goods Sold", "Avg Inventory", "='P_L'!B12/('Balance Sheet'!B34)"),
            ("Trade Receivables Turnover", "Revenue", "Avg Trade Receivables", "='P_L'!B7/('Balance Sheet'!B35)"),
            ("Trade Payables Turnover", "Cost of Goods Sold", "Avg Trade Payables", "='P_L'!B12/('Balance Sheet'!B27)"),
            ("Net Capital Turnover", "Revenue", "Net Capital Employed", "='P_L'!B7/('Balance Sheet'!B16)"),
            ("Debtors Collection Period", "Days", "TR Turnover", "=365/('Ratios'!D8)"),
        ]

        for ratio_name, numerator, denominator, formula in ratios:
            ws.cell(row=row, column=1).value = ratio_name
            ws.cell(row=row, column=2).value = numerator
            ws.cell(row=row, column=3).value = denominator
            ws.cell(row=row, column=4).value = formula
            row += 1

        # Set column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15

    def _create_tr_ageing_sheet(self):
        """Create Trade Receivables Ageing sheet."""
        ws = self.workbook.create_sheet("TR_Ageing", 7)

        ws["A1"] = "TRADE RECEIVABLES AGEING"
        ws["A1"].font = Font(name="Arial", size=12, bold=True)

        row = 3

        # Headers
        headers = ["Particulars", "Undisputed - Considered Good", "Undisputed - Doubtful", "Disputed - Considered Good", "Disputed - Doubtful"]
        buckets = ["Not Due", "< 6 months", "6 months - 1 year", "1 - 2 years", "2 - 3 years", "> 3 years"]

        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row, column=col_num).value = header
            ws.cell(row=row, column=col_num).font = self.subheader_font

        row += 1

        # Buckets
        tr_total = 0
        for bucket in buckets:
            ws.cell(row=row, column=1).value = bucket
            # Placeholder formulas - would require detailed ageing data
            ws.cell(row=row, column=2).value = 0
            row += 1

        ws.cell(row=row, column=1).value = "TOTAL"
        ws.cell(row=row, column=1).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=row, column=2).value = f"=SUM(B5:B10)"
        ws.cell(row=row, column=2).font = Font(name="Arial", size=10, bold=True)

        # Set column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _create_tp_ageing_sheet(self):
        """Create Trade Payables Ageing sheet."""
        ws = self.workbook.create_sheet("TP_Ageing", 8)

        ws["A1"] = "TRADE PAYABLES AGEING"
        ws["A1"].font = Font(name="Arial", size=12, bold=True)
        ws["A2"] = "(Including MSME)"

        row = 4

        # Headers
        headers = ["Particulars", "MSME", "Others", "Total"]
        buckets = ["Not Due", "< 1 year", "1 - 2 years", "2 - 3 years", "> 3 years"]

        for col_num, header in enumerate(headers, 1):
            ws.cell(row=row, column=col_num).value = header
            ws.cell(row=row, column=col_num).font = self.subheader_font

        row += 1

        # Buckets with formulas
        for bucket in buckets:
            ws.cell(row=row, column=1).value = bucket
            ws.cell(row=row, column=2).value = self._build_sumproduct_formula_with_sub("trade_payables", "msme")
            ws.cell(row=row, column=3).value = self._build_sumproduct_formula_with_sub("trade_payables", "others")
            ws.cell(row=row, column=4).value = f"=B{row}+C{row}"
            for col in range(2, 5):
                ws.cell(row=row, column=col).number_format = self.thousands_format
            row += 1

        ws.cell(row=row, column=1).value = "TOTAL"
        ws.cell(row=row, column=1).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=row, column=2).value = f"=SUM(B5:B9)"
        ws.cell(row=row, column=3).value = f"=SUM(C5:C9)"
        ws.cell(row=row, column=4).value = f"=SUM(D5:D9)"

        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _create_cfs_sheet(self):
        """Create Cash Flow Statement sheet (Indirect Method)."""
        ws = self.workbook.create_sheet("CFS", 9)

        ws["A1"] = self.company_name
        ws["A1"].font = Font(name="Arial", size=12, bold=True)
        ws["A2"] = "CASH FLOW STATEMENT (INDIRECT METHOD)"
        ws["A2"].font = Font(name="Arial", size=11, bold=True)
        ws["A3"] = f"for the year ended {self.as_on_date}"

        row = 5

        # Operating Activities
        ws[f"A{row}"] = "A. CASH FLOWS FROM OPERATING ACTIVITIES"
        ws[f"A{row}"].font = self.subheader_font
        row += 1

        ws[f"A{row}"] = "Profit After Tax"
        ws[f"B{row}"] = "='P_L'!B23"
        row += 1

        ws[f"A{row}"] = "Add: Depreciation"
        ws[f"B{row}"] = "='P_L'!B15"
        row += 1

        ws[f"A{row}"] = "Add: Finance Costs"
        ws[f"B{row}"] = "='P_L'!B18"
        row += 1

        ws[f"A{row}"] = "Less: Other Income"
        ws[f"B{row}"] = "=-'P_L'!B8"
        row += 1

        ws[f"A{row}"] = "Operating Profit Before WC Changes"
        ws[f"B{row}"] = f"=SUM(B{row-4}:B{row-1})"
        ws[f"B{row}"].font = Font(name="Arial", size=10, bold=True)
        op_profit_row = row
        row += 1

        ws[f"A{row}"] = "Working Capital Changes:"
        row += 1

        ws[f"A{row}"] = "  (Increase)/Decrease in Receivables"
        ws[f"B{row}"] = "='Balance Sheet'!B35-'Balance Sheet'!C35"
        row += 1

        ws[f"A{row}"] = "  (Increase)/Decrease in Inventory"
        ws[f"B{row}"] = "='Balance Sheet'!B34-'Balance Sheet'!C34"
        row += 1

        ws[f"A{row}"] = "  Increase/(Decrease) in Payables"
        ws[f"B{row}"] = "='Balance Sheet'!B27-'Balance Sheet'!C27"
        row += 1

        ws[f"A{row}"] = "Net Cash from Operating Activities"
        ws[f"B{row}"] = f"=B{op_profit_row}+SUM(B{op_profit_row+2}:B{row-1})"
        ws[f"B{row}"].font = Font(name="Arial", size=10, bold=True)
        net_op_row = row
        row += 2

        # Investing Activities
        ws[f"A{row}"] = "B. CASH FLOWS FROM INVESTING ACTIVITIES"
        ws[f"A{row}"].font = self.subheader_font
        row += 1

        ws[f"A{row}"] = "Purchase of Fixed Assets"
        ws[f"B{row}"] = 0
        row += 1

        ws[f"A{row}"] = "Sale of Fixed Assets"
        ws[f"B{row}"] = 0
        row += 1

        ws[f"A{row}"] = "Net Cash from Investing Activities"
        ws[f"B{row}"] = f"=SUM(B{row-2}:B{row-1})"
        ws[f"B{row}"].font = Font(name="Arial", size=10, bold=True)
        net_inv_row = row
        row += 2

        # Financing Activities
        ws[f"A{row}"] = "C. CASH FLOWS FROM FINANCING ACTIVITIES"
        ws[f"A{row}"].font = self.subheader_font
        row += 1

        ws[f"A{row}"] = "Proceeds from Borrowings"
        ws[f"B{row}"] = "='Balance Sheet'!B26-'Balance Sheet'!C26"
        row += 1

        ws[f"A{row}"] = "Repayment of Borrowings"
        ws[f"B{row}"] = 0
        row += 1

        ws[f"A{row}"] = "Finance Cost Paid"
        ws[f"B{row}"] = "-'P_L'!B18"
        row += 1

        ws[f"A{row}"] = "Dividend Paid"
        ws[f"B{row}"] = 0
        row += 1

        ws[f"A{row}"] = "Net Cash from Financing Activities"
        ws[f"B{row}"] = f"=SUM(B{row-4}:B{row-1})"
        ws[f"B{row}"].font = Font(name="Arial", size=10, bold=True)
        net_fin_row = row
        row += 2

        # Net Cash Movement
        ws[f"A{row}"] = "Net Increase/(Decrease) in Cash"
        ws[f"B{row}"] = f"=B{net_op_row}+B{net_inv_row}+B{net_fin_row}"
        ws[f"B{row}"].font = Font(name="Arial", size=10, bold=True)
        net_move_row = row
        row += 1

        ws[f"A{row}"] = "Opening Cash Balance"
        ws[f"B{row}"] = "='Balance Sheet'!C36"
        row += 1

        ws[f"A{row}"] = "Closing Cash Balance"
        ws[f"B{row}"] = f"=B{net_move_row}+B{row-1}"
        ws[f"B{row}"].font = Font(name="Arial", size=10, bold=True)

        # Set column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20

    def _build_sumproduct_formula(self, schedule_line: str) -> str:
        """
        Build SUMPRODUCT formula to sum all amounts for a specific Schedule III line.

        Formula: =SUMPRODUCT((TB_Tally!D:D="line_item")*(TB_Tally!H:H))/1000

        Args:
            schedule_line: Schedule III line item code

        Returns:
            Excel formula string
        """
        # For now, use a simpler SUMIF approach
        # This will sum column H (Display Amount) where column D (Schedule III Line) matches
        return f'=SUMIF(TB_Tally!D:D,"{schedule_line}",TB_Tally!H:H)'

    def _build_sumproduct_formula_with_sub(self, schedule_line: str, schedule_sub: str) -> str:
        """
        Build SUMPRODUCT formula for line + sub combination.

        Formula: =SUMPRODUCT((TB_Tally!D:D="line")*(TB_Tally!E:E="sub")*(TB_Tally!H:H))

        Args:
            schedule_line: Schedule III line item
            schedule_sub: Schedule III sub-item

        Returns:
            Excel formula string
        """
        # Using SUMIFS for two criteria
        return f'=SUMIFS(TB_Tally!H:H,TB_Tally!D:D,"{schedule_line}",TB_Tally!E:E,"{schedule_sub}")'


def generate_financial_statements(db_path: str, output_path: str, company_name: str = "ROHIT PHARMA"):
    """
    Main entry point to generate financial statements.

    Steps:
    1. Extract data from Tally SQLite
    2. Classify to Schedule III
    3. Add adjustments (if any)
    4. Generate Excel workbook

    Args:
        db_path: Path to Tally SQLite database
        output_path: Path to save output Excel file
        company_name: Company name for headers

    Returns:
        Path to generated Excel file
    """
    logger.info(f"Starting financial statement generation for {company_name}")

    # Layer 1: Extract
    extractor = TallyDataExtractor(db_path)
    extractor.connect()
    extractor.load_metadata()
    extractor.build_group_hierarchy()
    tb = extractor.extract_trial_balance()
    logger.info(f"Extracted trial balance with {tb.ledger_count} ledgers")

    # Layer 2: Classify
    classifier = ScheduleIIIClassifier(extractor)
    classified = classifier.classify_all()
    logger.info(f"Classified {len(classified)} ledgers to Schedule III")

    # Layer 3: Adjustments
    adjustments = YearEndAdjustments(classifier)
    logger.info("Adjustment framework ready")

    # Layer 4: Generate Excel
    generator = FinancialStatementsExcelGenerator(
        classifier, adjustments, company_name=company_name
    )
    output = generator.generate(output_path)

    extractor.disconnect()
    logger.info(f"Financial statements generated successfully: {output}")
    return output


if __name__ == "__main__":
    # Example usage
    import sys
    sys.path.insert(0, "/sessions/affectionate-stoic-cannon/mnt/Tally Automation/slv_app")

    db_path = "/sessions/affectionate-stoic-cannon/mnt/Tally Automation/slv_app/tally_data.db"
    output_path = "/sessions/affectionate-stoic-cannon/mnt/Tally Automation/slv_app/Schedule_III_ROHIT_PHARMA.xlsx"

    generate_financial_statements(db_path, output_path, company_name="ROHIT PHARMA")
    print(f"Excel file generated: {output_path}")
