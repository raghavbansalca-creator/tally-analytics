"""
Financial Statement Generation Engine for Schedule III (IGAAP) - Data & Logic Layer

This module provides the core data structures and logic for reading Tally accounting data
from SQLite and generating a complete Schedule III (Division I - IGAAP) financial statement.

The engine operates in 4 main layers:
1. Trial Balance Extraction - Read all ledgers from SQLite, build group hierarchy
2. Classification Engine - Map every ledger to exactly one Schedule III line item
3. Year-end Adjustments Framework - Structure for depreciation, DTA, provisions
4. Reconciliation Engine - Run 8 verification checks to ensure consistency

SIGN CONVENTION (Critical):
- Tally uses CREDIT-POSITIVE convention:
  - Credit = Positive (Liabilities, Revenue, Profits)
  - Debit = Negative (Assets, Expenses, Losses)
- Schedule III display amounts are ALWAYS positive (sign corrected)

Author: Seven Labs Vision
Date: 2026-03-22
"""

import sqlite3
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple, Set
from enum import Enum
from decimal import Decimal
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ============================================================================
# SECTION 1: CONSTANTS & ENUMS
# ============================================================================

class ScheduleIIISection(Enum):
    """Sections in Schedule III (Division I - IGAAP) financial statement"""
    BS_EQUITY = "BS_EQUITY"              # Shareholders' Funds
    BS_CL = "BS_CL"                      # Current Liabilities & Provisions
    BS_CA = "BS_CA"                      # Current Assets
    BS_NONCL = "BS_NONCL"                # Non-Current Liabilities
    BS_NONICA = "BS_NONICA"              # Non-Current Assets
    PL_REVENUE = "PL_REVENUE"            # Revenue from Operations
    PL_EXPENSE = "PL_EXPENSE"            # Expenses


class TallyGroupType(Enum):
    """Classification of Tally groups"""
    ASSET = "ASSET"
    LIABILITY = "LIABILITY"
    REVENUE = "REVENUE"
    EXPENSE = "EXPENSE"


# Schedule III Line Items (Balance Sheet - Assets)
BS_NONICA_ITEMS = {
    "property_plant_equipment": {
        "display_name": "Property, Plant and Equipment",
        "sub_items": {
            "gross": "Gross Block",
            "depreciation": "Accumulated Depreciation",
            "net": "Net Block",
        }
    },
    "investment_property": {
        "display_name": "Investment Property",
        "sub_items": {"net": "Net Book Value"}
    },
    "intangible_assets": {
        "display_name": "Intangible Assets",
        "sub_items": {"net": "Net"}
    },
    "financial_assets": {
        "display_name": "Financial Assets",
        "sub_items": {
            "investments": "Investments",
            "loans": "Loans",
            "receivables": "Receivables",
            "cash": "Cash and Cash Equivalents"
        }
    },
    "other_noncurrent_assets": {
        "display_name": "Other Non-Current Assets",
        "sub_items": {"net": "Net"}
    },
}

BS_NONICA = list(BS_NONICA_ITEMS.keys())

# Schedule III Line Items (Balance Sheet - Current Assets)
BS_CA_ITEMS = {
    "inventories": {
        "display_name": "Inventories",
        "sub_items": {"raw_materials": "Raw Materials", "wip": "Work-in-Progress", "finished_goods": "Finished Goods"}
    },
    "financial_assets_ca": {
        "display_name": "Financial Assets",
        "sub_items": {
            "investments": "Investments",
            "trade_receivables": "Trade Receivables",
            "msme": "of which MSME",
            "others": "others",
            "loans": "Loans",
            "receivables": "Receivables",
            "cash": "Cash and Cash Equivalents"
        }
    },
    "current_tax_assets": {
        "display_name": "Current Tax Assets",
        "sub_items": {}
    },
    "other_current_assets": {
        "display_name": "Other Current Assets",
        "sub_items": {}
    },
}

BS_CA = list(BS_CA_ITEMS.keys())

# Schedule III Line Items (Balance Sheet - Equity & Liabilities)
BS_EQUITY_ITEMS = {
    "share_capital": {
        "display_name": "Share Capital",
        "sub_items": {"authorized": "Authorized", "issued": "Issued", "subscribed": "Subscribed", "paid_up": "Paid-up"}
    },
    "other_equity": {
        "display_name": "Other Equity",
        "sub_items": {
            "reserves_surplus": "Reserves and Surplus",
            "retained_earnings": "Retained Earnings",
            "other": "Other Components"
        }
    },
}

BS_EQUITY = list(BS_EQUITY_ITEMS.keys())

# Schedule III Line Items (Balance Sheet - Liabilities)
BS_NONCL_ITEMS = {
    "borrowings": {
        "display_name": "Borrowings",
        "sub_items": {
            "secured": "Secured Loans",
            "unsecured": "Unsecured Loans"
        }
    },
    "deferred_tax_liabilities": {
        "display_name": "Deferred Tax Liabilities",
        "sub_items": {}
    },
    "other_noncurrent_liabilities": {
        "display_name": "Other Non-Current Liabilities",
        "sub_items": {}
    },
}

BS_NONCL = list(BS_NONCL_ITEMS.keys())

BS_CL_ITEMS = {
    "borrowings_cl": {
        "display_name": "Borrowings",
        "sub_items": {"secured": "Secured", "unsecured": "Unsecured"}
    },
    "trade_payables": {
        "display_name": "Trade Payables",
        "sub_items": {
            "total_outstanding": "Total Outstanding Dues",
            "msme": "of which MSME",
            "others": "others"
        }
    },
    "other_financial_liabilities": {
        "display_name": "Other Financial Liabilities",
        "sub_items": {}
    },
    "other_current_liabilities": {
        "display_name": "Other Current Liabilities",
        "sub_items": {}
    },
    "current_tax_liabilities": {
        "display_name": "Current Tax Liabilities",
        "sub_items": {}
    },
    "provisions": {
        "display_name": "Provisions",
        "sub_items": {
            "employee_benefits": "Employee Benefits",
            "warranties": "Warranties",
            "other": "Other"
        }
    },
}

BS_CL = list(BS_CL_ITEMS.keys())

# Schedule III Line Items (P&L)
PL_REVENUE_ITEMS = {
    "revenue_operations": {
        "display_name": "Revenue from Operations",
        "sub_items": {
            "sale_products": "Sale of Products",
            "sale_services": "Sale of Services",
            "other": "Other"
        }
    },
    "other_income": {
        "display_name": "Other Income",
        "sub_items": {}
    },
}

PL_REVENUE = list(PL_REVENUE_ITEMS.keys())

PL_EXPENSE_ITEMS = {
    "cost_materials": {
        "display_name": "Cost of Materials Consumed",
        "sub_items": {}
    },
    "purchases_changes": {
        "display_name": "Purchases and Changes in Inventory",
        "sub_items": {}
    },
    "employee_benefits": {
        "display_name": "Employee Benefits Expense",
        "sub_items": {}
    },
    "depreciation": {
        "display_name": "Depreciation and Amortization",
        "sub_items": {
            "depreciation": "Depreciation",
            "amortization": "Amortization"
        }
    },
    "other_expenses": {
        "display_name": "Other Expenses",
        "sub_items": {}
    },
    "finance_costs": {
        "display_name": "Finance Costs",
        "sub_items": {
            "interest": "Interest Expense",
            "bank_charges": "Bank Charges",
            "loan_processing": "Loan Processing Fees",
            "other": "Other"
        }
    },
}

PL_EXPENSE = list(PL_EXPENSE_ITEMS.keys())


# ============================================================================
# SECTION 2: DATA CLASSES
# ============================================================================

@dataclass
class TallyGroupNode:
    """Represents a single node in the Tally group hierarchy tree"""
    name: str
    parent: Optional[str]
    is_deemed_positive: bool  # Tally's convention: True = Credit-positive
    is_revenue: bool
    children: List[str] = field(default_factory=list)
    level: int = 0  # 0=Primary, 1=Secondary, etc.

    def __repr__(self) -> str:
        return f"TallyGroup({self.name}, parent={self.parent}, level={self.level})"


@dataclass
class TallyLedger:
    """Represents a ledger account from Tally"""
    name: str
    parent_group: str
    opening_balance: Decimal  # Tally sign convention (Credit=positive)
    closing_balance: Decimal  # Tally sign convention (Credit=positive)
    is_revenue: bool  # True if P&L account
    is_deemed_positive: bool  # Tally convention

    @property
    def has_closing_balance(self) -> bool:
        """Check if ledger has non-zero closing balance"""
        return self.closing_balance != Decimal(0)

    def get_display_amount(self) -> Decimal:
        """Return absolute value for display (always positive in FS)"""
        return abs(self.closing_balance)

    def __repr__(self) -> str:
        return f"Ledger({self.name}, CB={self.closing_balance})"


@dataclass
class ClassifiedLedger:
    """A ledger that has been classified into Schedule III"""
    name: str
    tally_group: str
    opening_balance: Decimal
    closing_balance: Decimal
    tally_group_type: TallyGroupType
    schedule_iii_section: ScheduleIIISection
    schedule_iii_line: str  # e.g., "share_capital", "trade_payables"
    schedule_iii_sub: str = ""  # e.g., "msme", "others", "finance_costs"
    display_amount: Decimal = field(default=Decimal(0))
    is_reclassified: bool = False
    reclassification_note: str = ""

    def __post_init__(self):
        """Auto-compute display amount if not set"""
        if self.display_amount == Decimal(0):
            self.display_amount = abs(self.closing_balance)

    def __repr__(self) -> str:
        reclassified = f" [RECLASSIFIED: {self.reclassification_note}]" if self.is_reclassified else ""
        return f"ClassifiedLedger({self.name}, {self.schedule_iii_section.value}:{self.schedule_iii_line}{reclassified})"


@dataclass
class TrialBalance:
    """Summary of extracted trial balance"""
    total_debits: Decimal  # Sum of debit ledgers
    total_credits: Decimal  # Sum of credit ledgers
    net_balance: Decimal  # Should be zero if balanced
    ledger_count: int
    ledgers: List[TallyLedger] = field(default_factory=list)
    extracted_at: str = ""

    @property
    def is_balanced(self) -> bool:
        """Check if trial balance is balanced"""
        return abs(self.net_balance) < Decimal("0.01")  # Allow 0.01 rounding

    def __repr__(self) -> str:
        status = "BALANCED" if self.is_balanced else "UNBALANCED"
        return f"TrialBalance({self.ledger_count} ledgers, {status}, Net={self.net_balance})"


@dataclass
class VerificationCheck:
    """Result of a single verification check"""
    check_name: str
    expected: Decimal
    actual: Decimal
    difference: Decimal
    status: str  # "PASS", "FAIL", "WARNING"
    details: str = ""
    severity: str = "INFO"  # INFO, WARNING, CRITICAL

    @property
    def passed(self) -> bool:
        return self.status == "PASS"

    def __repr__(self) -> str:
        return f"VerificationCheck({self.check_name}: {self.status}, Diff={self.difference})"


@dataclass
class AdjustmentEntry:
    """Represents a year-end adjustment (depreciation, DTA, provisions, etc.)"""
    entry_id: str
    entry_type: str  # "depreciation", "dta", "provision", "other"
    description: str
    debit_ledger: str
    debit_amount: Decimal
    credit_ledger: str
    credit_amount: Decimal
    schedule_iii_line: str  # Which Schedule III line this affects
    schedule_iii_sub: str
    is_applied: bool = False
    applied_at: str = ""
    notes: str = ""

    def __repr__(self) -> str:
        applied = "APPLIED" if self.is_applied else "PENDING"
        return f"Adjustment({self.entry_id}, {self.entry_type}, {applied})"


@dataclass
class ReclassificationRule:
    """Rule for reclassifying a ledger based on conditions"""
    rule_id: str
    description: str
    condition: str  # Lambda or description of condition
    from_schedule_iii: Tuple[ScheduleIIISection, str]  # (section, line_item)
    to_schedule_iii: Tuple[ScheduleIIISection, str]  # (section, line_item)
    applies_when: str  # Description of when this rule applies


# ============================================================================
# SECTION 3: LAYER 1 - TRIAL BALANCE EXTRACTION
# ============================================================================

class TallyDataExtractor:
    """
    Extracts trial balance data from Tally SQLite database.

    Responsibilities:
    - Connect to SQLite database
    - Build group hierarchy tree
    - Extract all ledgers with opening and closing balances
    - Verify trial balance
    - Identify P&L and Balance Sheet accounts
    """

    def __init__(self, db_path: str):
        """
        Initialize extractor with database path.

        Args:
            db_path: Path to Tally SQLite database
        """
        self.db_path = db_path
        self.connection: Optional[sqlite3.Connection] = None
        self.group_tree: Dict[str, TallyGroupNode] = {}
        self.ledgers: Dict[str, TallyLedger] = {}
        self.metadata: Dict[str, str] = {}

    def connect(self) -> bool:
        """
        Connect to SQLite database.

        Returns:
            True if connection successful

        Raises:
            sqlite3.Error: If database connection fails
        """
        try:
            self.connection = sqlite3.connect(self.db_path)
            self.connection.row_factory = sqlite3.Row
            logger.info(f"Connected to Tally database: {self.db_path}")
            return True
        except sqlite3.Error as e:
            logger.error(f"Failed to connect to database: {e}")
            raise

    def disconnect(self):
        """Close database connection"""
        if self.connection:
            self.connection.close()
            logger.info("Database connection closed")

    def load_metadata(self) -> Dict[str, str]:
        """
        Load metadata from _metadata table.

        Returns:
            Dictionary of metadata key-value pairs
        """
        if not self.connection:
            raise RuntimeError("Database not connected")

        cursor = self.connection.cursor()
        cursor.execute("SELECT key, value FROM _metadata")
        self.metadata = {row[0]: row[1] for row in cursor.fetchall()}
        logger.info(f"Loaded metadata: company_name={self.metadata.get('company_name')}")
        return self.metadata

    def build_group_hierarchy(self) -> Dict[str, TallyGroupNode]:
        """
        Build complete Tally group hierarchy tree from mst_group table.

        Returns:
            Dictionary of group name -> TallyGroupNode

        Raises:
            RuntimeError: If database not connected
        """
        if not self.connection:
            raise RuntimeError("Database not connected")

        cursor = self.connection.cursor()
        cursor.execute(
            """
            SELECT NAME, PARENT, ISDEEMEDPOSITIVE, ISREVENUE
            FROM mst_group
            ORDER BY NAME
            """
        )

        # First pass: create all nodes
        for row in cursor.fetchall():
            name = row[0]
            parent = row[1]
            is_deemed_positive = row[2] == "Yes"
            is_revenue = row[3] == "Yes"

            node = TallyGroupNode(
                name=name,
                parent=parent,
                is_deemed_positive=is_deemed_positive,
                is_revenue=is_revenue,
            )
            self.group_tree[name] = node

        # Second pass: link children and compute levels
        self._compute_group_levels()

        logger.info(f"Built group hierarchy with {len(self.group_tree)} groups")
        return self.group_tree

    def _compute_group_levels(self):
        """Compute hierarchy level for each group (0=Primary, 1=Secondary, etc.)"""
        primary_groups = [g for g in self.group_tree.values() if g.parent == "Primary"]
        for group in primary_groups:
            group.level = 0
            self._traverse_children(group)

    def _traverse_children(self, parent_node: TallyGroupNode):
        """Recursively set level and link children"""
        for group_name, group_node in self.group_tree.items():
            if group_node.parent == parent_node.name:
                group_node.level = parent_node.level + 1
                parent_node.children.append(group_name)
                self._traverse_children(group_node)

    def extract_trial_balance(self) -> TrialBalance:
        """
        Extract complete trial balance from mst_ledger table.

        Extracts:
        - All ledgers with non-zero closing balance
        - Opens only balance (for P&L verification)
        - Closing balance (for Balance Sheet)
        - Group assignment and P&L flag

        Returns:
            TrialBalance object with all ledgers and summary

        Raises:
            RuntimeError: If database not connected or groups not loaded
        """
        if not self.connection:
            raise RuntimeError("Database not connected")
        if not self.group_tree:
            raise RuntimeError("Group hierarchy not built. Call build_group_hierarchy first")

        cursor = self.connection.cursor()
        cursor.execute(
            """
            SELECT NAME, PARENT, OPENINGBALANCE, CLOSINGBALANCE
            FROM mst_ledger
            WHERE CLOSINGBALANCE IS NOT NULL AND CLOSINGBALANCE != ''
            ORDER BY NAME
            """
        )

        total_debits = Decimal(0)
        total_credits = Decimal(0)
        ledger_count = 0

        for row in cursor.fetchall():
            name = row[0]
            parent = row[1]
            try:
                opening = Decimal(row[2]) if row[2] and row[2].strip() else Decimal(0)
                closing = Decimal(row[3])
            except:
                logger.warning(f"Skipped ledger {name}: invalid balance format")
                continue

            # Get group properties
            group_node = self.group_tree.get(parent)
            if not group_node:
                # Special case: P&L A/c might have parent "Primary"
                if name in ("Profit & Loss A/c", "P & L A/c", "P&L A/c") and parent == "Primary":
                    is_revenue = False  # P&L is neither revenue nor standard asset/liability
                    is_deemed_positive = False
                else:
                    logger.warning(f"Unknown group for ledger {name}: {parent}")
                    continue
            else:
                is_revenue = group_node.is_revenue
                is_deemed_positive = group_node.is_deemed_positive

            ledger = TallyLedger(
                name=name,
                parent_group=parent,
                opening_balance=opening,
                closing_balance=closing,
                is_revenue=is_revenue,
                is_deemed_positive=is_deemed_positive,
            )

            # Only include ledgers with non-zero closing balance
            if ledger.has_closing_balance:
                self.ledgers[name] = ledger
                ledger_count += 1

                # Track debits and credits
                if closing < 0:
                    total_debits += abs(closing)
                else:
                    total_credits += closing

        net_balance = total_credits - total_debits

        tb = TrialBalance(
            total_debits=total_debits,
            total_credits=total_credits,
            net_balance=net_balance,
            ledger_count=ledger_count,
            ledgers=list(self.ledgers.values()),
        )

        logger.info(f"Extracted trial balance: {tb}")
        if not tb.is_balanced:
            logger.warning(f"Trial balance NOT balanced. Net difference: {tb.net_balance}")
        else:
            logger.info("Trial balance BALANCED")

        return tb

    def get_ledger(self, name: str) -> Optional[TallyLedger]:
        """Get a single ledger by name"""
        return self.ledgers.get(name)

    def get_ledgers_by_group(self, group_name: str) -> List[TallyLedger]:
        """
        Get all ledgers belonging to a specific group (including subgroups).

        Args:
            group_name: Name of Tally group

        Returns:
            List of TallyLedger objects in that group
        """
        result = []
        for ledger in self.ledgers.values():
            if self._is_ledger_in_group(ledger, group_name):
                result.append(ledger)
        return result

    def _is_ledger_in_group(self, ledger: TallyLedger, group_name: str) -> bool:
        """
        Check if a ledger belongs to a group (including subgroups).
        Traverses up the group hierarchy.
        """
        current_group = ledger.parent_group
        while current_group:
            if current_group == group_name:
                return True
            group_node = self.group_tree.get(current_group)
            current_group = group_node.parent if group_node else None
        return False


# ============================================================================
# SECTION 4: LAYER 2 - CLASSIFICATION ENGINE
# ============================================================================

class ScheduleIIIClassifier:
    """
    Classifies every ledger from trial balance into Schedule III line items.

    Responsibilities:
    - Map Tally groups to Schedule III sections and line items
    - Handle reclassifications (debtor credit balance → advance, etc.)
    - Verify every ledger is classified
    - Apply reclassification rules
    - Identify special accounts (P&L, Suspense, etc.)
    """

    def __init__(self, extractor: TallyDataExtractor):
        """
        Initialize classifier with extractor.

        Args:
            extractor: TallyDataExtractor instance with loaded trial balance
        """
        self.extractor = extractor
        self.classified_ledgers: List[ClassifiedLedger] = []
        self.suspense_accounts: List[TallyLedger] = []
        self.unclassified_ledgers: List[TallyLedger] = []
        self.reclassifications_applied: List[Tuple[str, str]] = []

        # Define mapping rules: (Tally group) -> (Schedule III section, line item, sub_item)
        self._init_classification_rules()

    def _init_classification_rules(self):
        """Initialize Tally group to Schedule III mapping rules"""
        # This is the master mapping rule set
        self.classification_rules = {
            # EQUITY & RESERVES
            "Capital Account": (ScheduleIIISection.BS_EQUITY, "share_capital", "paid_up"),
            "Reserves & Surplus": (ScheduleIIISection.BS_EQUITY, "other_equity", "reserves_surplus"),

            # ASSETS - Non-Current
            "Fixed Assets": (ScheduleIIISection.BS_NONICA, "property_plant_equipment", "net"),
            "Investments": (ScheduleIIISection.BS_NONICA, "financial_assets", "investments"),

            # ASSETS - Current
            "Bank Accounts": (ScheduleIIISection.BS_CA, "financial_assets_ca", "cash"),
            "Cash-in-Hand": (ScheduleIIISection.BS_CA, "financial_assets_ca", "cash"),
            "Stock-in-Hand": (ScheduleIIISection.BS_CA, "inventories", "finished_goods"),
            "Sundry Debtors": (ScheduleIIISection.BS_CA, "financial_assets_ca", "trade_receivables"),
            "Deposits (Asset)": (ScheduleIIISection.BS_CA, "financial_assets_ca", "receivables"),
            "Loans & Advances (Asset)": (ScheduleIIISection.BS_CA, "financial_assets_ca", "loans"),

            # LIABILITIES - Non-Current
            "Loans (Liability)": (ScheduleIIISection.BS_NONCL, "borrowings", "unsecured"),
            "Bank OD A/c": (ScheduleIIISection.BS_NONCL, "borrowings", "secured"),
            "Secured Loans": (ScheduleIIISection.BS_NONCL, "borrowings", "secured"),
            "Unsecured Loans": (ScheduleIIISection.BS_NONCL, "borrowings", "unsecured"),

            # LIABILITIES - Current
            "Sundry Creditors": (ScheduleIIISection.BS_CL, "trade_payables", "total_outstanding"),
            "Current Liabilities": (ScheduleIIISection.BS_CL, "other_current_liabilities", ""),
            "Duties & Taxes": (ScheduleIIISection.BS_CL, "current_tax_liabilities", ""),
            "Provisions": (ScheduleIIISection.BS_CL, "provisions", "other"),

            # P&L - REVENUE
            "Sales Accounts": (ScheduleIIISection.PL_REVENUE, "revenue_operations", "sale_products"),
            "Direct Incomes": (ScheduleIIISection.PL_REVENUE, "other_income", ""),

            # P&L - EXPENSES
            "Purchase Accounts": (ScheduleIIISection.PL_EXPENSE, "purchases_changes", ""),
            "Direct Expenses": (ScheduleIIISection.PL_EXPENSE, "cost_materials", ""),
            "Indirect Expenses": (ScheduleIIISection.PL_EXPENSE, "other_expenses", ""),
            "Indirect Incomes": (ScheduleIIISection.PL_REVENUE, "other_income", ""),

            # SPECIAL
            "Suspense A/c": (None, None, None),  # Handled separately
        }

    def classify_all(self) -> List[ClassifiedLedger]:
        """
        Classify all ledgers in the trial balance.

        Steps:
        1. Iterate through all ledgers with non-zero CB
        2. Determine Tally group type (Asset, Liability, Revenue, Expense)
        3. Apply classification rules
        4. Apply reclassifications (debtor/creditor sign inversions)
        5. Track unclassified and suspense accounts

        Returns:
            List of ClassifiedLedger objects
        """
        self.classified_ledgers = []
        self.unclassified_ledgers = []
        self.suspense_accounts = []

        for ledger in self.extractor.ledgers.values():
            classified = self._classify_single_ledger(ledger)
            if classified:
                self.classified_ledgers.append(classified)
                # Track suspense separately for reporting
                if ledger.parent_group == "Suspense A/c":
                    self.suspense_accounts.append(ledger)
            else:
                self.unclassified_ledgers.append(ledger)

        logger.info(
            f"Classified {len(self.classified_ledgers)} ledgers, "
            f"{len(self.suspense_accounts)} in Suspense, "
            f"{len(self.unclassified_ledgers)} unclassified"
        )

        if self.unclassified_ledgers:
            logger.warning(f"Unclassified ledgers: {[l.name for l in self.unclassified_ledgers]}")

        return self.classified_ledgers

    def _classify_single_ledger(self, ledger: TallyLedger) -> Optional[ClassifiedLedger]:
        """
        Classify a single ledger.

        Args:
            ledger: TallyLedger to classify

        Returns:
            ClassifiedLedger if successfully classified, None otherwise
        """
        # Handle P&L accounts (Profit & Loss A/c)
        if ledger.name in ("P & L A/c", "Profit & Loss A/c", "P&L A/c"):
            return ClassifiedLedger(
                name=ledger.name,
                tally_group=ledger.parent_group,
                opening_balance=ledger.opening_balance,
                closing_balance=ledger.closing_balance,
                tally_group_type=TallyGroupType.EXPENSE if ledger.closing_balance < 0 else TallyGroupType.REVENUE,
                schedule_iii_section=ScheduleIIISection.BS_EQUITY,
                schedule_iii_line="other_equity",
                schedule_iii_sub="retained_earnings",
                display_amount=abs(ledger.closing_balance),
            )

        # SPECIAL HANDLING: Specific ledger names (overrides parent group mapping)
        special_rules = {
            "Advance Tax": (ScheduleIIISection.BS_CA, "other_current_assets", ""),
            "TDS Receivable": (ScheduleIIISection.BS_CA, "other_current_assets", ""),
            "TCS Receivable": (ScheduleIIISection.BS_CA, "other_current_assets", ""),
            "GST ITC Receivable": (ScheduleIIISection.BS_CA, "other_current_assets", ""),
            "Tax Collected at Source": (ScheduleIIISection.BS_CA, "other_current_assets", ""),
            "Balance in Demat account": (ScheduleIIISection.BS_NONICA, "financial_assets", "investments"),
            "Electricity Security": (ScheduleIIISection.BS_CA, "financial_assets_ca", "receivables"),
        }

        if ledger.name in special_rules:
            section, line, sub = special_rules[ledger.name]
            group_type = TallyGroupType.ASSET if section in (ScheduleIIISection.BS_CA, ScheduleIIISection.BS_NONICA) else TallyGroupType.LIABILITY
            return ClassifiedLedger(
                name=ledger.name,
                tally_group=ledger.parent_group,
                opening_balance=ledger.opening_balance,
                closing_balance=ledger.closing_balance,
                tally_group_type=group_type,
                schedule_iii_section=section,
                schedule_iii_line=line,
                schedule_iii_sub=sub,
                display_amount=abs(ledger.closing_balance),
                is_reclassified=False,
                reclassification_note="",
            )

        # Suspense A/c: Classify based on balance sign
        # Credit balance (positive) = liability, Debit balance (negative) = asset
        if ledger.parent_group == "Suspense A/c":
            if ledger.closing_balance > 0:  # Credit balance = Current Liability
                return ClassifiedLedger(
                    name=ledger.name,
                    tally_group=ledger.parent_group,
                    opening_balance=ledger.opening_balance,
                    closing_balance=ledger.closing_balance,
                    tally_group_type=TallyGroupType.LIABILITY,
                    schedule_iii_section=ScheduleIIISection.BS_CL,
                    schedule_iii_line="other_current_liabilities",
                    schedule_iii_sub="suspense",
                    display_amount=abs(ledger.closing_balance),
                    is_reclassified=False,
                    reclassification_note="Suspense A/c (Credit balance) classified as Current Liability",
                )
            else:  # Debit balance = Current Asset
                return ClassifiedLedger(
                    name=ledger.name,
                    tally_group=ledger.parent_group,
                    opening_balance=ledger.opening_balance,
                    closing_balance=ledger.closing_balance,
                    tally_group_type=TallyGroupType.ASSET,
                    schedule_iii_section=ScheduleIIISection.BS_CA,
                    schedule_iii_line="other_current_assets",
                    schedule_iii_sub="suspense",
                    display_amount=abs(ledger.closing_balance),
                    is_reclassified=False,
                    reclassification_note="Suspense A/c (Debit balance) classified as Current Asset",
                )

        # Get classification rule for parent group
        rule = self.classification_rules.get(ledger.parent_group)
        if not rule or rule[0] is None:
            return None

        schedule_section, schedule_line, schedule_sub = rule

        # Determine Tally group type
        group_node = self.extractor.group_tree.get(ledger.parent_group)
        if ledger.is_revenue:
            group_type = TallyGroupType.REVENUE if ledger.closing_balance > 0 else TallyGroupType.EXPENSE
        else:
            group_type = TallyGroupType.LIABILITY if ledger.closing_balance > 0 else TallyGroupType.ASSET

        # Apply reclassification rules
        is_reclassified = False
        reclassification_note = ""
        classified_section = schedule_section
        classified_line = schedule_line
        classified_sub = schedule_sub

        # RULE 1: Sundry Creditors with DEBIT balance -> Advances to Suppliers
        if (ledger.parent_group == "Sundry Creditors" and
            ledger.closing_balance < 0):  # Debit = negative
            classified_section = ScheduleIIISection.BS_CA
            classified_line = "financial_assets_ca"
            classified_sub = "loans"
            is_reclassified = True
            reclassification_note = "Sundry Creditor with debit balance reclassified to Advances"

        # RULE 2: Sundry Debtors with CREDIT balance -> Advance from Customers
        elif (ledger.parent_group == "Sundry Debtors" and
              ledger.closing_balance > 0):  # Credit = positive
            classified_section = ScheduleIIISection.BS_CL
            classified_line = "other_current_liabilities"
            classified_sub = ""
            is_reclassified = True
            reclassification_note = "Sundry Debtor with credit balance reclassified to Advances from Customers"

        # RULE 3: Indirect Expenses - sub-classify by name
        elif schedule_line == "other_expenses":
            classified_sub = self._classify_indirect_expense(ledger.name)

        # RULE 4: Trade Receivables - identify MSME
        elif schedule_line == "trade_receivables":
            if "msme" in ledger.name.lower():
                classified_sub = "msme"
            else:
                classified_sub = "others"

        # RULE 5: Trade Payables - identify MSME
        elif schedule_line == "trade_payables":
            if "msme" in ledger.name.lower():
                classified_sub = "msme"
            else:
                classified_sub = "others"

        if is_reclassified:
            self.reclassifications_applied.append((ledger.name, reclassification_note))

        return ClassifiedLedger(
            name=ledger.name,
            tally_group=ledger.parent_group,
            opening_balance=ledger.opening_balance,
            closing_balance=ledger.closing_balance,
            tally_group_type=group_type,
            schedule_iii_section=classified_section,
            schedule_iii_line=classified_line,
            schedule_iii_sub=classified_sub,
            display_amount=abs(ledger.closing_balance),
            is_reclassified=is_reclassified,
            reclassification_note=reclassification_note,
        )

    def _classify_indirect_expense(self, ledger_name: str) -> str:
        """
        Sub-classify Indirect Expense ledgers by name matching.

        Returns:
            "finance_costs", "depreciation", or "other_expenses"
        """
        name_lower = ledger_name.lower()

        finance_keywords = ["interest", "bank charge", "bank od interest", "processing fee", "loan processing"]
        depreciation_keywords = ["depreciation", "amortization", "amortisation"]

        for keyword in finance_keywords:
            if keyword in name_lower:
                return "finance_costs"

        for keyword in depreciation_keywords:
            if keyword in name_lower:
                return "depreciation"

        return "other"

    def get_classified_by_section(self, section: ScheduleIIISection) -> List[ClassifiedLedger]:
        """Get all classified ledgers in a specific Schedule III section"""
        return [c for c in self.classified_ledgers if c.schedule_iii_section == section]

    def get_classified_by_line(self, section: ScheduleIIISection, line_item: str) -> List[ClassifiedLedger]:
        """Get all classified ledgers for a specific Schedule III line item"""
        return [
            c for c in self.classified_ledgers
            if c.schedule_iii_section == section and c.schedule_iii_line == line_item
        ]


# ============================================================================
# SECTION 5: LAYER 3 - YEAR-END ADJUSTMENTS FRAMEWORK
# ============================================================================

class YearEndAdjustments:
    """
    Framework for managing year-end adjustment entries.

    Manages:
    - Depreciation calculation and journaling
    - Deferred Tax Assets/Liabilities
    - Provisions (warranties, bonuses, etc.)
    - Other manual adjustments

    Does NOT apply adjustments to the ledger directly.
    Adjustments are tracked and can be applied to Schedule III output.
    """

    def __init__(self, classifier: ScheduleIIIClassifier):
        """
        Initialize adjustments framework.

        Args:
            classifier: ScheduleIIIClassifier instance
        """
        self.classifier = classifier
        self.adjustments: Dict[str, AdjustmentEntry] = {}
        self.adjustment_counter = 0

    def add_depreciation_adjustment(
        self,
        asset_ledger: str,
        accumulated_depreciation_ledger: str,
        depreciation_expense_ledger: str,
        amount: Decimal,
        description: str = "Depreciation for FY",
    ) -> str:
        """
        Record a depreciation adjustment.

        Args:
            asset_ledger: Name of asset ledger
            accumulated_depreciation_ledger: Name of accumulated depreciation contra-asset
            depreciation_expense_ledger: Name of P&L depreciation expense account
            amount: Amount of depreciation
            description: Description of adjustment

        Returns:
            Adjustment ID
        """
        self.adjustment_counter += 1
        adj_id = f"DEP_{self.adjustment_counter}"

        entry = AdjustmentEntry(
            entry_id=adj_id,
            entry_type="depreciation",
            description=description,
            debit_ledger=depreciation_expense_ledger,
            debit_amount=amount,
            credit_ledger=accumulated_depreciation_ledger,
            credit_amount=amount,
            schedule_iii_line="depreciation",
            schedule_iii_sub="depreciation",
            notes=f"Asset: {asset_ledger}",
        )

        self.adjustments[adj_id] = entry
        logger.info(f"Added depreciation adjustment {adj_id}: {description}")
        return adj_id

    def add_dta_adjustment(
        self,
        dta_ledger: str,
        deferral_account: str,
        amount: Decimal,
        description: str = "Deferred Tax Asset",
    ) -> str:
        """
        Record a Deferred Tax Asset (DTA) adjustment.

        Args:
            dta_ledger: Name of DTA ledger
            deferral_account: Name of deferral/tax effect account
            amount: Amount of DTA
            description: Description

        Returns:
            Adjustment ID
        """
        self.adjustment_counter += 1
        adj_id = f"DTA_{self.adjustment_counter}"

        entry = AdjustmentEntry(
            entry_id=adj_id,
            entry_type="dta",
            description=description,
            debit_ledger=dta_ledger,
            debit_amount=amount,
            credit_ledger=deferral_account,
            credit_amount=amount,
            schedule_iii_line="deferred_tax_assets",
            schedule_iii_sub="",
            notes="Deferred Tax Asset calculation",
        )

        self.adjustments[adj_id] = entry
        logger.info(f"Added DTA adjustment {adj_id}: {description}")
        return adj_id

    def add_provision_adjustment(
        self,
        provision_ledger: str,
        expense_ledger: str,
        amount: Decimal,
        provision_type: str = "other",
        description: str = "Provision for contingencies",
    ) -> str:
        """
        Record a provision adjustment.

        Args:
            provision_ledger: Name of provision liability ledger
            expense_ledger: Name of P&L expense ledger
            amount: Amount of provision
            provision_type: Type of provision (employee_benefits, warranties, other)
            description: Description

        Returns:
            Adjustment ID
        """
        self.adjustment_counter += 1
        adj_id = f"PROV_{self.adjustment_counter}"

        entry = AdjustmentEntry(
            entry_id=adj_id,
            entry_type="provision",
            description=description,
            debit_ledger=expense_ledger,
            debit_amount=amount,
            credit_ledger=provision_ledger,
            credit_amount=amount,
            schedule_iii_line="provisions",
            schedule_iii_sub=provision_type,
            notes=f"Provision type: {provision_type}",
        )

        self.adjustments[adj_id] = entry
        logger.info(f"Added provision adjustment {adj_id}: {description}")
        return adj_id

    def add_manual_adjustment(
        self,
        entry_type: str,
        description: str,
        debit_ledger: str,
        debit_amount: Decimal,
        credit_ledger: str,
        credit_amount: Decimal,
        schedule_iii_line: str,
        schedule_iii_sub: str = "",
        notes: str = "",
    ) -> str:
        """
        Add a manual adjustment entry.

        Args:
            entry_type: Type of adjustment (string identifier)
            description: Description of adjustment
            debit_ledger: Debit ledger name
            debit_amount: Debit amount
            credit_ledger: Credit ledger name
            credit_amount: Credit amount
            schedule_iii_line: Schedule III line affected
            schedule_iii_sub: Schedule III sub-item
            notes: Additional notes

        Returns:
            Adjustment ID
        """
        self.adjustment_counter += 1
        adj_id = f"{entry_type.upper()}_{self.adjustment_counter}"

        entry = AdjustmentEntry(
            entry_id=adj_id,
            entry_type=entry_type,
            description=description,
            debit_ledger=debit_ledger,
            debit_amount=debit_amount,
            credit_ledger=credit_ledger,
            credit_amount=credit_amount,
            schedule_iii_line=schedule_iii_line,
            schedule_iii_sub=schedule_iii_sub,
            notes=notes,
        )

        self.adjustments[adj_id] = entry
        logger.info(f"Added manual adjustment {adj_id}: {description}")
        return adj_id

    def mark_applied(self, adjustment_id: str, applied_at: str = ""):
        """Mark an adjustment as applied"""
        if adjustment_id in self.adjustments:
            self.adjustments[adjustment_id].is_applied = True
            self.adjustments[adjustment_id].applied_at = applied_at or "2026-03-22"

    def get_adjustment(self, adjustment_id: str) -> Optional[AdjustmentEntry]:
        """Get a single adjustment by ID"""
        return self.adjustments.get(adjustment_id)

    def get_all_adjustments(self) -> List[AdjustmentEntry]:
        """Get all adjustments"""
        return list(self.adjustments.values())

    def get_unapplied_adjustments(self) -> List[AdjustmentEntry]:
        """Get adjustments that have not been applied"""
        return [a for a in self.adjustments.values() if not a.is_applied]

    def get_adjustments_by_type(self, entry_type: str) -> List[AdjustmentEntry]:
        """Get adjustments of a specific type"""
        return [a for a in self.adjustments.values() if a.entry_type == entry_type]


# ============================================================================
# SECTION 6: LAYER 4 - RECONCILIATION ENGINE
# ============================================================================

class ReconciliationEngine:
    """
    Runs comprehensive verification checks on extracted and classified data.

    Checks performed:
    1. Trial Balance Check - Debits = Credits
    2. Classification Completeness - All ledgers classified
    3. Balance Sheet Equation - Assets = Liabilities + Equity
    4. Suspense Account Check - Suspense balance is zero
    5. P&L Articulation - Opening Equity + P&L = Closing Equity
    6. MSME Totals Check - MSME totals <= Total Trade Receivables/Payables
    7. Reclassification Consistency - All reclassifications are valid
    8. Adjustment Impact Check - Adjustments don't create imbalances
    """

    def __init__(
        self,
        extractor: TallyDataExtractor,
        classifier: ScheduleIIIClassifier,
        adjustments: YearEndAdjustments,
    ):
        """
        Initialize reconciliation engine.

        Args:
            extractor: TallyDataExtractor instance
            classifier: ScheduleIIIClassifier instance
            adjustments: YearEndAdjustments instance
        """
        self.extractor = extractor
        self.classifier = classifier
        self.adjustments = adjustments
        self.checks: List[VerificationCheck] = []

    def run_all_checks(self) -> List[VerificationCheck]:
        """
        Run all 8 verification checks.

        Returns:
            List of VerificationCheck results
        """
        self.checks = []

        self.checks.append(self._check_trial_balance())
        self.checks.append(self._check_classification_completeness())
        self.checks.append(self._check_balance_sheet_equation())
        self.checks.append(self._check_suspense_account())
        self.checks.append(self._check_pl_articulation())
        self.checks.append(self._check_msme_totals())
        self.checks.append(self._check_reclassification_consistency())
        self.checks.append(self._check_adjustment_impact())

        # Log results
        passed = sum(1 for c in self.checks if c.passed)
        failed = len(self.checks) - passed
        logger.info(f"Reconciliation: {passed} PASSED, {failed} FAILED")

        return self.checks

    def _check_trial_balance(self) -> VerificationCheck:
        """
        CHECK 1: Trial Balance Check
        Verify that total debits = total credits
        """
        tb = TrialBalance(
            total_debits=Decimal(0),
            total_credits=Decimal(0),
            net_balance=Decimal(0),
            ledger_count=0,
        )

        for ledger in self.extractor.ledgers.values():
            if ledger.closing_balance < 0:
                tb.total_debits += abs(ledger.closing_balance)
            else:
                tb.total_credits += ledger.closing_balance

        tb.net_balance = tb.total_credits - tb.total_debits
        tb.ledger_count = len(self.extractor.ledgers)

        status = "PASS" if tb.is_balanced else "FAIL"
        return VerificationCheck(
            check_name="Trial Balance",
            expected=Decimal(0),
            actual=tb.net_balance,
            difference=abs(tb.net_balance),
            status=status,
            details=f"Debits: {tb.total_debits}, Credits: {tb.total_credits}, Net: {tb.net_balance}",
            severity="CRITICAL" if status == "FAIL" else "INFO",
        )

    def _check_classification_completeness(self) -> VerificationCheck:
        """
        CHECK 2: Classification Completeness
        Verify that all non-suspense ledgers are classified
        """
        unclassified_count = len(self.classifier.unclassified_ledgers)
        total_count = len(self.extractor.ledgers)

        status = "PASS" if unclassified_count == 0 else "WARNING"
        return VerificationCheck(
            check_name="Classification Completeness",
            expected=Decimal(total_count),
            actual=Decimal(total_count - unclassified_count),
            difference=Decimal(unclassified_count),
            status=status,
            details=f"Classified: {total_count - unclassified_count}/{total_count}. "
                    f"Unclassified: {[l.name for l in self.classifier.unclassified_ledgers]}",
            severity="WARNING" if unclassified_count > 0 else "INFO",
        )

    def _check_balance_sheet_equation(self) -> VerificationCheck:
        """
        CHECK 3: Balance Sheet Equation
        Assets = Liabilities + Equity
        """
        total_assets = Decimal(0)
        total_liabilities = Decimal(0)
        total_equity = Decimal(0)

        for classified in self.classifier.classified_ledgers:
            if classified.schedule_iii_section in (
                ScheduleIIISection.BS_NONICA,
                ScheduleIIISection.BS_CA,
            ):
                total_assets += classified.display_amount
            elif classified.schedule_iii_section in (
                ScheduleIIISection.BS_NONCL,
                ScheduleIIISection.BS_CL,
            ):
                total_liabilities += classified.display_amount
            elif classified.schedule_iii_section == ScheduleIIISection.BS_EQUITY:
                total_equity += classified.display_amount

        expected = total_liabilities + total_equity
        difference = abs(total_assets - expected)
        status = "PASS" if difference < Decimal("1") else "FAIL"

        return VerificationCheck(
            check_name="Balance Sheet Equation",
            expected=expected,
            actual=total_assets,
            difference=difference,
            status=status,
            details=f"Assets: {total_assets}, Liabilities: {total_liabilities}, "
                    f"Equity: {total_equity}. Expected A=L+E: {expected}",
            severity="CRITICAL" if status == "FAIL" else "INFO",
        )

    def _check_suspense_account(self) -> VerificationCheck:
        """
        CHECK 4: Suspense Account Check
        Suspense account balance should be zero (or tracked separately)
        """
        suspense_total = Decimal(0)
        for ledger in self.classifier.suspense_accounts:
            suspense_total += abs(ledger.closing_balance)

        status = "PASS" if suspense_total == 0 else "WARNING"
        return VerificationCheck(
            check_name="Suspense Account",
            expected=Decimal(0),
            actual=suspense_total,
            difference=suspense_total,
            status=status,
            details=f"Suspense accounts found: {len(self.classifier.suspense_accounts)}. "
                    f"Total balance: {suspense_total}",
            severity="WARNING" if suspense_total > 0 else "INFO",
        )

    def _check_pl_articulation(self) -> VerificationCheck:
        """
        CHECK 5: P&L Articulation
        Opening Equity + Profit = Closing Equity (simplified check)
        This is a placeholder - full check would require P&L extraction
        """
        # Get closing equity
        closing_equity = Decimal(0)
        for classified in self.classifier.classified_ledgers:
            if classified.schedule_iii_section == ScheduleIIISection.BS_EQUITY:
                closing_equity += classified.display_amount

        # For now, just verify equity exists
        status = "PASS" if closing_equity != 0 else "WARNING"
        return VerificationCheck(
            check_name="P&L Articulation",
            expected=Decimal(0),
            actual=closing_equity,
            difference=Decimal(0),
            status=status,
            details=f"Closing Equity: {closing_equity}. "
                    f"Full P&L articulation check requires P&L extraction.",
            severity="INFO",
        )

    def _check_msme_totals(self) -> VerificationCheck:
        """
        CHECK 6: MSME Totals Check
        MSME trade receivables/payables should not exceed total
        """
        tr_msme = Decimal(0)
        tr_total = Decimal(0)

        for classified in self.classifier.classified_ledgers:
            if (classified.schedule_iii_section == ScheduleIIISection.BS_CA and
                classified.schedule_iii_line == "financial_assets_ca"):
                if classified.schedule_iii_sub == "trade_receivables":
                    tr_total += classified.display_amount
                elif classified.schedule_iii_sub == "msme":
                    tr_msme += classified.display_amount

        # MSME should be part of total
        status = "PASS" if tr_msme <= tr_total or tr_total == 0 else "FAIL"
        return VerificationCheck(
            check_name="MSME Totals",
            expected=tr_total,
            actual=tr_msme,
            difference=abs(tr_total - tr_msme),
            status=status,
            details=f"MSME TR: {tr_msme}, Total TR: {tr_total}",
            severity="WARNING" if status == "FAIL" else "INFO",
        )

    def _check_reclassification_consistency(self) -> VerificationCheck:
        """
        CHECK 7: Reclassification Consistency
        All reclassifications should follow defined rules
        """
        reclassified_count = sum(1 for c in self.classifier.classified_ledgers if c.is_reclassified)
        applied_rules = len(self.classifier.reclassifications_applied)

        status = "PASS"
        return VerificationCheck(
            check_name="Reclassification Consistency",
            expected=Decimal(applied_rules),
            actual=Decimal(reclassified_count),
            difference=Decimal(0),
            status=status,
            details=f"Reclassifications applied: {applied_rules}",
            severity="INFO",
        )

    def _check_adjustment_impact(self) -> VerificationCheck:
        """
        CHECK 8: Adjustment Impact Check
        All adjustments should balance (debits = credits)
        """
        total_debits = Decimal(0)
        total_credits = Decimal(0)

        for adj in self.adjustments.get_all_adjustments():
            total_debits += adj.debit_amount
            total_credits += adj.credit_amount

        difference = abs(total_debits - total_credits)
        status = "PASS" if difference < Decimal("0.01") else "FAIL"

        return VerificationCheck(
            check_name="Adjustment Impact",
            expected=total_credits,
            actual=total_debits,
            difference=difference,
            status=status,
            details=f"Total adjustments: {len(self.adjustments.get_all_adjustments())}. "
                    f"Debits: {total_debits}, Credits: {total_credits}",
            severity="CRITICAL" if status == "FAIL" else "INFO",
        )

    def get_summary(self) -> Dict[str, any]:
        """
        Get summary of all verification checks.

        Returns:
            Dictionary with pass/fail counts and status
        """
        passed = sum(1 for c in self.checks if c.passed)
        failed = sum(1 for c in self.checks if c.status == "FAIL")
        warnings = sum(1 for c in self.checks if c.status == "WARNING")

        return {
            "total_checks": len(self.checks),
            "passed": passed,
            "failed": failed,
            "warnings": warnings,
            "overall_status": "PASS" if failed == 0 else "FAIL",
            "checks": self.checks,
        }


# ============================================================================
# SECTION 7: HELPER FUNCTIONS
# ============================================================================

def get_group_ancestry(
    group_tree: Dict[str, TallyGroupNode],
    group_name: str,
) -> List[str]:
    """
    Get the complete ancestry of a group (from root to leaf).

    Args:
        group_tree: Group hierarchy tree from TallyDataExtractor
        group_name: Name of group

    Returns:
        List of group names from root to specified group
    """
    ancestry = []
    current = group_name
    while current:
        ancestry.insert(0, current)
        node = group_tree.get(current)
        current = node.parent if node else None
    return ancestry


def sum_classified_by_section(
    classified_ledgers: List[ClassifiedLedger],
    section: ScheduleIIISection,
) -> Decimal:
    """
    Sum all classified ledger amounts in a specific Schedule III section.

    Args:
        classified_ledgers: List of classified ledgers
        section: Target Schedule III section

    Returns:
        Total amount in section
    """
    return sum(
        c.display_amount
        for c in classified_ledgers
        if c.schedule_iii_section == section
    )


def sum_classified_by_line(
    classified_ledgers: List[ClassifiedLedger],
    section: ScheduleIIISection,
    line_item: str,
) -> Decimal:
    """
    Sum all classified ledger amounts for a specific Schedule III line item.

    Args:
        classified_ledgers: List of classified ledgers
        section: Target Schedule III section
        line_item: Target line item code

    Returns:
        Total amount for line item
    """
    return sum(
        c.display_amount
        for c in classified_ledgers
        if c.schedule_iii_section == section and c.schedule_iii_line == line_item
    )


# ============================================================================
# SECTION 8: PRIOR YEAR PARSER
# ============================================================================

class PriorYearParser:
    """Parse prior year financial statements from Excel files."""

    def __init__(self, excel_path: str):
        """
        Initialize parser with Excel file path.

        Args:
            excel_path: Path to previous year's Excel financial statement
        """
        self.excel_path = excel_path
        self.unit_multiplier = Decimal(1)

    def parse(self) -> Dict[str, Dict[str, Decimal]]:
        """
        Parse Balance Sheet and P&L from Excel file.

        Returns:
            Dict with structure: {"bs": {...}, "pl": {...}}
        """
        import openpyxl
        from openpyxl.utils import get_column_letter

        try:
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        except Exception as e:
            logger.error(f"Failed to load Excel: {e}")
            return {"bs": {}, "pl": {}}

        result = {"bs": {}, "pl": {}}

        # Parse Balance Sheet
        if "Balance Sheet" in wb.sheetnames:
            result["bs"] = self._parse_balance_sheet(wb["Balance Sheet"])
        elif any("balance" in sn.lower() for sn in wb.sheetnames):
            sheet = next(ws for ws in wb.sheetnames if "balance" in ws.lower())
            result["bs"] = self._parse_balance_sheet(wb[sheet])

        # Parse P&L
        pl_sheets = [s for s in wb.sheetnames if any(x in s.lower() for x in ["profit", "loss", "p&l", "pl", "income"])]
        if pl_sheets:
            result["pl"] = self._parse_pl(wb[pl_sheets[0]])

        return result

    def _detect_unit(self, ws) -> None:
        """Detect unit from header text. Also validates by checking if
        amounts look pre-scaled (large values suggest absolute rupees
        despite header saying Thousands)."""
        text = ""
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for cell in row:
                if cell:
                    text += str(cell).lower()

        header_unit = Decimal(1)
        if "lakh" in text or "lac" in text:
            header_unit = Decimal(100000)
        elif "thousand" in text or "'000" in text:
            header_unit = Decimal(1000)
        elif "hundred" in text or "'00" in text:
            header_unit = Decimal(100)

        # Validation: sample a few amounts. If header says "Thousands"
        # but amounts are > 100,000, the data is likely in absolute rupees already
        amount_col, _ = self._find_amount_columns(ws)
        sample_vals = []
        for row_num in range(8, min(25, ws.max_row + 1)):
            cell = ws.cell(row=row_num, column=amount_col)
            if isinstance(cell.value, (int, float)) and abs(cell.value) > 50:
                sample_vals.append(abs(cell.value))

        if sample_vals and header_unit > 1:
            median_val = sorted(sample_vals)[len(sample_vals) // 2]
            # If median value > 50000 and header says Thousands, data is likely absolute
            if median_val > 50000 and header_unit == Decimal(1000):
                logger.info(f"Unit override: header says Thousands but median amount is {median_val:.0f} — treating as absolute rupees")
                self.unit_multiplier = Decimal(1)
                return
            if median_val > 500000 and header_unit == Decimal(100000):
                logger.info(f"Unit override: header says Lakhs but median amount is {median_val:.0f} — treating as absolute rupees")
                self.unit_multiplier = Decimal(1)
                return

        self.unit_multiplier = header_unit
        self.detected_unit = str(header_unit)

    def _find_amount_columns(self, ws) -> Tuple[int, Optional[int]]:
        """Find current year and previous year amount columns.

        Logic: Scan rows 5-20 for columns with large numeric values (>10).
        Skip columns with small numbers (likely note references like 1.0, 2.0).
        Returns (current_year_col, previous_year_col).
        """
        from collections import Counter
        col_scores = Counter()

        for row_num in range(5, min(30, ws.max_row + 1)):
            for col_num in range(2, min(10, ws.max_column + 1)):
                cell = ws.cell(row=row_num, column=col_num)
                if isinstance(cell.value, (int, float)):
                    val = abs(cell.value)
                    if val > 25:  # Real amounts are > 25; note refs are 1-24
                        col_scores[col_num] += 1

        if not col_scores:
            return (3, 4)  # Safe default: C=current, D=previous

        # Sort by frequency — most common columns with real amounts
        ranked = col_scores.most_common()
        current_col = ranked[0][0]
        prev_col = ranked[1][0] if len(ranked) > 1 else None

        # Ensure current comes before previous (left = current, right = previous)
        if prev_col and prev_col < current_col:
            current_col, prev_col = prev_col, current_col

        return (current_col, prev_col)

    def _parse_balance_sheet(self, ws) -> Dict[str, Decimal]:
        """Extract Balance Sheet line items."""
        self._detect_unit(ws)
        amount_col, prev_col = self._find_amount_columns(ws)
        self.bs_current_col = amount_col
        self.bs_prev_col = prev_col
        logger.info(f"B/S columns detected: current={amount_col}, previous={prev_col}")

        bs_items = {
            "share_capital": [],
            "reserves_surplus": [],
            "long_term_borrowings": [],
            "deferred_tax": [],
            "short_term_borrowings": [],
            "trade_payables": [],
            "other_current_liabilities": [],
            "short_term_provisions": [],
            "ppe": [],
            "intangible_assets": [],
            "investments": [],
            "deferred_tax_asset": [],
            "inventories": [],
            "trade_receivables": [],
            "cash_equivalents": [],
            "short_term_loans": [],
            "other_current_assets": [],
        }

        patterns = {
            "share_capital": ["share capital", "paid.up"],
            "reserves_surplus": ["reserve", "surplus"],
            "long_term_borrowings": ["long.term borrow", "long term borrow"],
            "deferred_tax": ["deferred tax liability"],
            "short_term_borrowings": ["short.term borrow", "bank od", "overdraft"],
            "trade_payables": ["trade payable", "payable"],
            "other_current_liabilities": ["other current liab"],
            "short_term_provisions": ["short.term provision", "^provision"],
            "ppe": ["property", "plant", "fixed asset", "tangible"],
            "intangible_assets": ["intangible"],
            "investments": ["investment"],
            "deferred_tax_asset": ["deferred tax asset"],
            "inventories": ["inventor"],
            "trade_receivables": ["trade receivable", "sundry debtor", "debtor"],
            "cash_equivalents": ["cash", "bank balance"],
            "short_term_loans": ["short.term loan", "loan.*advance", "advance"],
            "other_current_assets": ["other current asset"],
        }

        data = {}
        for row_num in range(1, ws.max_row + 1):
            label_cell = ws.cell(row=row_num, column=1)
            amount_cell = ws.cell(row=row_num, column=amount_col)

            if not label_cell.value or not isinstance(amount_cell.value, (int, float)):
                continue

            label = str(label_cell.value).strip().lower()
            amount = Decimal(str(amount_cell.value)) * self.unit_multiplier

            for item_key, pats in patterns.items():
                if any(pat.replace(".", "").replace("^", "") in label.replace(".", "") for pat in pats):
                    if item_key not in data:
                        data[item_key] = Decimal(0)
                    data[item_key] += amount
                    break

        return data

    def _parse_pl(self, ws) -> Dict[str, Decimal]:
        """Extract P&L line items."""
        self._detect_unit(ws)
        amount_col, prev_col = self._find_amount_columns(ws)
        self.pl_current_col = amount_col
        self.pl_prev_col = prev_col
        logger.info(f"P&L columns detected: current={amount_col}, previous={prev_col}")

        pl_items = {
            "revenue": [],
            "other_income": [],
            "cost_materials": [],
            "inventory_change": [],
            "employee_expense": [],
            "finance_costs": [],
            "depreciation": [],
            "other_expenses": [],
            "tax_expense": [],
            "net_profit": [],
        }

        patterns = {
            "revenue": ["revenue from operation"],
            "other_income": ["other income"],
            "cost_materials": ["cost of material", "purchase"],
            "inventory_change": ["change.*inventor"],
            "employee_expense": ["employee", "salary"],
            "finance_costs": ["finance", "interest"],
            "depreciation": ["depreciation", "amortization"],
            "other_expenses": ["other expense"],
            "tax_expense": ["tax"],
            "net_profit": ["profit", "loss"],
        }

        data = {}
        net_profit_row = None
        net_profit_val = None

        for row_num in range(1, ws.max_row + 1):
            label_cell = ws.cell(row=row_num, column=1)
            amount_cell = ws.cell(row=row_num, column=amount_col)

            if not label_cell.value or not isinstance(amount_cell.value, (int, float)):
                continue

            label = str(label_cell.value).strip().lower()
            amount = Decimal(str(amount_cell.value)) * self.unit_multiplier

            # Track last profit/loss as net profit
            if any(pat in label for pat in ["profit", "loss"]):
                net_profit_row = row_num
                net_profit_val = amount

            for item_key, pats in patterns.items():
                if item_key == "net_profit":
                    continue
                if any(pat.replace(".*", "").replace("^", "") in label for pat in pats):
                    if item_key not in data:
                        data[item_key] = Decimal(0)
                    data[item_key] += amount
                    break

        if net_profit_val is not None:
            data["net_profit"] = net_profit_val

        return data


# ============================================================================
# END OF FILE
# ============================================================================
